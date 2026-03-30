#!/usr/bin/env python3
"""
Pricing Table Extractor

Reads an extracted section JSON and produces an Excel file with
pricing line items mapped to a standardised schema.

Two-pass scanning:
  Pass 1 — Document-level: extract common fields (currency, dates,
           contract ID, company) from header + all content. These
           become defaults for EVERY row.
  Pass 2 — Per-table: scan each pricing table's local context
           (parent section, surrounding paragraphs) for overrides.
           These apply to all rows from that specific table.

Then for each table, row-level extraction fills in item-specific
fields (description, price, UOM, quantity).

Usage:
    python scripts/pricing_extractor.py --input output/final/doc.json
    python scripts/pricing_extractor.py --input doc.json --output pricing.xlsx
    python scripts/pricing_extractor.py --input doc.json --no-llm
    python scripts/pricing_extractor.py --input doc.json --provider azure_openai
"""
import argparse
import json
import re
import sys
import os
from pathlib import Path
from typing import Any, Dict, List, Optional

# ---------------------------------------------------------------------------
# Target output fields
# ---------------------------------------------------------------------------

OUTPUT_FIELDS = [
    "CONTRACT_ID",
    "VALID_FROM",
    "VALID_TO",
    "COMPANY_CODE",
    "CURRENCY",
    "MATERIAL_ID",
    "ITEM_DESCRIPTION",
    "ORDER_UOM",
    "TARGET_QUANTITY",
    "MIN_RATE",
    "MAX_RATE",
    "PRICE_UNIT_UOM",
    "NET_PRICE",
    "CLAUSE_DESCRIPTION",
]

# Pricing keywords for table detection
_PRICING_KW = [
    "price", "pricing", "rate", "rates", "fee", "fees", "cost", "costs",
    "charge", "charges", "schedule of rates", "bill of quantities",
    "tariff", "lump sum", "unit rate", "remuneration", "payment",
    "$/hr", "$/day", "$/tonne", "per hour", "per day", "per unit",
    "rate card", "price list", "price schedule", "rate schedule",
    "amount", "value", "quotation",
]


def _is_header_row(row: List) -> bool:
    """
    Check if a row looks like column headers rather than data.

    A header row is predominantly text (non-numeric) cells.
    At least 60% of non-empty cells should be text.
    """
    if not row:
        return False
    text_count = 0
    numeric_count = 0
    for cell in row:
        s = str(cell).strip()
        if not s:
            continue
        if re.match(r'^[$€£]?\s*[\d,]+\.?\d*$', s):
            numeric_count += 1
        else:
            text_count += 1
    total = text_count + numeric_count
    if total == 0:
        return False
    return text_count / total >= 0.6


# ======================================================================
# Pass 1 — Document-level scan
# ======================================================================

def extract_document_defaults(document: Dict, invoke_fn=None) -> Dict[str, str]:
    """
    Scan the ENTIRE document for common fields that apply globally.

    Three sources, checked in priority order (later overrides earlier):
      1. Document header (lowest priority — often just creation date)
      2. ALL paragraphs — regex finds candidate regions, then LLM
         extracts the actual value from a wide text window
      3. ALL tables — key-value pair tables (highest priority)

    When invoke_fn is available, regex matches in sources 1-2 are
    verified by sending surrounding text to the LLM for extraction.

    Returns a dict keyed by OUTPUT_FIELDS names.
    """
    defaults: Dict[str, str] = {f: "" for f in OUTPUT_FIELDS}

    header = document.get("document_header", {})
    all_text = _collect_all_text(document)

    # --- Source 1 (lowest priority): Document header ---
    _extract_from_header(header, defaults, all_text, invoke_fn)

    # --- Source 2: Paragraph text (regex + LLM) ---
    _extract_from_text(all_text, defaults, invoke_fn)

    # --- Source 3 (highest priority): Key-value tables ---
    # all_kv = _scan_all_tables_for_kv(document)
    # _apply_kv_pairs(all_kv, defaults)

    return defaults


def _extract_from_header(header: Dict, defaults: Dict[str, str],
                        all_text: str, invoke_fn=None) -> None:
    """
    Extract CONTRACT_ID and COMPANY_CODE from the document header.

    When LLM is available, sends the entire header as-is to the LLM
    for extraction in a single call. No key/value checking — let the
    LLM figure out which fields contain the contract ID and company.

    When no LLM, falls back to key-name matching.
    """
    if invoke_fn and header:
        header_json = json.dumps(header, indent=2, default=str)
        prompt = f"""Extract the Contract ID and Company Name from this document header.

DOCUMENT HEADER:
{header_json}

RULES:
1. CONTRACT_ID: the contract number, agreement number, reference number, or document number. Return ONLY the identifier (digits code), not labels.
2. COMPANY_CODE: the company code. NOT company name.
3. If a field is not found, return empty string "".
4. Return ONLY a JSON object.

Return JSON:"""

        try:
            response = invoke_fn(prompt, 256)
            result = _parse_defaults_response(response, ["CONTRACT_ID", "COMPANY_CODE"])
            if result.get("CONTRACT_ID"):
                defaults["CONTRACT_ID"] = _clean_contract_id(result["CONTRACT_ID"])
            if result.get("COMPANY_CODE"):
                defaults["COMPANY_CODE"] = result["COMPANY_CODE"]
        except Exception as e:
            print(f"    LLM header extraction failed ({e}), falling back to key matching")
            _extract_from_header_fallback(header, defaults)
    else:
        _extract_from_header_fallback(header, defaults)


def _extract_from_header_fallback(header: Dict, defaults: Dict[str, str]) -> None:
    """Fallback: extract from header using key-name matching (no LLM)."""
    for key in ("document_number", "reference", "contract_number",
                "agreement_number", "contract_id"):
        val = _header_val(header, key)
        if val:
            defaults["CONTRACT_ID"] = _clean_contract_id(val)
            break

    for key in ("organisation", "company", "principal", "client"):
        val = _header_val(header, key)
        if val:
            defaults["COMPANY_CODE"] = val
            break


def _scan_all_tables_for_kv(document: Dict) -> List[Dict[str, str]]:
    """
    Walk ALL tables in the document and extract key-value pairs.

    A key-value table is one where:
      - It has 2 columns (or the first 2 columns are label+value)
      - Column 1 contains recognisable field labels
      - Column 2 contains the corresponding values

    Also handles wider tables where a known label appears in any cell
    and the adjacent cell contains the value.

    Returns a list of {label: value} dicts found.
    """
    kv_pairs: List[Dict[str, str]] = []
    _walk_for_tables(document, kv_pairs)
    return kv_pairs


def _walk_for_tables(data: Any, kv_pairs: List[Dict[str, str]]) -> None:
    """Recursively find all table blocks and extract key-value pairs."""
    if isinstance(data, dict):
        if data.get("type") == "table":
            _extract_kv_from_table(data, kv_pairs)
        else:
            for k, v in data.items():
                if not k.startswith("_") and k != "image":
                    _walk_for_tables(v, kv_pairs)
    elif isinstance(data, list):
        for item in data:
            _walk_for_tables(item, kv_pairs)


# Pattern-based label matching — each entry is (regex_pattern, target_field)
# Patterns are checked in order; first match wins.
# The regex is applied to the NORMALISED label (lowercased, stripped, colon removed).
_KV_LABEL_PATTERNS: List[tuple] = [
    # VALID_FROM — broad matching for any "effective date" variant
    (r'\beffective\s+date\b', "VALID_FROM"),
    (r'\bdate\s+of\s+effect\b', "VALID_FROM"),
    (r'\bdate\s+effective\b', "VALID_FROM"),
    (r'\bcommencement\s+date\b', "VALID_FROM"),
    (r'\bstart\s+date\b', "VALID_FROM"),
    (r'\bvalid\s+from\b', "VALID_FROM"),
    (r'\bcommencing\b', "VALID_FROM"),
    (r'\bdate\s+of\s+commencement\b', "VALID_FROM"),
    (r'\bagreement\s+date\b', "VALID_FROM"),
    (r'\bcontract\s+date\b', "VALID_FROM"),
    (r'\bexecution\s+date\b', "VALID_FROM"),
    (r'\bdate\s+of\s+agreement\b', "VALID_FROM"),
    (r'\bdate\s+of\s+contract\b', "VALID_FROM"),
    (r'\bdeed\s+date\b', "VALID_FROM"),
    (r'\bdated\b', "VALID_FROM"),
    (r'\bvariation\s+date\b', "VALID_FROM"),
    # VALID_TO
    (r'\bexpiry\s+date\b', "VALID_TO"),
    (r'\bexpiration\s+date\b', "VALID_TO"),
    (r'\bend\s+date\b', "VALID_TO"),
    (r'\bvalid\s+(?:to|until)\b', "VALID_TO"),
    (r'\btermination\s+date\b', "VALID_TO"),
    (r'\bexpires?\b', "VALID_TO"),
    (r'\bdate\s+of\s+expiry\b', "VALID_TO"),
    # CONTRACT_ID
    (r'\bcontract\s+(?:no\.?|number|ref|reference|id)\b', "CONTRACT_ID"),
    (r'\bagreement\s+(?:no\.?|number|ref|reference|id)\b', "CONTRACT_ID"),
    (r'\breference\s+(?:no\.?|number)\b', "CONTRACT_ID"),
    (r'\bdocument\s+(?:no\.?|number)\b', "CONTRACT_ID"),
    (r'\border\s+(?:no\.?|number)\b', "CONTRACT_ID"),
    (r'\bproject\s+(?:no\.?|number|ref)\b', "CONTRACT_ID"),
    # CURRENCY
    (r'\bcurrency\b', "CURRENCY"),
    (r'\bpayment\s+currency\b', "CURRENCY"),
    (r'\bcontract\s+currency\b', "CURRENCY"),
    # COMPANY_CODE
    (r'\bcompany(?:\s+name)?\b', "COMPANY_CODE"),
    (r'\bprincipal\b', "COMPANY_CODE"),
    (r'\bclient\b', "COMPANY_CODE"),
    (r'\borgani[sz]ation\b', "COMPANY_CODE"),
    (r'\bbuyer\b', "COMPANY_CODE"),
    (r'\bpurchaser\b', "COMPANY_CODE"),
    (r'\bvendor\b', "COMPANY_CODE"),
    (r'\bcontractor\b', "COMPANY_CODE"),
    (r'\bsupplier\b', "COMPANY_CODE"),
    (r'\bparty\b', "COMPANY_CODE"),
]


def _match_label_to_field(label: str) -> Optional[str]:
    """
    Match a table cell label to an output field using pattern matching.

    The label is normalised (lowercased, stripped, trailing colon removed,
    parenthetical hints removed) before matching.
    """
    if not label or not label.strip():
        return None

    # Normalise: lowercase, strip whitespace and colons, remove parentheticals
    normalised = label.lower().strip().rstrip(":").strip()
    normalised = re.sub(r'\([^)]*\)', '', normalised).strip()

    for pattern, field in _KV_LABEL_PATTERNS:
        if re.search(pattern, normalised):
            return field
    return None


def _extract_kv_from_table(table: Dict, kv_pairs: List[Dict[str, str]]) -> None:
    """
    Extract key-value pairs from a single table.

    Strategy 1: 2-column table — row[0] is label, row[1] is value.
    Strategy 2: Wider table — scan all adjacent cell pairs for labels.
    Strategy 3: Header row — if a header matches a known label and
                there are few data rows, the cell below is the value.
    Strategy 4: Single-cell — label and value in one cell separated
                by colon (e.g. "Effective Date: 1 July 2025").
    """
    headers = table.get("headers", [])
    rows = table.get("rows", [])
    if not rows:
        return

    found_fields: set = set()  # avoid duplicate matches

    # Strategy 1 & 2: Scan rows for label-value pairs
    for row in rows:
        if not isinstance(row, list) or len(row) < 1:
            continue

        # Check adjacent pairs (covers both 2-col and wider tables)
        for ci in range(len(row)):
            cell_text = str(row[ci]).strip()

            # Strategy 4: single cell with "Label: Value"
            if ":" in cell_text and ci == 0:
                parts = cell_text.split(":", 1)
                label_part = parts[0].strip()
                value_part = parts[1].strip() if len(parts) > 1 else ""
                field = _match_label_to_field(label_part)
                if field and value_part and field not in found_fields:
                    kv_pairs.append({"label": label_part, "value": value_part, "field": field})
                    found_fields.add(field)

            # Strategy 1 & 2: this cell is a label, next cell is value
            if ci + 1 < len(row):
                label = cell_text.rstrip(":").strip()
                value = str(row[ci + 1]).strip()
                field = _match_label_to_field(label)
                if field and value and field not in found_fields:
                    kv_pairs.append({"label": label, "value": value, "field": field})
                    found_fields.add(field)

    # Strategy 3: Headers as labels (summary tables)
    if headers and len(rows) <= 3:
        for row in rows:
            if not isinstance(row, list):
                continue
            for ci, header in enumerate(headers):
                if ci < len(row):
                    label = str(header).strip().rstrip(":").strip()
                    value = str(row[ci]).strip()
                    field = _match_label_to_field(label)
                    if field and value and field not in found_fields:
                        kv_pairs.append({"label": label, "value": value, "field": field})
                        found_fields.add(field)


def _clean_contract_id(value: str) -> str:
    """
    Extract just the contract ID/number from a value that may contain
    surrounding label text.

    Examples:
        "MSA-2025-001"                    -> "MSA-2025-001"
        "Agreement No. CN-4500012345"     -> "CN-4500012345"
        "Contract Number: 4500067890"     -> "4500067890"
        "Ref: PRJ-2025/789"              -> "PRJ-2025/789"
        "DOC 12345"                       -> "12345"
        "No. 4500012345"                  -> "4500012345"
        "See contract 4500012345 for..."  -> "4500012345"
    """
    if not value or not value.strip():
        return ""

    v = value.strip()

    # If the value is already just an ID (alphanumeric with dashes/slashes), return it
    if re.match(r'^[0-9]$', v, re.IGNORECASE):
        return v

    # Strip common label prefixes that precede the actual ID
    # Match the full label phrase, then separator, then capture the rest
    m_prefix = re.match(
        r'(?:contract|agreement|document|order|project|purchase\s+order|po|'
        r'ref(?:erence)?|deed)\s*'
        r'(?:no\.?|number|ref(?:erence)?|id)?\s*[:\s.\-]*\s*'
        r'(.+)$',
        v, flags=re.IGNORECASE,
    )
    if m_prefix:
        v = m_prefix.group(1).strip()

    # If we now have a clean ID, return it
    if v and re.match(r'^[0-9]$', v, re.IGNORECASE):
        return v

    # Otherwise, find the first alphanumeric+dash/slash token that contains digits
    m = re.search(r'(\d*)', v, re.IGNORECASE)
    if m:
        return m.group(1)

    # Last resort: return the cleaned value
    return v.strip()


def _apply_kv_pairs(kv_pairs: List[Dict[str, str]], defaults: Dict[str, str]) -> None:
    """
    Apply extracted key-value pairs to defaults.
    KV tables are highest priority so they override existing values.
    """
    for kv in kv_pairs:
        field = kv.get("field")
        value = kv.get("value", "")
        if field and value and field in defaults:
            if defaults[field] != '':
                continue
            # Clean contract IDs to extract just the identifier
            if field == "CONTRACT_ID":
                value = _clean_contract_id(value)
            defaults[field] = value


def _extract_from_text(all_text: str, defaults: Dict[str, str],
                       invoke_fn=None) -> None:
    """
    Extract fields from paragraph text.

    Strategy: use regex to find candidate regions in the text, then
    send a wide window around each match to the LLM for reliable
    extraction. Falls back to regex-only when no LLM is available.
    """
    # Define regex patterns to find candidate regions for each field
    _FIELD_PATTERNS = {
        "CONTRACT_ID": [
            r'[Cc]ontract\s+(?:[Nn]o\.?|[Nn]umber|[Rr]ef(?:erence)?)[:\s]+([A-Z0-9][\w\-/]+)',
            r'[Aa]greement\s+(?:[Nn]o\.?|[Nn]umber|[Rr]ef)[:\s]+([A-Z0-9][\w\-/]+)',
            r'[Cc]ontract\s+(?:[Nn]o\.?|[Nn]umber|[Rr]ef)',
            r'[Aa]greement\s+(?:[Nn]o\.?|[Nn]umber|[Rr]ef)',
        ],
        "COMPANY_CODE": [
            r'between\s+([A-Z][A-Za-z\s&]+?(?:Ltd|Pty|Inc|Corp|Limited|LLC)\.?)',
            r'(?:principal|client|company|buyer|purchaser)[:\s]+',
        ],
        "CURRENCY": [
            r'\b(AUD|USD|EUR|GBP|NZD|CAD|SGD)\b',
            r'[Aa]ustralian\s+[Dd]ollars?',
            r'US\s+[Dd]ollars?',
        ],
        "VALID_FROM": [
            r'[Ee]ffective\s+(?:[Dd]ate|[Ff]rom)',
            r'[Cc]ommencement\s+[Dd]ate',
            r'[Vv]alid\s+[Ff]rom',
            r'[Ss]tart\s+[Dd]ate',
            r'[Dd]ate\s+[Oo]f\s+[Ss]igning',
            r'[Ee]xecution\s+[Dd]ate',
        ],
        "VALID_TO": [
            r'[Ee]xpir(?:y|es|ation)\s+[Dd]ate',
            r'[Vv]alid\s+[Tt]o',
            r'[Ee]nd\s+[Dd]ate',
            r'[Tt]ermination\s+[Dd]ate',
        ],
    }

    WINDOW = 500  # chars either side of the match

    for field, patterns in _FIELD_PATTERNS.items():
        if defaults.get(field):
            continue  # already set by header

        # Find candidate region using regex
        match_pos = None
        regex_value = None
        for pat in patterns:
            chunk = all_text[:50000]
            m = re.search(pat, chunk)
            if m:
                match_pos = m.start()
                # Try to extract value from capturing group
                if m.lastindex and m.lastindex >= 1:
                    regex_value = m.group(1).strip()
                break

        if match_pos is None:
            continue  # no candidate found in text

        if invoke_fn:
            # Extract a wide window around the match
            start = max(0, match_pos - WINDOW)
            end = min(len(all_text), match_pos + WINDOW)
            text_window = all_text[start:end]

            llm_result = _llm_extract_defaults(
                invoke_fn, [field], "", text_window, "document text"
            )
            llm_value = llm_result.get(field, "")
            if llm_value:
                if field == "CONTRACT_ID":
                    llm_value = _clean_contract_id(llm_value)
                defaults[field] = llm_value
                continue

        # Fallback: use the regex-extracted value directly
        if regex_value:
            if field == "CONTRACT_ID":
                regex_value = _clean_contract_id(regex_value)
                defaults[field] = regex_value
        elif field == "CURRENCY":
            defaults["CURRENCY"] = _detect_currency(all_text)
        elif field in ("VALID_FROM", "VALID_TO"):
            dates = _detect_dates(all_text)
            if field == "VALID_FROM" and dates.get("valid_from"):
                defaults["VALID_FROM"] = dates["valid_from"]
            elif field == "VALID_TO" and dates.get("valid_to"):
                defaults["VALID_TO"] = dates["valid_to"]


def _llm_extract_defaults(
    invoke_fn, fields: List[str], header_text: str,
    context_text: str, source_label: str,
) -> Dict[str, str]:
    """
    Ask the LLM to extract specific field values from a text window.

    Returns a dict of {field: value} for each field found.
    """
    field_descriptions = {
        "CONTRACT_ID": "Contract ID / number / reference (digits and alphanumeric code only, no labels)",
        "VALID_FROM": "Start date / Signing date / effective date / commencement date (exact date as written)",
        "VALID_TO": "End date / expiry date / termination date (exact date as written)",
        "COMPANY_CODE": "Company code",
        "CURRENCY": "Currency code (e.g. AUD, USD, EUR, GBP)",
    }

    fields_desc = "\n".join(
        f"- {f}: {field_descriptions.get(f, f)}"
        for f in fields
    )

    prompt = f"""Extract the following fields from the {source_label} text below.

FIELDS TO EXTRACT:
{fields_desc}

TEXT:
\"\"\"
{header_text}

{context_text}
\"\"\"

RULES:
1. Extract the EXACT value as it appears in the text.
2. For CONTRACT_ID, return ONLY the identifier (digits code), not labels like "Contract No." or "Agreement Number".
3. For dates, MUST extract dates only, return the date exactly as written but in "dd/mm/yyyy" format (e.g. "01/01/2025").
4. For CURRENCY, return the 3-letter code (e.g. AUD, USD).
5. For COMPANY CODE, return only the company code NOT company name.
6. Validate the field to see if it make sense. If a field is not found or does not make sense in the text, return empty string "".
7. Return ONLY a JSON object with the field names as keys.

Return JSON:"""

    try:
        response = invoke_fn(prompt, 256)
        return _parse_defaults_response(response, fields)
    except Exception as e:
        print(f"    LLM defaults extraction failed ({e})")
        return {}


def _parse_defaults_response(response: str, fields: List[str]) -> Dict[str, str]:
    """Parse the LLM response for default field extraction."""
    cleaned = response.strip()

    # Strip markdown fences
    if cleaned.startswith("```"):
        first_nl = cleaned.find("\n")
        if first_nl > 0:
            cleaned = cleaned[first_nl + 1:]
    if cleaned.endswith("```"):
        cleaned = cleaned[:-3]
    cleaned = cleaned.strip()

    # Try to parse as JSON
    try:
        obj = json.loads(cleaned)
        if isinstance(obj, dict):
            result = {}
            for f in fields:
                val = obj.get(f, "")
                if val and str(val).strip() not in ("null", "None", "N/A", ""):
                    result[f] = str(val).strip()
            return result
    except json.JSONDecodeError:
        pass

    # Fallback: find JSON object anywhere in response
    m = re.search(r'\{[\s\S]*\}', cleaned)
    if m:
        try:
            obj = json.loads(m.group(0))
            if isinstance(obj, dict):
                result = {}
                for f in fields:
                    val = obj.get(f, "")
                    if val and str(val).strip() not in ("null", "None", "N/A", ""):
                        result[f] = str(val).strip()
                return result
        except json.JSONDecodeError:
            pass

    return {}


# ======================================================================
# Pass 2 — Per-table context scan
# ======================================================================

def extract_table_context(
    table_info: Dict,
    doc_defaults: Dict[str, str],
) -> Dict[str, str]:
    """
    Scan a pricing table's local context for values that override
    or supplement the document defaults.

    The local context includes:
      - The section heading / name
      - Paragraphs before and after the table
      - The table caption
    """
    overrides: Dict[str, str] = {}

    # Build local text from context
    parts = []
    parts.append(table_info.get("section_heading", ""))
    parts.append(table_info.get("source_section", ""))
    parts.append(table_info["table"].get("caption", ""))
    parts.extend(table_info.get("context_before", []))
    parts.extend(table_info.get("context_after", []))
    local_text = " ".join(p for p in parts if isinstance(p, str) and p.strip())

    # --- CURRENCY (local override) ---
    local_currency = _detect_currency(local_text)
    if local_currency:
        overrides["CURRENCY"] = local_currency

    # --- VALID_FROM / VALID_TO (local override) ---
    local_dates = _detect_dates(local_text)
    if local_dates.get("valid_from"):
        overrides["VALID_FROM"] = local_dates["valid_from"]
    if local_dates.get("valid_to"):
        overrides["VALID_TO"] = local_dates["valid_to"]

    # --- CONTRACT_ID (local override) ---
    m = re.search(
        r'[Cc]ontract\s+(?:[Nn]o\.?|[Nn]umber|[Rr]ef)[:\s]+([A-Z0-9][\w\-/]+)',
        local_text,
    )
    if m:
        overrides["CONTRACT_ID"] = m.group(1).strip()

    # --- CLAUSE_DESCRIPTION ---
    section = table_info.get("source_section", "")
    caption = table_info["table"].get("caption", "")
    clause = section
    if caption and caption != section:
        clause = f"{section} - {caption}" if section else caption
    overrides["CLAUSE_DESCRIPTION"] = clause

    # --- PRICE_UNIT_UOM from headers or context ---
    headers_text = " ".join(
        str(h) for h in table_info["table"].get("headers", [])
    ).lower()
    uom = _detect_uom(headers_text + " " + local_text.lower())
    if uom:
        overrides["PRICE_UNIT_UOM"] = uom

    return overrides


# ======================================================================
# Detection helpers
# ======================================================================

def _header_val(header: Dict, key: str) -> str:
    """Get a string value from the document header."""
    val = header.get(key)
    if isinstance(val, dict):
        return val.get("text", "").strip()
    if isinstance(val, str):
        return val.strip()
    return ""


def _detect_currency(text: str) -> str:
    """Detect currency from text."""
    patterns = [
        (r'\b(AUD|USD|EUR|GBP|NZD|CAD|SGD|HKD|JPY|CNY|INR)\b', None),
        (r'[Aa]ustralian\s+[Dd]ollars?', "AUD"),
        (r'US\s+[Dd]ollars?', "USD"),
        (r'[Ss]terling|[Pp]ounds?\s+[Ss]terling', "GBP"),
    ]
    for pat, override in patterns:
        m = re.search(pat, text)
        if m:
            return override or m.group(1)
    return ""


def _detect_dates(text: str) -> Dict[str, str]:
    """Detect valid_from / valid_to dates from text."""
    dates: Dict[str, str] = {}
    patterns = [
        # "Effective Date: 1 January 2025" / "effective from 1/1/2025"
        (r'[Ee]ffective\s+(?:[Dd]ate|[Ff]rom)[:\s]+(\d{1,2}[\s/\-]\w+[\s/\-]\d{2,4})', "valid_from"),
        (r'[Cc]ommencement\s+[Dd]ate[:\s]+(\d{1,2}[\s/\-]\w+[\s/\-]\d{2,4})', "valid_from"),
        (r'[Vv]alid\s+[Ff]rom[:\s]+(\d{1,2}[\s/\-]\w+[\s/\-]\d{2,4})', "valid_from"),
        (r'[Ss]tart\s+[Dd]ate[:\s]+(\d{1,2}[\s/\-]\w+[\s/\-]\d{2,4})', "valid_from"),
        (r'[Ee]xpir(?:y|es|ation)\s+[Dd]ate[:\s]+(\d{1,2}[\s/\-]\w+[\s/\-]\d{2,4})', "valid_to"),
        (r'[Vv]alid\s+[Tt]o[:\s]+(\d{1,2}[\s/\-]\w+[\s/\-]\d{2,4})', "valid_to"),
        (r'[Ee]nd\s+[Dd]ate[:\s]+(\d{1,2}[\s/\-]\w+[\s/\-]\d{2,4})', "valid_to"),
        (r'[Tt]ermination\s+[Dd]ate[:\s]+(\d{1,2}[\s/\-]\w+[\s/\-]\d{2,4})', "valid_to"),
        # ISO-ish: 2025-01-01
        (r'[Ee]ffective[:\s]+(\d{4}-\d{2}-\d{2})', "valid_from"),
        (r'[Ee]xpir[:\s]+(\d{4}-\d{2}-\d{2})', "valid_to"),
    ]
    for pat, field in patterns:
        m = re.search(pat, text)
        if m and field not in dates:
            dates[field] = m.group(1).strip()
    return dates


def _detect_uom(text: str) -> str:
    """Detect unit of measure from text."""
    patterns = [
        (r'\$/\s*hr|per\s+hour|hourly', "Hour"),
        (r'\$/\s*day|per\s+day|daily', "Day"),
        (r'\$/\s*week|per\s+week|weekly', "Week"),
        (r'\$/\s*month|per\s+month|monthly', "Month"),
        (r'\$/\s*year|per\s+annum|annual', "Year"),
        (r'per\s+tonne|per\s+ton|\$/t\b', "Tonne"),
        (r'per\s+kg|per\s+kilogram', "Kilogram"),
        (r'per\s+unit|each|per\s+item', "Each"),
        (r'per\s+metre|per\s+meter|\$/m\b', "Metre"),
        (r'per\s+sq\s*m|per\s+m2|per\s+square\s+met', "Square Metre"),
        (r'lump\s+sum|fixed\s+(fee|price)', "Lump Sum"),
        (r'per\s+trip', "Trip"),
        (r'per\s+shift', "Shift"),
        (r'per\s+lot', "Lot"),
    ]
    for pat, uom in patterns:
        if re.search(pat, text):
            return uom
    return ""


def _collect_all_text(data: Any) -> str:
    """Recursively collect all text from a document structure."""
    parts: List[str] = []
    _walk(data, parts)
    return " ".join(parts)


def _walk(data: Any, acc: List[str]) -> None:
    if isinstance(data, str) and data.strip():
        acc.append(data.strip())
    elif isinstance(data, dict):
        for k, v in data.items():
            if not k.startswith("_") and k != "image":
                _walk(v, acc)
    elif isinstance(data, list):
        for item in data:
            _walk(item, acc)


# ======================================================================
# Table scanning
# ======================================================================

class DocumentScanner:
    """Find pricing tables and their local context."""

    def __init__(self, document: Dict, invoke_fn=None):
        self.document = document
        self.invoke_fn = invoke_fn

    def find_pricing_tables(self) -> List[Dict]:
        """
        Returns list of dicts with:
          - table: {caption, headers, rows}
          - source_section: section name
          - section_heading: heading text
          - context_before: paragraphs before the table
          - context_after: paragraphs after the table
        """
        results = []
        sections = self.document.get("sections", [])
        if not sections:
            sections = self._from_old_format()
        for sec in sections:
            name = sec.get("section_name", sec.get("heading", ""))
            heading = sec.get("heading", name)
            content = sec.get("content", [])
            if isinstance(content, list):
                self._scan(content, name, heading, results)
        return results

    def _scan(self, content, section_name, section_heading, results,
              parent_ctx=None, scan_state=None):
        ctx = list(parent_ctx or [])

        # Mutable container shared across recursive subsection calls
        # so headerless continuations are detected across boundaries
        if scan_state is None:
            scan_state = {"last_pricing_table": None}

        i = 0
        while i < len(content):
            block = content[i]
            if not isinstance(block, dict):
                i += 1
                continue

            bt = block.get("type", "")

            if bt == "paragraph":
                t = block.get("text", "")
                if t:
                    ctx.append(t)
                i += 1

            elif bt == "table":
                is_pricing = self._is_pricing(block, ctx, section_name)

                # If not detected as pricing, check if it's a headerless
                # continuation of a recent pricing table in this section
                if not is_pricing and scan_state["last_pricing_table"] is not None:
                    if self._is_orphan_continuation(block, scan_state["last_pricing_table"]):
                        is_pricing = True

                if is_pricing:
                    # Absorb adjacent continuation tables
                    consolidated, absorbed_end = self._absorb_continuations(content, i)

                    # Validate and fix headers (e.g. real headers in rows[0])
                    self._fix_table_headers(consolidated)

                    # If headerless, inherit headers from last pricing table
                    if (scan_state["last_pricing_table"] is not None
                            and not consolidated.get("headers")
                            and self._col_count_compatible(
                                consolidated, scan_state["last_pricing_table"])):
                        prev_headers = scan_state["last_pricing_table"].get("headers", [])
                        if prev_headers:
                            consolidated["headers"] = list(prev_headers)
                            consolidated["_inherited_headers"] = True
                            print(f"    Inherited headers from previous table: "
                                  f"{' | '.join(str(h) for h in prev_headers)}")

                    # Remember this as the last pricing table
                    scan_state["last_pricing_table"] = consolidated

                    # Collect context after absorbed content
                    after_idx = absorbed_end

                    after = []
                    for j in range(after_idx, min(after_idx + 6, len(content))):
                        if j >= len(content):
                            break
                        nxt = content[j]
                        if isinstance(nxt, dict):
                            if nxt.get("type") == "paragraph":
                                after.append(nxt.get("text", ""))
                            elif nxt.get("type") == "subsection":
                                break

                    results.append({
                        "table": consolidated,
                        "source_section": section_name,
                        "section_heading": section_heading,
                        "context_before": ctx[-10:],
                        "context_after": after,
                    })
                    i = after_idx
                else:
                    i += 1

            elif bt == "subsection":
                sub_name = block.get("heading", "") or section_name
                self._scan(
                    block.get("content", []),
                    sub_name, section_heading,
                    results, ctx, scan_state,
                )
                i += 1

            else:
                i += 1

    @staticmethod
    def _fix_table_headers(table: Dict) -> None:
        """
        Validate and fix table headers.

        Detects cases where the real header row ended up in rows[0]
        while the headers field contains a title or is malformed.

        Cases handled:
          1. headers has fewer columns than rows → headers is a title,
             rows[0] is the real header (if rows[0] is mostly text)
          2. headers is a single element → likely a title/caption,
             check if rows[0] is the real header
          3. headers is empty but rows[0] is all text → promote rows[0]
        """
        headers = table.get("headers", [])
        rows = table.get("rows", [])
        if not rows:
            return

        first_row = rows[0]
        if not isinstance(first_row, list) or not first_row:
            return

        # Determine the typical row width from data rows
        data_widths = [len(r) for r in rows if isinstance(r, list)]
        if not data_widths:
            return
        typical_width = max(set(data_widths), key=data_widths.count)

        needs_fix = False
        title_text = ""

        # Case 1: headers has fewer columns than the data rows
        if headers and len(headers) < typical_width:
            # Check if rows[0] matches the typical width and looks like headers
            if len(first_row) == typical_width and _is_header_row(first_row):
                needs_fix = True
                title_text = " | ".join(str(h) for h in headers)

        # Case 2: headers is a single element (likely a title)
        elif headers and len(headers) == 1:
            if len(first_row) >= 2 and _is_header_row(first_row):
                needs_fix = True
                title_text = str(headers[0]).strip()

        # Case 3: no headers but rows[0] is all text
        elif not headers or not any(str(h).strip() for h in headers):
            if len(first_row) >= 2 and _is_header_row(first_row):
                # Only promote if this looks like a header row and subsequent
                # rows look like data rows (have numeric values)
                if len(rows) >= 2:
                    second_row = rows[1] if isinstance(rows[1], list) else []
                    has_numeric = any(
                        re.match(r'^[$€£]?\s*[\d,]+\.?\d*$', str(c).strip())
                        for c in second_row
                    )
                    if has_numeric:
                        needs_fix = True

        if needs_fix:
            old_headers = list(headers) if headers else []
            table["headers"] = [str(h).strip() for h in first_row]
            table["rows"] = rows[1:]  # remove promoted row

            # If old headers contained a title, merge it into caption
            if title_text:
                existing_caption = (table.get("caption", "") or "").strip()
                if title_text and title_text not in existing_caption:
                    if existing_caption:
                        table["caption"] = f"{existing_caption} - {title_text}"
                    else:
                        table["caption"] = title_text

            print(f"    Fixed headers: {old_headers} -> {table['headers']}")

    def _is_orphan_continuation(self, table: Dict, last_pricing: Dict) -> bool:
        """
        Check if a headerless table is a separated continuation of a
        previous pricing table (not immediately adjacent).

        Matches if: no headers, no caption, compatible column count,
        and has rows with data.
        """
        headers = table.get("headers", [])
        caption = (table.get("caption", "") or "").strip()
        rows = table.get("rows", [])

        if not rows:
            return False
        if headers and any(str(h).strip() for h in headers):
            return False
        if caption:
            return False

        return self._col_count_compatible(table, last_pricing)

    @staticmethod
    def _col_count_compatible(table_a: Dict, table_b: Dict) -> bool:
        """Check if two tables have compatible column counts (within +/- 1)."""
        def _count(t):
            h = t.get("headers", [])
            if h:
                return len(h)
            rows = t.get("rows", [])
            if rows:
                return max((len(r) for r in rows if isinstance(r, list)), default=0)
            return 0
        a, b = _count(table_a), _count(table_b)
        if a == 0 or b == 0:
            return True
        return abs(a - b) <= 1

    def _absorb_continuations(self, content: List, start_idx: int) -> tuple:
        """
        Starting from a pricing table at start_idx, look forward for
        continuation tables (no headers, compatible column count) and
        merge their rows into the base table.

        Skips past intervening paragraphs to find continuations that
        are separated by notes or sub-headings.

        Returns (consolidated_table_dict, end_idx) where end_idx is
        the index past the last absorbed element.
        """
        base = content[start_idx]
        consolidated = {
            "type": "table",
            "caption": base.get("caption", ""),
            "headers": list(base.get("headers", [])),
            "rows": [list(r) for r in base.get("rows", []) if isinstance(r, list)],
        }

        j = start_idx + 1
        while j < len(content):
            nxt = content[j]
            if not isinstance(nxt, dict):
                break

            nxt_type = nxt.get("type", "")

            if nxt_type == "table" and self._is_continuation_table(nxt, consolidated):
                cont_rows = nxt.get("rows", [])
                consolidated["rows"].extend(cont_rows)
                j += 1
                continue

            elif nxt_type == "paragraph":
                # Look ahead past paragraph(s) for another continuation table
                lookahead = j + 1
                while lookahead < len(content):
                    la = content[lookahead]
                    if isinstance(la, dict) and la.get("type") == "paragraph":
                        lookahead += 1  # skip multiple paragraphs
                        continue
                    break

                if (lookahead < len(content)
                        and isinstance(content[lookahead], dict)
                        and content[lookahead].get("type") == "table"
                        and self._is_continuation_table(content[lookahead], consolidated)):
                    # Skip the paragraph(s) and absorb the table
                    cont_rows = content[lookahead].get("rows", [])
                    consolidated["rows"].extend(cont_rows)
                    j = lookahead + 1
                    continue
                break
            else:
                break

        absorbed = len(consolidated["rows"]) - len(base.get("rows", []))
        if absorbed > 0:
            print(f"    Absorbed {absorbed} continuation rows into "
                  f"'{consolidated['caption']}' ({len(consolidated['rows'])} total)")

        return consolidated, j  # return end index

    @staticmethod
    def _is_continuation_table(table: Dict, base: Dict) -> bool:
        """
        Check if a table is a headerless continuation of a base table.

        A continuation table:
          - Has no headers (or empty headers list)
          - Has no caption (or empty caption)
          - Has a compatible column count (same as base, or within ±1)
          - Has rows with data
        """
        headers = table.get("headers", [])
        caption = (table.get("caption", "") or "").strip()
        rows = table.get("rows", [])

        # Must have rows
        if not rows:
            return False

        # Must have no headers or empty headers
        if headers and any(str(h).strip() for h in headers):
            return False

        # Must have no caption
        if caption:
            return False

        # Column count should be compatible
        base_cols = 0
        base_headers = base.get("headers", [])
        base_rows = base.get("rows", [])
        if base_headers:
            base_cols = len(base_headers)
        elif base_rows:
            base_cols = max(len(r) for r in base_rows if isinstance(r, list))

        if base_cols > 0:
            cont_cols = max(len(r) for r in rows if isinstance(r, list))
            if abs(cont_cols - base_cols) > 1:
                return False

        return True

    @staticmethod
    def _table_col_count(table: Dict) -> int:
        """Get the column count of a table."""
        headers = table.get("headers", [])
        if headers:
            return len(headers)
        rows = table.get("rows", [])
        if rows:
            return max((len(r) for r in rows if isinstance(r, list)), default=0)
        return 0

    def _is_pricing(self, table, context, section_name):
        # First: exclude key-value metadata tables
        if self._is_kv_table(table):
            return False

        # If LLM is available, ask it to classify the table
        if self.invoke_fn:
            return self._is_pricing_llm(table, context, section_name)

        # Fallback: rule-based detection
        return self._is_pricing_rules(table, context, section_name)

    def _is_pricing_rules(self, table, context, section_name):
        """Rule-based pricing table detection (fallback when no LLM)."""
        headers = table.get("headers", [])
        caption = (table.get("caption", "") or "").lower()
        rows = table.get("rows", [])
        section_lower = section_name.lower()
        headers_text = " ".join(str(h).lower() for h in headers)

        # Check caption and headers for pricing keywords
        combined = f"{caption} {headers_text} {section_lower}"
        if any(kw in combined for kw in _PRICING_KW):
            return True

        # Check for numeric columns + pricing context
        has_numeric = False
        for row in rows[:5]:
            if isinstance(row, list):
                for cell in row:
                    if re.match(r'^[$€£]?\s*[\d,]+\.?\d*$', str(cell).strip()):
                        has_numeric = True
                        break
            if has_numeric:
                break

        if has_numeric:
            if any(kw in section_lower for kw in _PRICING_KW):
                return True
            ctx_text = " ".join(context[-5:]).lower()
            if any(kw in ctx_text for kw in _PRICING_KW):
                return True

        return False

    def _is_pricing_llm(self, table, context, section_name):
        """LLM-based pricing table detection."""
        headers = table.get("headers", [])
        caption = (table.get("caption", "") or "")
        rows = table.get("rows", [])

        # Build a compact table summary for the LLM
        headers_str = " | ".join(str(h) for h in headers) if headers else "(no headers)"
        sample_rows = rows[:3]  # send at most 3 rows
        rows_str = "\n".join(
            "  " + " | ".join(str(c) for c in r)
            for r in sample_rows if isinstance(r, list)
        )
        if len(rows) > 3:
            rows_str += f"\n  ... ({len(rows)} rows total)"

        ctx_str = "\n".join(context[-3:]) if context else "(none)"

        prompt = f"""Is this table a PRICING table (containing rates, fees, costs, prices for goods or services)?

Section: "{section_name}"
Caption: "{caption}"
Headers: {headers_str}
Sample rows:
{rows_str}

Answer ONLY "YES" or "NO". A pricing table contains line items with monetary values, rates, or costs.
Tables that are just metadata (dates, contract details, definitions, scope descriptions) are NOT pricing tables."""

        try:
            response = self.invoke_fn(prompt, 16)
            answer = response.strip().upper()
            # Parse: accept YES/NO possibly followed by explanation
            is_pricing = True if "YES" in answer else False
            label = "pricing" if is_pricing else "non-pricing"
            table_id = caption or headers_str[:40] or f"{len(rows)} rows"
            print(f"    LLM classified [{table_id}] as {label}")
            return is_pricing
        except Exception as e:
            print(f"    LLM classification failed ({e}), falling back to rules")
            return self._is_pricing_rules(table, context, section_name)

    @staticmethod
    def _is_kv_table(table: Dict) -> bool:
        """
        Check if a table is a key-value metadata table (not pricing).

        A KV table has most rows where column 0 is a known metadata label.
        """
        rows = table.get("rows", [])
        if not rows:
            return False
        kv_count = 0
        for row in rows:
            if isinstance(row, list) and len(row) >= 2:
                label = str(row[0]).strip().rstrip(":").strip()
                if _match_label_to_field(label):
                    kv_count += 1
        # If more than 30% of rows are KV pairs, it's a metadata table
        return kv_count >= 2 or (len(rows) > 0 and kv_count / len(rows) > 0.3)

    def _from_old_format(self):
        sections = []
        for key, val in self.document.items():
            if key in ("document_id", "document_header"):
                continue
            if isinstance(val, list):
                for item in val:
                    if isinstance(item, dict):
                        sections.append({
                            "section_name": item.get("heading", key),
                            "heading": item.get("heading", ""),
                            "content": item.get("content", []),
                        })
        return sections


# ======================================================================
# Row-level mapping (fallback — no LLM)
# ======================================================================

# Patterns for detecting price-like column headers
_PRICE_COL_RE = re.compile(
    r'rate|price|cost|amount|value|net|fee|charge|\$|total',
    re.IGNORECASE,
)

# Patterns for non-price field columns
_FIELD_PATTERNS = {
    "ITEM_DESCRIPTION": re.compile(
        r'desc|item|service|material|product|name|scope|activity|role|position|title|type|category',
        re.IGNORECASE,
    ),
    "ORDER_UOM": re.compile(r'uom|unit|basis|measure|\bper\b|frequency', re.IGNORECASE),
    "PRICE_UNIT_UOM": re.compile(r'price.?unit|price.?uom|price.?basis', re.IGNORECASE),
    "TARGET_QUANTITY": re.compile(r'qty|quantity|volume|target', re.IGNORECASE),
    "MATERIAL_ID": re.compile(r'\bcode\b|\bid\b|number|ref|sku|part|material.?id', re.IGNORECASE),
    "MIN_RATE": re.compile(r'\bmin\b|minimum|lower|floor', re.IGNORECASE),
    "MAX_RATE": re.compile(r'\bmax\b|maximum|upper|ceil', re.IGNORECASE),
    "CURRENCY": re.compile(r'curr|ccy|currency', re.IGNORECASE),
}

# Currency codes that might appear in price column headers
_CURRENCY_RE = re.compile(r'\b(AUD|USD|EUR|GBP|NZD|CAD|SGD)\b', re.IGNORECASE)


def _classify_columns(headers: List[str]) -> Dict:
    """
    Classify each column header into: price columns, field columns,
    or unclassified.

    Returns:
        {
            "price_cols": [(col_idx, header_text, qualifier, currency), ...],
            "field_cols": {col_idx: field_name, ...},
            "desc_col": col_idx or None,
        }

    For price columns, the "qualifier" is the non-price part of the
    header (e.g. "Dayshift" from "Dayshift AUD", "Night" from "Night Rate").
    """
    price_cols: List[tuple] = []
    field_cols: Dict[int, str] = {}
    desc_col: Optional[int] = None

    used_fields: set = set()

    for ci, header in enumerate(headers):
        h = header.strip()
        h_lower = h.lower()

        # Check if it's a price column
        if _PRICE_COL_RE.search(h_lower) or _is_pure_currency_col(h):
            # Extract qualifier: strip price keywords and currency
            qualifier = _extract_price_qualifier(h)
            currency = ""
            cm = _CURRENCY_RE.search(h)
            if cm:
                currency = cm.group(1).upper()
            price_cols.append((ci, h, qualifier, currency))
            continue

        # Check against field patterns
        matched = False
        for field, pattern in _FIELD_PATTERNS.items():
            if field not in used_fields and pattern.search(h_lower):
                field_cols[ci] = field
                used_fields.add(field)
                if field == "ITEM_DESCRIPTION":
                    desc_col = ci
                matched = True
                break

        # If not matched and no description yet, mark as potential description
        if not matched and desc_col is None:
            # First unclassified text-like column is likely the description
            if not re.match(r'^[\d$€£]', h.strip()):
                desc_col = ci
                field_cols[ci] = "ITEM_DESCRIPTION"
                used_fields.add("ITEM_DESCRIPTION")

    return {
        "price_cols": price_cols,
        "field_cols": field_cols,
        "desc_col": desc_col,
    }


def _is_pure_currency_col(header: str) -> bool:
    """Check if header is just a currency code or currency + qualifier."""
    h = header.strip()
    # "AUD", "USD", "Dayshift AUD", "AUD Rate"
    if _CURRENCY_RE.search(h):
        # Strip currency and see if what's left is empty or a qualifier
        stripped = _CURRENCY_RE.sub('', h).strip()
        if not stripped or not re.search(r'desc|item|name|uom|unit|qty', stripped, re.IGNORECASE):
            return True
    return False


def _extract_price_qualifier(header: str) -> str:
    """
    Extract the qualifier part from a price column header.

    Examples:
        "Dayshift AUD"       -> "Dayshift"
        "Nightshift AUD"     -> "Nightshift"
        "Rate ($/hr)"        -> ""
        "Day Rate"           -> "Day"
        "Night Rate"         -> "Night"
        "Standard Price"     -> "Standard"
        "Premium Cost AUD"   -> "Premium"
        "AUD"                -> ""
        "Rate"               -> ""
    """
    h = header.strip()

    # Remove currency codes
    h = _CURRENCY_RE.sub('', h).strip()

    # Remove common price keywords
    h = re.sub(
        r'\b(rate|price|cost|amount|value|net|fee|charge|total)\b',
        '', h, flags=re.IGNORECASE,
    ).strip()

    # Remove parentheticals like ($/hr)
    h = re.sub(r'\([^)]*\)', '', h).strip()

    # Remove $, punctuation leftovers
    h = re.sub(r'[$€£/]', '', h).strip()
    h = h.strip(' -_')

    return h


def map_table_fallback(
    table_info: Dict,
    doc_defaults: Dict[str, str],
    table_overrides: Dict[str, str],
) -> List[Dict]:
    """
    Map table rows using column-name matching. No LLM needed.

    Handles multi-price columns: if a table has multiple price columns
    (e.g. "Dayshift AUD", "Nightshift AUD"), each row is pivoted into
    N output rows, one per price column, with the price qualifier
    appended to the item description.
    """
    table = table_info["table"]
    headers = table.get("headers", [])
    rows = table.get("rows", [])

    classification = _classify_columns(headers)
    price_cols = classification["price_cols"]
    field_cols = classification["field_cols"]
    desc_col = classification["desc_col"]

    # If still no price columns, try first unclassified column
    if not price_cols:
        for ci, header in enumerate(headers):
            if ci not in field_cols:
                price_cols.append((ci, header.strip(), "", ""))
                break

    result = []
    for ri, row in enumerate(rows, 1):
        if not isinstance(row, list):
            continue

        # Build base item from defaults + field columns
        base_item = dict(doc_defaults)
        base_item.update({k: v for k, v in table_overrides.items() if v})

        # Map non-price columns
        base_desc = ""
        for ci, cell in enumerate(row):
            field = field_cols.get(ci)
            if field:
                val = str(cell).strip()
                base_item[field] = val
                if field == "ITEM_DESCRIPTION":
                    base_desc = val

        # If no description column was mapped, find first non-numeric cell
        if not base_desc:
            for ci, cell in enumerate(row):
                if ci not in {pc[0] for pc in price_cols} and ci not in field_cols:
                    s = str(cell).strip()
                    if s and not re.match(r'^[$€£]?\s*[\d,]+\.?\d*$', s):
                        base_desc = s
                        base_item["ITEM_DESCRIPTION"] = s
                        break

        # Pivot: one output row per price column
        if len(price_cols) > 1:
            # Multi-price: create separate rows with qualifier
            for pc_idx, (ci, header, qualifier, currency) in enumerate(price_cols):
                if ci >= len(row):
                    continue
                cell_val = str(row[ci]).strip()
                if not cell_val:
                    continue

                item = dict(base_item)
                item["NET_PRICE"] = cell_val

                # Append qualifier to description
                if qualifier:
                    if base_desc:
                        item["ITEM_DESCRIPTION"] = f"{base_desc} - {qualifier}"
                    else:
                        item["ITEM_DESCRIPTION"] = qualifier

                # Currency from column header overrides defaults
                if currency:
                    item["CURRENCY"] = currency

                result.append(item)
        else:
            # Single price column (or none)
            item = dict(base_item)
            if price_cols:
                ci = price_cols[0][0]
                if ci < len(row):
                    item["NET_PRICE"] = str(row[ci]).strip()
                currency = price_cols[0][3]
                if currency:
                    item["CURRENCY"] = currency
            else:
                # No price column found — try first numeric cell
                for ci, cell in enumerate(row):
                    s = str(cell).strip()
                    if re.match(r'^[$€£]?\s*[\d,]+\.?\d*$', s):
                        item["NET_PRICE"] = s
                        break

            if not item.get("ITEM_DESCRIPTION"):
                item["ITEM_DESCRIPTION"] = base_desc
            result.append(item)

    return result


# ======================================================================
# Row-level mapping (LLM)
# ======================================================================

_SYSTEM_PROMPT = """You are a contract pricing data extraction expert.
Your task is to extract structured pricing data from contract documents.
You always respond with valid JSON arrays containing the requested fields.
Never respond with empty content — always return at least an empty array []."""

_USER_PROMPT = """Map the pricing table rows below to the target output fields.

TARGET FIELDS:
{fields_desc}

DOCUMENT DEFAULTS (pre-filled for all rows):
{defaults_text}

TABLE CONTEXT:
Section: "{section_name}"
Caption: "{caption}"

Context before table:
{ctx_before}

Context after table:
{ctx_after}

TABLE DATA:
Headers: {headers}
Rows:
{rows_text}

RULES:
1. Map columns by meaning: "Description"/"Item"/"Role" -> ITEM_DESCRIPTION,
   "Rate"/"Price"/"Cost"/"$/hr" -> NET_PRICE, "Unit"/"UOM"/"Per" -> ORDER_UOM, etc.
2. For fields not in the table, use the document defaults above.
3. CLAUSE_DESCRIPTION = section name and table caption.
4. Use empty string "" for unknown fields. Never use null.
5. Copy cell values EXACTLY as they appear.
6. CRITICAL - MULTIPLE PRICE COLUMNS: If the table has more than one
   price/rate/cost column (e.g. "Dayshift AUD" and "Nightshift AUD",
   or "Standard Rate" and "Premium Rate"), you MUST create SEPARATE
   output rows for each price column. For each:
   - Set NET_PRICE to that column's value
   - Append the column qualifier to ITEM_DESCRIPTION
     (e.g. "Senior Engineer - Dayshift", "Senior Engineer - Nightshift")
   - If the column header contains a currency code, use it for CURRENCY
   So a table row with 2 price columns produces 2 output objects.

Example output format:
[
  {example_row}
]

Return ONLY the JSON array now:"""


def map_table_with_llm(
    table_info: Dict,
    doc_defaults: Dict[str, str],
    table_overrides: Dict[str, str],
    invoke_fn,
    max_tokens: int = 32000,
) -> List[Dict]:
    """Map table rows using LLM."""
    table = table_info["table"]
    headers = table.get("headers", [])
    rows = table.get("rows", [])

    rows_text = "\n".join(
        f"  Row {i}: {' | '.join(str(c) for c in r)}"
        for i, r in enumerate(rows, 1) if isinstance(r, list)
    )

    # Merge defaults + overrides
    merged_defaults = dict(doc_defaults)
    merged_defaults.update({k: v for k, v in table_overrides.items() if v})
    defaults_text = "\n".join(
        f"  {k}: {v}" for k, v in merged_defaults.items() if v
    ) or "  (none)"

    fields_desc = "\n".join(f"- {f}" for f in OUTPUT_FIELDS)

    example_obj = json.dumps(
        {f: merged_defaults.get(f, "") or "<value>" for f in OUTPUT_FIELDS},
        indent=4,
    )

    user_prompt = _USER_PROMPT.format(
        fields_desc=fields_desc,
        defaults_text=defaults_text,
        section_name=table_info.get("source_section", ""),
        caption=table.get("caption", ""),
        ctx_before="\n".join(table_info.get("context_before", [])[-5:]),
        ctx_after="\n".join(table_info.get("context_after", [])[:3]),
        headers=" | ".join(str(h) for h in headers),
        rows_text=rows_text,
        example_row=example_obj,
    )

    retries = 3
    response = ""
    for attempt in range(retries):
        try:
            response = invoke_fn(user_prompt, max_tokens)
            items = _parse_llm_response(response)
            # Apply defaults for any empty fields
            for item in items:
                for f in OUTPUT_FIELDS:
                    if not item.get(f):
                        item[f] = merged_defaults.get(f, "")
            return items
        except Exception as e:
            print(f"  LLM attempt {attempt + 1}/{retries} failed: {e}")
            if attempt == 0 and response:
                print(f"  Response preview: {repr(response[:300])}")

    print(f"  All LLM attempts failed. Falling back to column matching.")
    return map_table_fallback(table_info, doc_defaults, table_overrides)


def _parse_llm_response(response: str) -> List[Dict]:
    if not response or not response.strip():
        raise ValueError("Empty LLM response")

    cleaned = response.strip()

    # Strip markdown fences
    if cleaned.startswith("```"):
        # Find end of opening fence line
        first_newline = cleaned.find("\n")
        if first_newline > 0:
            cleaned = cleaned[first_newline + 1:]
        else:
            cleaned = cleaned[3:]
    if cleaned.endswith("```"):
        cleaned = cleaned[:-3]
    cleaned = cleaned.strip()

    # Try to find a JSON array
    # First try: the whole thing is a JSON array
    if cleaned.startswith("["):
        try:
            items = json.loads(cleaned)
            if isinstance(items, list):
                return _validate_items(items)
        except json.JSONDecodeError:
            pass

    # Second try: find [...] anywhere in the response
    m = re.search(r'\[[\s\S]*\]', cleaned)
    if m:
        try:
            items = json.loads(m.group(0))
            if isinstance(items, list):
                return _validate_items(items)
        except json.JSONDecodeError:
            pass

    # Third try: find individual {...} objects and wrap in array
    objects = []
    depth = 0
    start = None
    for i, ch in enumerate(cleaned):
        if ch == '{':
            if depth == 0:
                start = i
            depth += 1
        elif ch == '}':
            depth -= 1
            if depth == 0 and start is not None:
                try:
                    obj = json.loads(cleaned[start:i + 1])
                    objects.append(obj)
                except json.JSONDecodeError:
                    pass
                start = None

    if objects:
        return _validate_items(objects)

    raise ValueError(
        f"No JSON array in LLM response (length={len(response)}, "
        f"starts_with='{response[:80]}')"
    )


def _validate_items(items: List) -> List[Dict]:
    """Ensure all items are dicts with the right keys."""
    result = []
    for item in items:
        if not isinstance(item, dict):
            continue
        row = {f: "" for f in OUTPUT_FIELDS}
        for k, v in item.items():
            if k in row:
                row[k] = str(v).strip() if v and str(v).strip() not in ("null", "None", "N/A") else ""
        result.append(row)
    if not result:
        raise ValueError("No valid items parsed from LLM response")
    return result


# ======================================================================
# Excel writer
# ======================================================================

def write_pricing_excel(
    items: List[Dict],
    output_path: Path,
    doc_id: str = "",
    tables_info: List[Dict] = None,
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing Data"

    H_BG = "FF2B5797"
    H_FG = "FFFFFFFF"
    ALT = "FFF5F7FA"
    thin = Side(style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = Font(name="Arial", bold=True, color=H_FG, size=9)
    hfill = PatternFill("solid", fgColor=H_BG)
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bf = Font(name="Arial", size=9)
    wrap = Alignment(vertical="top", wrap_text=True)

    for ci, field in enumerate(OUTPUT_FIELDS, 1):
        c = ws.cell(row=1, column=ci, value=field)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, border

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for ri, item in enumerate(items, 2):
        for ci, field in enumerate(OUTPUT_FIELDS, 1):
            c = ws.cell(row=ri, column=ci, value=item.get(field, ""))
            c.font, c.alignment, c.border = bf, wrap, border
        if ri % 2 == 0:
            f = PatternFill("solid", fgColor=ALT)
            for ci in range(1, len(OUTPUT_FIELDS) + 1):
                ws.cell(row=ri, column=ci).fill = f

    widths = {
        "CONTRACT_ID": 18, "VALID_FROM": 16, "VALID_TO": 16,
        "COMPANY_CODE": 20, "CURRENCY": 10, "MATERIAL_ID": 14,
        "ITEM_DESCRIPTION": 45, "ORDER_UOM": 12, "TARGET_QUANTITY": 14,
        "MIN_RATE": 12, "MAX_RATE": 12, "PRICE_UNIT_UOM": 14,
        "NET_PRICE": 14, "CLAUSE_DESCRIPTION": 40,
    }
    for ci, field in enumerate(OUTPUT_FIELDS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(field, 14)

    # Source Tables sheet
    if tables_info:
        ws2 = wb.create_sheet("Source Tables")
        cols = ["#", "Section", "Caption", "Headers", "Rows"]
        for ci, col in enumerate(cols, 1):
            c = ws2.cell(row=1, column=ci, value=col)
            c.font, c.fill, c.border = hf, hfill, border
        for ti, t in enumerate(tables_info, 1):
            tbl = t["table"]
            r = ti + 1
            ws2.cell(row=r, column=1, value=ti).border = border
            ws2.cell(row=r, column=2, value=t.get("source_section", "")).border = border
            ws2.cell(row=r, column=3, value=tbl.get("caption", "")).border = border
            ws2.cell(row=r, column=4, value=" | ".join(str(h) for h in tbl.get("headers", []))).border = border
            ws2.cell(row=r, column=5, value=len(tbl.get("rows", []))).border = border
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 35
        ws2.column_dimensions["C"].width = 35
        ws2.column_dimensions["D"].width = 50
        ws2.column_dimensions["E"].width = 8

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


# ======================================================================
# Main
# ======================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Extract pricing tables from section JSON to Excel"
    )
    # parser.add_argument("--input", "-i", required=True, help="Extracted section JSON")
    parser.add_argument("--output", "-o", help="Output Excel path")
    parser.add_argument("--provider", choices=["aws_bedrock", "azure_openai"], help="LLM provider")
    parser.add_argument("--no-llm", action="store_true", help="Skip LLM, use column matching")
    args = parser.parse_args()
    args.input = "output/Contract_9100075152_ Supply of Explosives AD3  - Executed 18.12.2025/final/Contract_9100075152_Supply_of_Explosives_AD3_-_Executed_18.12.2025.json"
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: {input_path} not found")
        sys.exit(1)

    with open(input_path, encoding="utf-8") as f:
        document = json.load(f)

    doc_id = document.get("document_id", input_path.stem)
    output_path = Path(args.output) if args.output else input_path.with_name(f"{input_path.stem}_pricing.xlsx")

    # === Set up LLM ===
    invoke_fn = None
    if not args.no_llm:
        if args.provider:
            os.environ["LLM_PROVIDER"] = args.provider
        try:
            project_root = Path(__file__).resolve().parent.parent
            if str(project_root) not in sys.path:
                sys.path.insert(0, str(project_root))
            from src.tools.llm_provider import invoke_text as _invoke_text

            def _invoke_with_debug(prompt: str, max_tokens: int) -> str:
                """Wrapper with debug logging and system prompt support."""
                try:
                    response = _invoke_text(
                        prompt=prompt,
                        system_prompt=_SYSTEM_PROMPT,
                        max_tokens=max_tokens,
                    )
                except TypeError:
                    response = _invoke_text(
                        prompt=_SYSTEM_PROMPT + "\n\n" + prompt,
                        max_tokens=max_tokens,
                    )

                if not response or not response.strip():
                    print(f"  DEBUG: LLM returned empty response")
                    print(f"  DEBUG: Prompt length: {len(prompt)} chars (~{len(prompt)//4} tokens)")
                    raise ValueError("Empty LLM response")
                return response

            invoke_fn = _invoke_with_debug
            print(f"Using LLM for extraction, table detection, and field mapping")
        except Exception as e:
            print(f"WARNING: LLM unavailable ({e}). Using rule-based extraction.")

    # === Pass 1: Document-level defaults ===
    print(f"\nDocument: {doc_id}")
    print(f"\nPass 1: Scanning document for common fields...")
    doc_defaults = extract_document_defaults(document, invoke_fn=invoke_fn)
    for k, v in doc_defaults.items():
        if v:
            print(f"  {k}: {v}")

    # === Find pricing tables ===
    scanner = DocumentScanner(document, invoke_fn=invoke_fn)
    pricing_tables = scanner.find_pricing_tables()
    print(f"\nFound {len(pricing_tables)} pricing table(s)")
    for i, pt in enumerate(pricing_tables, 1):
        tbl = pt["table"]
        print(f"  {i}. [{pt['source_section']}] '{tbl.get('caption', '')}' ({len(tbl.get('rows', []))} rows)")

    if not pricing_tables:
        print("No pricing tables found.")
        sys.exit(0)

    # === Pass 2 + row mapping per table ===
    all_items: List[Dict] = []
    for i, pt in enumerate(pricing_tables, 1):
        tbl = pt["table"]
        n_rows = len(tbl.get("rows", []))
        print(f"\nPass 2: Table {i}/{len(pricing_tables)} — '{tbl.get('caption', '')}' ({n_rows} rows)")

        # Per-table context scan
        table_overrides = extract_table_context(pt, doc_defaults)
        for k, v in table_overrides.items():
            if v and v != doc_defaults.get(k, ""):
                print(f"  Override: {k} = {v}")

        # Row mapping
        if invoke_fn:
            items = map_table_with_llm(pt, doc_defaults, table_overrides, invoke_fn)
        else:
            items = map_table_fallback(pt, doc_defaults, table_overrides)

        print(f"  Mapped {len(items)} line items")
        all_items.extend(items)

    # === Write Excel ===
    print(f"\nTotal line items: {len(all_items)}")
    result = write_pricing_excel(all_items, output_path, doc_id, pricing_tables)
    print(f"Output: {result}")


if __name__ == "__main__":
    main()