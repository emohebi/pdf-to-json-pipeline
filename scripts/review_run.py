#!/usr/bin/env python3
"""
Run extraction review as a standalone step.

Takes a final extracted document JSON and the source PDF file,
compares the extracted text against each PDF page image using the LLM,
and produces per-page extraction accuracy statistics.

Output:
  - JSON report  : per-page coverage, missed content, errors
  - Excel report : colour-coded per-page statistics
  - Console      : summary statistics

Usage:
    python scripts/review_run.py \\
        --input ./output/final/my_doc.json \\
        --pdf ./input/my_doc.pdf

    python scripts/review_run.py \\
        --input ./output/final/my_doc.json \\
        --pdf ./input/my_doc.pdf \\
        --output ./review_report.json \\
        --provider azure_openai

    python scripts/review_run.py \\
        --input ./output/final/my_doc.json \\
        --pdf ./input/my_doc.pdf \\
        --pages 1-20

    python scripts/review_run.py \\
        --input ./output/final/my_doc.json \\
        --pdf ./input/my_doc.pdf \\
        --excel ./review_report.xlsx
"""
import argparse
import json
import os
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


# ====================================================================
# Reconstruct sections from the final assembled document
# ====================================================================

def _reconstruct_sections_from_document(document: Dict) -> List[Dict]:
    """
    Convert a final pipeline document JSON into the list-of-dicts
    format that the review agent expects.

    Handles both the new flat format and old bucketed format.
    """
    # New flat format
    flat_sections = document.get("sections")
    if isinstance(flat_sections, list) and flat_sections:
        return _from_flat_sections(flat_sections)

    # Old bucketed format
    return _from_bucketed_sections(document)


def _from_flat_sections(flat_sections: List[Dict]) -> List[Dict]:
    """Convert new flat sections array to agent format."""
    result = []
    for entry in flat_sections:
        if not isinstance(entry, dict):
            continue
        name = entry.get("section_name", entry.get("heading", "Unknown"))
        stype = entry.get("section_type", "section")
        page_range = entry.get("page_range", [])

        data = {
            "heading": entry.get("heading", name),
            "heading_level": entry.get("heading_level", ""),
            "content": entry.get("content", []),
        }

        result.append({
            "section_name": name,
            "page_range": page_range,
            "data": data,
            "_metadata": {"section_type": stype},
        })
    return result


def _from_bucketed_sections(document: Dict) -> List[Dict]:
    """Convert old bucketed format to agent format (backward compat)."""
    sections: List[Dict] = []
    skip_keys = {"document_id", "document_header"}
    header = document.get("document_header", {})
    header_section_names = header.get("sections", [])

    for key, value in document.items():
        if key in skip_keys:
            continue
        if isinstance(value, list):
            for item in value:
                name = _derive_name(item, key)
                sections.append({
                    "section_name": name,
                    "page_range": item.get("page_range", []),
                    "data": item,
                    "_metadata": {"section_type": key},
                })
        elif isinstance(value, dict) and value:
            name = _derive_name(value, key)
            sections.append({
                "section_name": name,
                "page_range": value.get("page_range", []),
                "data": value,
                "_metadata": {"section_type": key},
            })

    if header_section_names and sections:
        name_order = {n: i for i, n in enumerate(header_section_names)}
        sections.sort(
            key=lambda s: name_order.get(s["section_name"], 9999)
        )

    return sections


def _derive_name(data: Any, fallback_type: str) -> str:
    if isinstance(data, dict):
        for key in ("heading", "section_name", "section", "caption"):
            val = data.get(key)
            if isinstance(val, str) and val.strip():
                return val.strip()
    return fallback_type.replace("_", " ").title()


# ====================================================================
# Excel report writer
# ====================================================================

def write_review_excel(
    report: Dict[str, Any],
    output_path: Path,
) -> Path:
    """
    Write an Excel report with per-page review statistics.

    Sheet 1: Per-page results (colour-coded by coverage)
    Sheet 2: Summary statistics
    Sheet 3: Missed content details
    Sheet 4: Errors details
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed — skipping Excel report")
        return output_path

    wb = Workbook()

    # Styles
    thin = Side(style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="FF2B5797")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Arial", size=10)
    wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="top")

    # Coverage colour coding
    def _coverage_fill(pct: float) -> PatternFill:
        if pct >= 95:
            return PatternFill("solid", fgColor="FF92D050")  # green
        elif pct >= 80:
            return PatternFill("solid", fgColor="FFC6EFCE")  # light green
        elif pct >= 60:
            return PatternFill("solid", fgColor="FFFFC000")  # amber
        elif pct >= 30:
            return PatternFill("solid", fgColor="FFFF9900")  # orange
        else:
            return PatternFill("solid", fgColor="FFFF0000")  # red

    # ── Sheet 1: Per-page results ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Page Results"

    columns = [
        "Page", "Coverage %", "Elements on Page",
        "Elements Extracted", "Elements Missed",
        "Has Tables", "Table Accuracy %",
        "Missed Items", "Errors", "Notes",
    ]
    for ci, col in enumerate(columns, 1):
        c = ws1.cell(row=1, column=ci, value=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = border

    ws1.freeze_panes = "A2"

    for ri, review in enumerate(report.get("page_reviews", []), 2):
        cov = review.get("coverage_pct", 0)
        missed = review.get("missed_content", [])
        errors = review.get("incorrect_content", [])

        row_data = [
            review.get("page_number", ""),
            cov,
            review.get("total_elements_on_page", 0),
            review.get("elements_extracted", 0),
            review.get("elements_missed", 0),
            "Yes" if review.get("has_tables") else "No",
            review.get("table_accuracy_pct", ""),
            len(missed),
            len(errors),
            review.get("notes", ""),
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.font = body_font
            cell.border = border

            if ci == 2:  # Coverage column
                cell.fill = _coverage_fill(cov)
                cell.alignment = center
            elif ci in (1, 3, 4, 5, 8, 9):
                cell.alignment = center
            else:
                cell.alignment = wrap

    widths = [8, 14, 16, 18, 16, 12, 16, 12, 10, 50]
    for ci, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Summary ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    stats = report.get("statistics", {})
    summary_rows = [
        ("Document", report.get("document_id", "")),
        ("Total Pages", stats.get("total_pages_reviewed", 0)),
        ("", ""),
        ("Average Coverage %", stats.get("average_coverage_pct", 0)),
        ("Median Coverage %", stats.get("median_coverage_pct", 0)),
        ("Min Coverage %", stats.get("min_coverage_pct", 0)),
        ("Max Coverage %", stats.get("max_coverage_pct", 0)),
        ("", ""),
        ("Perfect Coverage (99%+)", stats.get("pages_with_perfect_coverage", 0)),
        ("High Coverage (80-99%)", stats.get("pages_with_high_coverage", 0)),
        ("Low Coverage (<50%)", stats.get("pages_with_low_coverage", 0)),
        ("No Extraction (0%)", stats.get("pages_with_no_extraction", 0)),
        ("", ""),
        ("Pages with Missed Content", stats.get("pages_with_missed_content", 0)),
        ("Pages with Errors", stats.get("pages_with_errors", 0)),
        ("Total Missed Elements", stats.get("total_missed_elements", 0)),
        ("Total Incorrect Elements", stats.get("total_incorrect_elements", 0)),
        ("", ""),
        ("Tables Found", stats.get("total_tables_found", 0)),
        ("Avg Table Accuracy %", stats.get("average_table_accuracy_pct", 0)),
        ("", ""),
        ("Quality Grade", stats.get("quality_grade", "?")),
    ]

    label_font = Font(name="Arial", bold=True, size=11)
    val_font = Font(name="Arial", size=11)

    for ri, (label, value) in enumerate(summary_rows, 1):
        lc = ws2.cell(row=ri, column=1, value=label)
        vc = ws2.cell(row=ri, column=2, value=value)
        lc.font = label_font
        vc.font = val_font

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 25

    # ── Sheet 3: Missed content ───────────────────────────────────────
    ws3 = wb.create_sheet("Missed Content")

    miss_cols = ["Page", "Type", "Description"]
    for ci, col in enumerate(miss_cols, 1):
        c = ws3.cell(row=1, column=ci, value=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = border

    mr = 2
    for review in report.get("page_reviews", []):
        page = review.get("page_number", "")
        for m in review.get("missed_content", []):
            ws3.cell(row=mr, column=1, value=page).border = border
            ws3.cell(row=mr, column=2, value=m.get("type", "")).border = border
            ws3.cell(row=mr, column=3, value=m.get("description", "")).border = border
            mr += 1

    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 80

    # ── Sheet 4: Errors ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Errors")

    err_cols = ["Page", "Type", "Description", "Extracted", "Actual"]
    for ci, col in enumerate(err_cols, 1):
        c = ws4.cell(row=1, column=ci, value=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = border

    er = 2
    for review in report.get("page_reviews", []):
        page = review.get("page_number", "")
        for e in review.get("incorrect_content", []):
            ws4.cell(row=er, column=1, value=page).border = border
            ws4.cell(row=er, column=2, value=e.get("type", "")).border = border
            ws4.cell(row=er, column=3, value=e.get("description", "")).border = border
            ws4.cell(row=er, column=4, value=e.get("extracted", "")).border = border
            ws4.cell(row=er, column=5, value=e.get("actual", "")).border = border
            er += 1

    ws4.column_dimensions["A"].width = 8
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 50
    ws4.column_dimensions["D"].width = 40
    ws4.column_dimensions["E"].width = 40

    # Save
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


# ====================================================================
# Main
# ====================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Review extraction quality by comparing extracted "
                    "JSON against source PDF pages",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    # parser.add_argument(
    #     "--input", "-i", required=True,
    #     help="Path to the final extracted JSON from the pipeline",
    # )
    # parser.add_argument(
    #     "--pdf", "-p", required=True,
    #     help="Path to the source PDF file",
    # )
    parser.add_argument(
        "--output", "-o",
        help="Path to write the JSON review report. "
             "If omitted, saves to the pipeline's intermediate directory.",
    )
    parser.add_argument(
        "--excel", "-e",
        help="Path to write an Excel review report. "
             "If omitted, saves alongside the JSON report.",
    )
    parser.add_argument(
        "--pages",
        help="Page range to review, e.g. '5-20' or '10'. "
             "If omitted, all pages are reviewed.",
    )
    parser.add_argument(
        "--provider",
        choices=["aws_bedrock", "azure_openai"],
        help="Override LLM provider",
    )
    args = parser.parse_args()
    args.input = r"output\OD-AM-AD_1_HOC_MAINTENANCE_SERVICES_Executed\final\OD-AM-AD_1_HOC_MAINTENANCE_SERVICES_Executed.json"
    args.pdf = r"input\OD-AM-AD 1 HOC MAINTENANCE SERVICES_Executed.pdf"
    # --- Provider override (must happen before config imports) ---
    if args.provider:
        os.environ["LLM_PROVIDER"] = args.provider

    from config.config_loader import load_config
    from config import settings
    from src.agents.review_agent import ReviewAgent
    from src.utils.pdf_processor import extract_pages

    # --- Validate paths ---
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: Input JSON not found: {input_path}")
        sys.exit(1)

    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        print(f"ERROR: PDF not found: {pdf_path}")
        sys.exit(1)

    # --- Load the document JSON ---
    try:
        with open(input_path, encoding="utf-8") as f:
            document = json.load(f)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON in {input_path}: {e}")
        sys.exit(1)

    if not isinstance(document, dict):
        print(
            f"ERROR: Expected a JSON object, "
            f"got {type(document).__name__}"
        )
        sys.exit(1)

    document_id = document.get("document_id", input_path.stem)

    # --- Reconstruct sections ---
    section_jsons = _reconstruct_sections_from_document(document)
    if not section_jsons:
        print("ERROR: No sections found in the document JSON")
        sys.exit(1)

    print(f"Document: {document_id}")
    print(f"Sections found: {len(section_jsons)}")
    for s in section_jsons:
        stype = s.get("_metadata", {}).get("section_type", "?")
        pr = s.get("page_range", [])
        pr_str = f"pp {pr[0]}-{pr[1]}" if len(pr) == 2 else "pp ?"
        print(f"  [{stype}] {s['section_name']} ({pr_str})")

    # --- Extract PDF pages ---
    print(f"\nExtracting PDF pages: {pdf_path}")
    pages_data = extract_pages(str(pdf_path))
    total_pages = len(pages_data)
    print(f"  {total_pages} pages extracted")

    # --- Filter by page range if specified ---
    review_pages = pages_data
    page_start = 1
    page_end = total_pages

    if args.pages:
        parts = args.pages.strip().split("-")
        page_start = max(1, int(parts[0]))
        page_end = min(total_pages, int(parts[-1]))
        review_pages = pages_data[page_start - 1:page_end]

        # Re-number pages to match the original page numbers
        for i, p in enumerate(review_pages):
            p["page_number"] = page_start + i

        print(f"  Filtered to pages {page_start}-{page_end} "
              f"({len(review_pages)} pages)")

        # Also filter section_jsons to only those with overlapping pages
        filtered_sections = []
        for s in section_jsons:
            pr = s.get("page_range", [])
            if len(pr) == 2:
                if pr[1] >= page_start and pr[0] <= page_end:
                    filtered_sections.append(s)
            else:
                filtered_sections.append(s)
        section_jsons = filtered_sections
        print(f"  {len(section_jsons)} sections overlap with page range")

    print(f"\nProvider: {settings.PROVIDER_NAME}")

    # --- Determine output paths ---
    if args.output:
        json_out = Path(args.output)
    else:
        json_out = settings.FINAL_DIR / f"{document_id}_review.json"

    if args.excel:
        excel_out = Path(args.excel)
    else:
        excel_out = json_out.with_suffix(".xlsx")

    print(f"JSON output:  {json_out}")
    print(f"Excel output: {excel_out}")
    print()

    # --- Run review ---
    reviewer = ReviewAgent()
    report = reviewer.review_document(
        section_jsons, document_id, review_pages,
    )

    # --- Save JSON report ---
    json_out.parent.mkdir(parents=True, exist_ok=True)
    with open(json_out, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"\nJSON report saved: {json_out}")

    # --- Save Excel report ---
    try:
        write_review_excel(report, excel_out)
        print(f"Excel report saved: {excel_out}")
    except Exception as e:
        print(f"WARNING: Excel report failed: {e}")

    # --- Print summary ---
    stats = report.get("statistics", {})
    page_reviews = report.get("page_reviews", [])

    print(f"\n{'=' * 70}")
    print("EXTRACTION REVIEW SUMMARY")
    print(f"{'=' * 70}")
    print(f"  Document:              {document_id}")
    print(f"  Pages reviewed:        {stats.get('total_pages_reviewed', 0)}")
    print(f"  Quality grade:         {stats.get('quality_grade', '?')}")
    print()
    print(f"  COVERAGE:")
    print(f"    Average:             {stats.get('average_coverage_pct', 0):.1f}%")
    print(f"    Median:              {stats.get('median_coverage_pct', 0):.1f}%")
    print(f"    Min:                 {stats.get('min_coverage_pct', 0):.1f}%")
    print(f"    Max:                 {stats.get('max_coverage_pct', 0):.1f}%")
    print()
    print(f"  PAGE BREAKDOWN:")
    print(f"    Perfect (99%+):      {stats.get('pages_with_perfect_coverage', 0)}")
    print(f"    High (80-99%):       {stats.get('pages_with_high_coverage', 0)}")
    print(f"    Low (<50%):          {stats.get('pages_with_low_coverage', 0)}")
    print(f"    No extraction (0%):  {stats.get('pages_with_no_extraction', 0)}")
    print()
    print(f"  ISSUES:")
    print(f"    Pages with misses:   {stats.get('pages_with_missed_content', 0)}")
    print(f"    Pages with errors:   {stats.get('pages_with_errors', 0)}")
    print(f"    Total missed items:  {stats.get('total_missed_elements', 0)}")
    print(f"    Total errors:        {stats.get('total_incorrect_elements', 0)}")

    if stats.get("total_tables_found", 0) > 0:
        print()
        print(f"  TABLES:")
        print(f"    Tables found:        {stats.get('total_tables_found', 0)}")
        print(f"    Avg table accuracy:  {stats.get('average_table_accuracy_pct', 0):.1f}%")

    # Print worst pages
    worst = stats.get("worst_pages", [])
    if worst:
        print()
        print(f"  WORST PAGES (lowest coverage):")
        for wp in worst[:10]:
            flag = ""
            if wp.get("errors", 0) > 0:
                flag = f"  [!{wp['errors']} error(s)]"
            print(
                f"    Page {wp['page']:>4}: "
                f"{wp['coverage_pct']:>5.1f}% coverage, "
                f"{wp['missed']:>2} missed{flag}"
            )

    # Print per-page detail (compact)
    print()
    print(f"  PER-PAGE COVERAGE:")
    for review in page_reviews:
        pn = review.get("page_number", "?")
        cov = review.get("coverage_pct", 0)
        missed = len(review.get("missed_content", []))
        errors = len(review.get("incorrect_content", []))
        status = review.get("review_status", "?")

        bar_len = int(cov / 2)  # 50 chars = 100%
        bar = "#" * bar_len + "." * (50 - bar_len)

        flags = []
        if missed > 0:
            flags.append(f"{missed}m")
        if errors > 0:
            flags.append(f"{errors}e")
        if status == "failed":
            flags.append("FAIL")

        flag_str = f" [{','.join(flags)}]" if flags else ""

        print(f"    p{pn:>4}: [{bar}] {cov:>5.1f}%{flag_str}")

    print(f"\n{'=' * 70}")


if __name__ == "__main__":
    main()