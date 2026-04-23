"""
Field Mapping Agent — classifies extracted tables and maps source fields
to target database columns using LLM structured outputs.

Takes an extracted document JSON, finds all tables, and for each table:
  1. Classifies it as TblInvoice, TblTimesheets, or SKIP
  2. Maps every source field (table columns + section header keys) to the
     most appropriate target database column
  3. Tags each mapping with a transform hint (parse_currency, parse_date, etc.)

Uses Azure OpenAI structured outputs (response_format) to return validated
Pydantic models directly. This is critical for accuracy — structured output
mode constrains the LLM to produce exactly the fields and types defined in
the Pydantic model, avoiding JSON parsing errors entirely.

This agent is standalone — it is NOT part of the main extraction pipeline.
"""
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional
import pandas as pd
import httpx
from enum import Enum
from pydantic import BaseModel, Field

from src.utils import setup_logger, StorageManager

logger = setup_logger("field_mapper")


# =====================================================================
# Pydantic models for structured LLM output
# =====================================================================

class TransformHint(str, Enum):
    """Hint for the normalisation pipeline on how to treat a mapped field."""
    NONE = "none"
    EXPAND_DAYWORK = "expand_daywork"
    PARSE_CURRENCY = "parse_currency"
    PARSE_DATE = "parse_date"


class FieldMapping(BaseModel):
    """A single source -> target field mapping produced by the LLM."""
    source: str = Field(
        description="Source column name or header key exactly as it appears in the extracted JSON."
    )
    target: str = Field(
        description="Target database column name (e.g. 'vendor_name', 'quantity') or 'UNMAPPED' if no match."
    )
    confidence: float = Field(
        ge=0.0, le=1.0,
        description="Confidence score 0.0-1.0."
    )
    transform_hint: TransformHint = Field(
        description="Processing hint for the normalisation pipeline."
    )
    reasoning: str = Field(
        description="Brief explanation of why this mapping was chosen."
    )


class MappingProposal(BaseModel):
    """Complete mapping proposal for one table — returned by the LLM as structured output."""
    target_table: Literal["TblInvoice", "TblTimesheets", "SKIP"] = Field(
        description=(
            "Classification: TblInvoice for invoice data (including supply/delivery documents), "
            "TblTimesheets for timesheet/labour data, "
            "SKIP for tables that should be ignored "
            "(totals_summary, how_to_pay, remittance, tax_summary, signatures)."
        )
    )
    classification_reasoning: str = Field(
        description="Why this table was classified as invoice, timesheet, or skip."
    )
    mappings: list[FieldMapping] = Field(
        description="One mapping per source field (columns + header keys)."
    )
    unmapped_target_fields: list[str] = Field(
        description="Target schema fields that no source field was mapped to (will be NULL in DB)."
    )


# =====================================================================
# Target field schemas — mirror the DDL columns exactly
# =====================================================================

INVOICE_HEADER_TARGETS: List[str] = [
    "vendor_name",
    "invoice_number",
    "invoice_date",
    "due_date",
    "currency",
    "net_amount",
    "gst",
    "gross_amount",
    "description",
    "pricing_type",
    "primary_invoice_flag",
    "discount_rebate",
    "pr_number",
]

INVOICE_LINE_ITEM_TARGETS: List[str] = [
    "description",
    "quantity",
    "uom",
    "unit_price",
    "amount",
    "item_description",
]

TIMESHEET_HEADER_TARGETS: List[str] = [
    "vendor_name",
    "period_start",
    "period_end",
]

TIMESHEET_ENTRY_TARGETS: List[str] = [
    "work_date",
    "staff_equipment_name",
    "quantity",
    "uom",
    "shift_type",
    "rate",
    "charge",
    "position_make",
    "location",
    "approved_date",
    "unique_identifier",
    "item_description",
    "employee_number",
    "trade",
    "pay_type",
    "cost_centre",
    "task_description",
    "work_place",
]

# Descriptions for each target field — included in the prompt so the LLM
# knows what kind of values each column expects and common source-column
# aliases that map to it.
_TARGET_FIELD_DESCRIPTIONS = {
    # Invoice header
    "vendor_name": (
        "Supplier/vendor company name. "
        "Common source names: Supplier, Vendor, Company, From, Billed By, Contractor"
    ),
    "invoice_number": (
        "Invoice or document reference number. "
        "Common source names: Invoice No, Inv #, Invoice Ref, Document No, Tax Invoice Number"
    ),
    "invoice_date": (
        "Date the invoice was issued OR the service period start date. "
        "For SES (Service Entry Sheet) documents, the service period start date "
        "(period_from) is the invoice date. "
        "Common source names: Date, Invoice Date, Tax Date, Document Date, Issued, "
        "Period From, Period Start, Service Date, Service Period From"
    ),
    "due_date": (
        "Payment due date OR the service period end date. "
        "For SES (Service Entry Sheet) documents, the service period end date "
        "(period_to) is the due date. "
        "Common source names: Due Date, Payment Due, Pay By, Terms Date, "
        "Period To, Period End, Service Period To"
    ),
    "currency": (
        "ISO 4217 currency code ONLY (AUD, USD, GBP, NZD). "
        "Must be a 3-letter code, NOT descriptive text like 'Australian Dollars'. "
        "Common source names: Currency, Ccy, Cur"
    ),
    "net_amount": (
        "Invoice total BEFORE tax (ex-GST). This is a HEADER-LEVEL total, not a line item. "
        "Common source names: Subtotal, Net, Total Ex GST, Amount Ex Tax, Net Amount"
    ),
    "gst": (
        "GST or tax amount on the invoice. "
        "Common source names: GST, Tax, VAT, Tax Amount, GST Amount"
    ),
    "gross_amount": (
        "Invoice total INCLUDING tax (inc-GST). This is a HEADER-LEVEL total, not a line item. "
        "Common source names: Total, Total Inc GST, Gross, Amount Inc Tax, Invoice Total, Grand Total, Total Due"
    ),
    "description": (
        "General invoice description or purpose at the HEADER level. "
        "NOT for individual line item descriptions — use line-level 'description' for those. "
        "Common source names: Description, Particulars, Re, Subject, Regarding, Scope"
    ),
    "pricing_type": (
        "A CATEGORY label: LABOUR, MATERIALS, EQUIPMENT, SUBCONTRACTOR, or CONSUMABLES. "
        "Do NOT map worker names, role titles, or item descriptions to this field. "
        "Common source names: Type, Category, Cost Type, Pricing Category"
    ),
    "primary_invoice_flag": (
        "Whether this is the primary invoice (Y/N). Rarely present in source data."
    ),
    "discount_rebate": (
        "Any discount or rebate amount. "
        "Common source names: Discount, Rebate, Adjustment, Credit"
    ),
    "pr_number": (
        "Purchase requisition number. "
        "Common source names: PR, PR No, Requisition, PR Number"
    ),
    # Invoice line items
    "quantity": (
        "Quantity of items, hours, or units for a LINE ITEM. "
        "Common source names: Qty, Quantity, No., Units, Hrs, Hours, Vol, Volume, Count"
    ),
    "uom": (
        "Unit of measure for a line item. "
        "Common source names: UOM, Unit, Per, Basis, Measure, U/M, Each"
    ),
    "unit_price": (
        "Price PER SINGLE UNIT. This is the rate, NOT the line total. "
        "Common source names: Rate, Price, Unit Price, Unit Rate, $/Unit, Unit Cost, Price Each"
    ),
    "amount": (
        "LINE ITEM total = quantity x unit_price. This is a per-row total, NOT the invoice total. "
        "Common source names: Amount, Total, Ext., Extension, Value, Line Total, Amt, Net"
    ),
    "item_description": (
        "Pricing schedule item CODE or reference (e.g. '10.10 Supervisor Mech,DS', 'Item 3.2.1'). "
        "This is an identifier that links back to a contract rate card, NOT free-text description. "
        "Common source names: Item, Item Code, Schedule Item, Ref, Code, Line Ref, Rate Code"
    ),
    # Timesheet header
    "period_start": (
        "Timesheet period start date. "
        "Common source names: Period Start, W/E, Week Ending, From, Start Date, Week Commencing"
    ),
    "period_end": (
        "Timesheet period end date. "
        "Common source names: Period End, To, End Date, Week Ending"
    ),
    # Timesheet entries
    "work_date": (
        "Date the work was performed. "
        "Common source names: Date, Work Date, Day, Shift Date"
    ),
    "staff_equipment_name": (
        "Worker's full name or equipment identifier. "
        "Common source names: Name, Employee, Worker, Staff, Personnel, Operator, Equipment"
    ),
    "shift_type": (
        "Day shift (DS) or night shift (NS). "
        "Common source names: Shift, D/N, Shift Type, DS/NS"
    ),
    "rate": (
        "Hourly or daily PAY RATE for a timesheet entry. "
        "Common source names: Rate, Hourly Rate, $/Hr, Pay Rate, Charge Rate, Daily Rate"
    ),
    "charge": (
        "Total charge for a timesheet entry = quantity x rate. "
        "Common source names: Charge, Total, Amount, Cost, Ext, Value"
    ),
    "position_make": (
        "Job role/position title (for labour) or equipment make/model (for plant). "
        "Common source names: Position, Role, Classification, Trade, Title, Make, Model, Equipment Type"
    ),
    "location": (
        "Work site, pit, or location name. "
        "Common source names: Location, Site, Area, Pit, Mine, Project"
    ),
    "approved_date": (
        "Date the timesheet was approved or signed off. "
        "Common source names: Approved, Approved Date, Sign Off, Authorised"
    ),
    "unique_identifier": (
        "Unique row identifier, employee ID, or reference number. "
        "Common source names: ID, Emp ID, Employee No, Ref, Badge, TS No"
    ),
    # Optional timesheet fields
    "employee_number": (
        "Employee ID, badge number, or personnel number. "
        "Common source names: Employee Number, Emp No, Personnel No, Badge No, "
        "Employee No, Emp ID, Staff ID"
    ),
    "trade": (
        "Trade classification or skill category (e.g. Fitter, Boilermaker, Rigger). "
        "Common source names: Trade, Classification, Skill, Trade Classification"
    ),
    "pay_type": (
        "Pay category or type code (e.g. Normal, OT, TL). "
        "Common source names: Pay Type, Pay Code, Pay Category, Allowance"
    ),
    "cost_centre": (
        "Cost centre or cost code for billing allocation. "
        "Common source names: Cost Centre, Cost Code, CC, Cost Center, GL Code"
    ),
    "task_description": (
        "Description of the work task performed (e.g. 'Fixed Plant UG', 'Rail Maintenance'). "
        "Common source names: Task Description, Task, Work Description, Job Description, Activity"
    ),
    "work_place": (
        "Specific work location within the site (e.g. 'D/S', 'UG', 'Surface'). "
        "Common source names: Work Place, Workplace, Area, Work Area, Work Location"
    ),
}


# =====================================================================
# System prompt
# =====================================================================

_SYSTEM_PROMPT = """\
You are a data-mapping specialist for Australian mining industry invoices and timesheets.

## YOUR TASK

Given a table extracted from a PDF document, you must:

1. **CLASSIFY** the table as TblInvoice, TblTimesheets, or SKIP.
2. **MAP** every source field to the best target database column, or "UNMAPPED".
3. **TAG** each mapping with a transform_hint.

## CLASSIFICATION RULES

**TblInvoice** — The table contains invoice LINE ITEMS with some combination of:
  descriptions, quantities, unit prices, amounts/totals, UOM.
  This INCLUDES supply/delivery documents, purchase orders, and credit notes.
  IMPORTANT: Only tables from section_type "invoice" or "supply_document" should
  be classified as TblInvoice. Tables from "attachment" sections that contain
  labour or equipment breakdowns are TblTimesheets (see below).

**TblTimesheets** — The table contains timesheet/labour/equipment entries with some combination of:
  worker names, dates, hours, shifts (DS/NS), rates, charges, positions/roles.
  IMPORTANT: section_type "other" can contain timesheets — classify by DATA CONTENT, not section_type.
  This INCLUDES:
  - Weekly daywork timesheets with day-of-week columns (Mon, Tue, ..., Fri/Sat/Sun)
  - Individual daily work detail cards (one person, one day, with PM ORDER NUMBER / HOURS columns)
  - Equipment usage tables in timesheet sections (equipment names/IDs with hours/rates)
  - Shutdown timesheets with date+shift sub-columns (e.g. "26/10/2024 Dayshift")
  - labour_charges / equipment_charges tables from "attachment" sections ONLY WHEN they
    contain individual worker/equipment names per row with day-by-day or per-date hours.
  CAUTION: If an attachment table has rate-card codes (e.g. "10.10 Coded Welder DS") with
  aggregated hours and dollar amounts but NO individual worker names and NO per-day columns,
  it is an invoice billing summary — classify as TblInvoice, NOT TblTimesheets.
  The key test: do the rows show INDIVIDUAL PEOPLE/EQUIPMENT with daily detail, or do they
  show BILLING CATEGORIES with total hours and amounts? If billing categories → TblInvoice.

  **MANDATORY TblTimesheets fields** — Every TblTimesheets mapping MUST populate these
  from either table columns or header keys (whichever is available):
    - work_date (from header key "date" or a date column or date+shift column names)
    - staff_equipment_name (from a name/employee column or header key "name")
    - quantity (from hours/qty column or day-of-week columns)
    - position_make (from role/position/occupation/classification column or header)
    - task_description (from task/description column if present)
    - shift_type (from shift columns, DS/NS indicators, or date+shift column names)
    - uom (if a unit-of-measure column exists, map it)
    - approved_date (if an approval/sign-off date exists, map it)
  If a source for any of these exists in the table or headers, you MUST map it.
  Do NOT leave them unmapped when the data is available.

  **staff_equipment_name — STRICT RULE:**
  This field is for human names (e.g. "John Smith") or equipment descriptions
  (e.g. "20T Franna Crane"). NEVER map rate-card codes like "10.10", "10.50",
  schedule item numbers, or billing category codes to staff_equipment_name.
  If the column header says "Employee Name" but the sample data contains codes
  like "10.10", it is item_description, NOT staff_equipment_name.

**SKIP** — The table should be IGNORED. Skip these:
  - Summary/totals tables (e.g. "Total Ex GST / GST / Total Inc GST" with no line items)
  - How-to-pay / remittance / bank details tables
  - Tax summaries (e.g. "Tax Code / Net / Tax" breakdown)
  - Signature blocks / approval tables
  - Tables with fewer than 2 data rows AND no meaningful line items
  - Metadata tables (key-value pairs like "Invoice No: 12345 / Date: 01/01/2025")

## MAPPING RULES

### Source fields
You will receive TWO types of source fields:
  - **Table columns**: Column headers from the extracted table (e.g. "Description", "Qty", "Amount")
  - **Header keys**: Metadata fields from the document/section header (e.g. "title", "date", "organisation").
    These are NOT table columns — they come from the document's metadata block.
    Map them if they correspond to a target field (e.g. header key "date" -> invoice_date).

You must produce ONE mapping entry for every source field (columns + header keys).

### Disambiguation rules (critical for accuracy)

**description vs item_description:**
  - "description" = free-text description of what was supplied/done (e.g. "Labour hire for shutdown week 12")
  - "item_description" = a schedule/rate-card CODE that references a pricing agreement
    (e.g. "10.10 Supervisor Mech,DS", "Item 3.2.1", "SOR-001")
  - If the source column contains free text → map to "description"
  - If the source column contains codes/references → map to "item_description"
  - When in doubt, check sample data: codes are short, alphanumeric, with dots/dashes

**amount vs net_amount vs gross_amount:**
  - "amount" = a LINE-LEVEL total (one per row, = qty x unit_price)
  - "net_amount" = the INVOICE-LEVEL total before tax (appears once, not per row)
  - "gross_amount" = the INVOICE-LEVEL total after tax (appears once, not per row)
  - If the column has different values per row → it's "amount" (line level)
  - If the column has one value for the whole table → it's "net_amount" or "gross_amount"

**document_total / Document Total — STRICT RULE:**
  - If the document shows BOTH a subtotal AND a GST line AND a total (3 separate values),
    then: subtotal → net_amount, GST → gst, total → gross_amount.
  - If the document shows ONLY ONE total with NO separate GST line (common in supply
    documents, delivery documents, and inter-company transfers), then that single total
    is the NET amount (ex-GST). Map it to "net_amount", NOT "gross_amount".
  - "gross_amount" should ONLY be used when you can confirm GST/tax IS included in the value.
  - When in doubt, prefer "net_amount" — most Australian mining supply documents
    show amounts ex-GST.

**unit_price vs rate vs amount:**
  - "unit_price" (invoice) = price per single unit BEFORE multiplication
  - "rate" (timesheet) = hourly/daily charge rate BEFORE multiplication
  - "amount" / "charge" = the RESULT of multiplication (qty x price or hrs x rate)
  - Look at sample data: if Column A x Column B = Column C, then A and B are qty/price, C is amount

**pricing_type — STRICT RULE:**
  Map ONLY if the source contains category labels like LABOUR, MATERIALS, EQUIPMENT.
  Do NOT map worker role names (e.g. "Boilermaker", "Supervisor N/S") to pricing_type.
  Do NOT map item descriptions to pricing_type.

**currency — STRICT RULE:**
  Map ONLY if the source contains ISO 4217 codes (AUD, USD, EUR, GBP).
  Do NOT map disclaimer text like "All amounts shown are in Australian Dollars".
  Do NOT map currency symbols ($, A$) — those are handled by parse_currency transform.

### Fields to NEVER map
Do NOT produce mappings for these — they are extracted by a separate deterministic process:
  - contract_number / contract_no / contract_ref
  - po_number / purchase_order / PO / customer_po / customer_p_o
  - ses_number / SES / service_entry_sheet
  - sap_invoice_reference / sap_ref / sap_invoice

If a source field matches these, map it to "UNMAPPED" with reasoning "handled by reference extractor".

### Fields to NEVER map — additional rules

**sap_invoice_reference / sap_ref:**
  ALWAYS map to UNMAPPED. This is a SAP system reference number — it is NOT an
  invoice number. Even if it looks like it could be an invoice number, it is
  handled separately by the reference extractor. Do NOT map it to invoice_number.

**item_no / Item No. / line number / sequence number:**
  ALWAYS map to UNMAPPED. This is a line SEQUENCE number (e.g. "010", "020", "030")
  that identifies the row position within the table. It is NOT a pricing schedule
  item code. Do NOT map it to item_description.
  - item_description is for schedule CODES like "10.10 Supervisor Mech,DS" or "SOR-001"
  - A line sequence number like "010" or "1" is just a row counter — mark it UNMAPPED.

**quantity_ordered vs quantity_supplied:**
  In supply/delivery documents, there are often TWO quantity columns:
  - "Quantity Ordered" = what was requested (the order amount)
  - "Quantity Supplied" = what was actually delivered
  Map ONLY "Quantity Supplied" (or "Qty Supplied", "Qty Delivered") to "quantity".
  Map "Quantity Ordered" (or "Qty Ordered", "Order Qty") to UNMAPPED — it is
  reference information, not the billable quantity.

**document_title / document_type:**
  ALWAYS map to UNMAPPED. Values like "SUPPLY DOCUMENT", "TAX INVOICE",
  "TIMESHEET" are structural labels that describe the document type, NOT a
  meaningful description of what was supplied or invoiced. Do NOT map to
  "description".

**header keys that duplicate table columns:**
  When a header key provides the same information as a table column (e.g.
  header "amount" = "53,896.68" and table column "Price" = "53,896.68"),
  prefer the TABLE COLUMN mapping. Map the header key to UNMAPPED with
  reasoning "duplicates table column [column_name]".

**price_units / price_unit / pricing_units / price_uom:**
  ALWAYS map to UNMAPPED. These indicate the pricing basis (e.g. "per 1000 KG").
  The pipeline uses these values automatically to normalise unit_price to a
  per-single-unit value during materialisation. Do NOT map to any target field.

### Transform hints
  - **"parse_currency"** — The value contains currency formatting: "$1,234.56", "A$7,573.50", "$-500.00"
  - **"parse_date"** — The value contains a date in any format: "04.01.2026", "25/02/2025", "Sat 29-11-2025", "W/E 10.10.2025"
  - **"expand_daywork"** — The column represents a specific day in a daywork timesheet (Mon, Tue, Wed, Thu, Fri, Sat, Sun)
    OR a composite date+shift column (e.g. "26/10/2024 Dayshift", "Sat 11/01/2025").
    Each such column contains hours worked. Map to "quantity" with hint "expand_daywork".
    The pipeline will pivot these into separate rows per day.
  - **"none"** — No special processing needed.

### Confidence calibration
  - **0.95–1.0** — Exact or near-exact name match with matching data type (e.g. "Invoice Date" -> invoice_date)
  - **0.80–0.94** — Strong semantic match, confirmed by sample data (e.g. "Ext." with "$1,234" values -> amount)
  - **0.60–0.79** — Reasonable match but some ambiguity (e.g. "Total" could be amount or gross_amount)
  - **0.40–0.59** — Uncertain, multiple plausible targets (e.g. "Value" with no clear context)
  - **Below 0.40** — Use UNMAPPED instead. Do not force a low-confidence mapping.

## WORKED EXAMPLES

### Example 1: Invoice line items
Columns: ["Item No", "Description", "Qty", "Unit", "Rate", "Amount"]
Sample: {"Item No": "3.1", "Description": "Mobilisation", "Qty": "1", "Unit": "EA", "Rate": "$5,000.00", "Amount": "$5,000.00"}
→ TblInvoice
  Item No → item_description (0.85, none) — schedule reference code like "3.1"
  Description → description (0.95, none) — free-text line description
  Qty → quantity (0.95, none)
  Unit → uom (0.95, none)
  Rate → unit_price (0.90, parse_currency)
  Amount → amount (0.95, parse_currency)

NOTE: "Item No" maps to item_description ONLY because the sample value "3.1" is a
schedule reference code. If the value were "010" or "1" (a line sequence number),
it should be UNMAPPED instead.

### Example 2: Timesheet daywork
Columns: ["Name", "Position", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Total Hrs", "Rate", "Amount"]
Sample: {"Name": "J Smith", "Position": "Boilermaker", "Mon": "10", "Tue": "10", ..., "Rate": "$85.50", "Amount": "$855.00"}
→ TblTimesheets
  Name → staff_equipment_name (0.95, none)
  Position → position_make (0.90, none) — NOT pricing_type
  Mon → quantity (0.90, expand_daywork)
  Tue → quantity (0.90, expand_daywork)
  ... (each day column → quantity with expand_daywork)
  Total Hrs → UNMAPPED — derived field, not stored
  Rate → rate (0.95, parse_currency)
  Amount → charge (0.90, parse_currency)

### Example 3: SKIP table
Columns: ["", "Amount"]
Rows: [["Subtotal Ex GST", "$45,000.00"], ["GST", "$4,500.00"], ["Total Inc GST", "$49,500.00"]]
→ SKIP — summary totals table, not line items

### Example 4: Supply document
Columns: ["Item No.", "Material Description", "Quantity Ordered", "Quantity Supplied", "UOM", "Unit Price", "Price"]
Sample: {"Item No.": "010", "Material Description": "AMMONIUM NITRATE PRILLED BUL", "Quantity Ordered": "5000,000.00", "Quantity Supplied": "59,530.00", "UOM": "KG", "Unit Price": "905.37", "Price": "53,896.68"}
Header keys include: document_title, sap_invoice_reference, vendor_name, doc_number, date, customer_po, amount, price_units, price_uom
→ TblInvoice
  Item No. → UNMAPPED — line sequence number "010", not a schedule code
  Material Description → description (0.95, none) — what was supplied
  Quantity Ordered → UNMAPPED — order qty, not the delivered/billable qty
  Quantity Supplied → quantity (0.95, none) — actual delivered quantity
  UOM → uom (0.95, none)
  Unit Price → unit_price (0.95, parse_currency) — note: pipeline will auto-adjust using price_units
  Price → amount (0.95, parse_currency) — line total
  document_title → UNMAPPED — structural label "SUPPLY DOCUMENT", not a description
  sap_invoice_reference → UNMAPPED — handled by reference extractor
  vendor_name → vendor_name (0.95, none)
  doc_number → invoice_number (0.90, none) — document reference number
  date → invoice_date (0.90, parse_date)
  customer_po → UNMAPPED — handled by reference extractor (PO number)
  amount → UNMAPPED — duplicates table column "Price"
  price_units → UNMAPPED — pricing basis handled by pipeline during materialisation
  price_uom → UNMAPPED — pricing basis handled by pipeline during materialisation

### Example 5: SES (Service Entry Sheet) document
Columns: ["Service Item No", "Service Description", "GL Account", "WBS Element/Cost Centre", "Qty", "UOM", "Price/unit (Tax excl.)", "Tax (%)", "Tax Amt", "Cost excl tax"]
Sample: {"Service Item No": "10", "Service Description": "FY25 Q1 AN & ANE", "GL Account": "4200010", "WBS Element/Cost Centre": "I-MN19-0031.01.01", "Qty": "119620.200", "UOM": "KG", "Price/unit (Tax excl.)": "0.91", "Tax (%)": "10.0", "Tax Amt": "10833.36", "Cost excl tax": "108333.56"}
Header keys include: vendor_name, reference_number, po_number, period_from, period_to, currency, net_amount, tax_amount, gross_amount, ses_type, po_short_description, claim_submission_date
→ TblInvoice
  Service Item No → UNMAPPED — line sequence number, not a schedule code
  Service Description → description (0.95, none) — what service was provided
  GL Account → UNMAPPED — internal accounting code
  WBS Element/Cost Centre → UNMAPPED — internal cost centre
  Qty → quantity (0.95, none)
  UOM → uom (0.95, none)
  Price/unit (Tax excl.) → unit_price (0.95, parse_currency)
  Tax (%) → UNMAPPED — percentage, not a monetary amount
  Tax Amt → UNMAPPED — line-level tax, use header tax_amount for gst
  Cost excl tax → amount (0.95, parse_currency) — line total ex-tax
  vendor_name → vendor_name (0.95, none)
  reference_number → invoice_number (0.85, none) — SES reference number
  po_number → UNMAPPED — handled by reference extractor
  period_from → invoice_date (0.95, parse_date) — service period start = invoice date
  period_to → due_date (0.90, parse_date) — service period end = due date
  currency → currency (0.95, none) — already ISO code "AUD"
  net_amount → net_amount (0.95, parse_currency) — header-level total ex-tax
  tax_amount → gst (0.95, parse_currency) — header-level GST
  gross_amount → gross_amount (0.95, parse_currency) — header-level total inc-tax
  ses_type → UNMAPPED — structural metadata ("Create"), not a description
  po_short_description → UNMAPPED — PO reference text, not line description
  claim_submission_date → UNMAPPED — internal submission date, not invoice date

## TIMESHEET FORMAT RULES

These are the timesheet formats you will encounter. Apply the correct mapping approach for each:

### Format A: Shutdown / multi-person weekly timesheet with date+shift sub-columns
Columns have composite names like "26/10/2024 Dayshift", "26/10/2024 Training", "WEEKLY TOTALS Dayshift".
- Each date+shift column → quantity (expand_daywork). The date in the column name provides work_date, the shift type (Dayshift/Nightshift/Training/Mob-Demob/Standdown) provides shift_type.
- "WEEKLY TOTALS *" columns → UNMAPPED (derived totals).
- INDIVIDUALS NAME / Employee Name → staff_equipment_name
- ROLE IN SHUTDOWN / Role → position_make
- EMPLOYMENT GROUP / Employment Company → UNMAPPED
- SITE ACCESS NUMBER / BHP Access No → unique_identifier

### Format B: Weekly daywork timesheet with day-name columns
Columns use day names, optionally with dates: "Mon", "Tue", "Sat 11/01/2025", etc.
- Each day column → quantity (expand_daywork)
- Total Hours / Total Hrs → UNMAPPED (derived)
- Employee Name / Name → staff_equipment_name
- Employee Number → employee_number
- Trade → trade
- Pay Type → pay_type
- Rate $ → rate (parse_currency)
- Charge $ → charge (parse_currency)

### Format C: Individual daily work detail card (e.g. BHPOD 665)
A single-person-per-day card with columns like: PM ORDER NUMBER, COST CENTRE, HOURS, TASK DESCRIPTION, WORK PLACE.
The worker's name, date, and occupation come from header keys (timesheet_header), NOT from table columns.
- PM ORDER NUMBER → UNMAPPED (reference field)
- COST CENTRE → cost_centre
- HOURS → quantity
- TASK DESCRIPTION → task_description
- WORK PLACE → work_place
- TOTAL HOURS (row) → UNMAPPED (derived)
- Header key "name" → staff_equipment_name
- Header key "date" → work_date (parse_date)
- Header key "occupation" → position_make
- Header key "personnel_no" → employee_number
- Header key "serial_no" → unique_identifier
- Header key "supervisor_name" → approved_date → UNMAPPED (it's a name, not a date)

### Format D: Equipment usage table in timesheet sections
Tables with equipment descriptions, day-of-week columns, and rates/charges.
Classify as TblTimesheets (NOT SKIP). Equipment description/ID maps to staff_equipment_name.
- Equipment Description / Equipment Number → staff_equipment_name
- Day columns → quantity (expand_daywork)
- Rate $ → rate (parse_currency)
- Charge $ → charge (parse_currency)
- Total → UNMAPPED (derived)

### Example 6: Individual daily work detail card
Columns: ["PM ORDER NUMBER", "COST CENTRE", "HOURS", "TASK DESCRIPTION", "WORK PLACE"]
Sample: {"PM ORDER NUMBER": "1001589", "COST CENTRE": "", "HOURS": "12", "TASK DESCRIPTION": "Fixed Plant UG", "WORK PLACE": ""}
Header keys include: date, name, personnel_no, occupation, serial_no, contractor, supervisor_name, pm_order_number_1, company_name, site_name
→ TblTimesheets
  PM ORDER NUMBER → UNMAPPED — reference field, same as header key pm_order_number_1
  COST CENTRE → cost_centre (0.90, none)
  HOURS → quantity (0.95, none)
  TASK DESCRIPTION → task_description (0.95, none)
  WORK PLACE → work_place (0.90, none)
  date → work_date (0.95, parse_date) — the date this card covers
  name → staff_equipment_name (0.95, none) — worker's name from header
  personnel_no → employee_number (0.90, none) — employee ID from header
  occupation → position_make (0.90, none) — trade/role from header
  serial_no → unique_identifier (0.85, none) — card serial number
  contractor → vendor_name (0.85, none) — contracting company
  supervisor_name → UNMAPPED — name, not a mappable field
  pm_order_number_1 → UNMAPPED — reference field
  company_name → UNMAPPED — client company (BHP), not the vendor
  site_name → location (0.85, none) — work site

### Example 7: Shutdown timesheet with composite date+shift columns
Columns: ["INDIVIDUALS NAME", "ROLE IN SHUTDOWN", "26/10/2024 Dayshift", "26/10/2024 Nightshift", "27/10/2024 Dayshift", "WEEKLY TOTALS Dayshift", "WEEKLY TOTALS Nightshift"]
Sample: {"INDIVIDUALS NAME": "Stuart Irrgang", "ROLE IN SHUTDOWN": "Fitter", "26/10/2024 Dayshift": "12", "26/10/2024 Nightshift": "", "27/10/2024 Dayshift": "12", "WEEKLY TOTALS Dayshift": "60", "WEEKLY TOTALS Nightshift": "0"}
→ TblTimesheets
  INDIVIDUALS NAME → staff_equipment_name (0.95, none)
  ROLE IN SHUTDOWN → position_make (0.90, none)
  26/10/2024 Dayshift → quantity (0.90, expand_daywork) — date=26/10/2024, shift=Dayshift
  26/10/2024 Nightshift → quantity (0.90, expand_daywork) — date=26/10/2024, shift=Nightshift
  27/10/2024 Dayshift → quantity (0.90, expand_daywork) — date=27/10/2024, shift=Dayshift
  WEEKLY TOTALS Dayshift → UNMAPPED — derived weekly total
  WEEKLY TOTALS Nightshift → UNMAPPED — derived weekly total

### Example 8: Equipment usage in timesheet section
Columns: ["Equipment Number", "Equipment Description", "Sat 11/01/2025", "Sun 12/01/2025", "Mon 13/01/2025", "Total", "Rate $", "Charge $"]
Sample: {"Equipment Number": "EQ-001", "Equipment Description": "20T Franna Crane", "Sat 11/01/2025": "", "Sun 12/01/2025": "8", "Mon 13/01/2025": "10", "Total": "18", "Rate $": "$250.00", "Charge $": "$4,500.00"}
→ TblTimesheets
  Equipment Number → unique_identifier (0.85, none)
  Equipment Description → staff_equipment_name (0.90, none) — equipment name goes here
  Sat 11/01/2025 → quantity (0.90, expand_daywork)
  Sun 12/01/2025 → quantity (0.90, expand_daywork)
  Mon 13/01/2025 → quantity (0.90, expand_daywork)
  Total → UNMAPPED — derived total
  Rate $ → rate (0.95, parse_currency)
  Charge $ → charge (0.90, parse_currency)

### Example 9: Labour charges from attachment — billing summary (TblInvoice)
section_type: "attachment", table_id: "labour_charges"
Columns: ["Description", "Hours", "Rate Label", "Rate", "Rate Unit", "Amount"]
Sample: {"Description": "10.10 Coded Welder DS", "Hours": "28.5", "Rate Label": "Hours @ $", "Rate": "$ 112.01", "Rate Unit": "per hour", "Amount": "$ 3,192.29"}
→ TblInvoice (rows are billing CATEGORIES like "10.10 Coded Welder DS", NOT individual worker names)
  Description → item_description (0.90, none) — schedule rate code
  Hours → quantity (0.95, none)
  Rate Label → UNMAPPED — label text
  Rate → unit_price (0.95, parse_currency)
  Rate Unit → uom (0.80, none)
  Amount → amount (0.95, parse_currency)

### Example 10: Labour charges from attachment — per-worker detail (TblTimesheets)
section_type: "attachment", table_id: "labour_charges"
Columns: ["Employee Name", "Position", "Mon", "Tue", "Wed", "Thu", "Fri", "Total Hrs", "Rate", "Charge"]
Sample: {"Employee Name": "J Smith", "Position": "Boilermaker", "Mon": "10", "Tue": "10", "Wed": "10", "Thu": "10", "Fri": "10", "Total Hrs": "50", "Rate": "$85.50", "Charge": "$4,275.00"}
→ TblTimesheets (rows are INDIVIDUAL WORKERS with day-by-day hours)
  Employee Name → staff_equipment_name (0.95, none)
  Position → position_make (0.90, none)
  Mon → quantity (0.90, expand_daywork)
  Tue → quantity (0.90, expand_daywork)
  Wed → quantity (0.90, expand_daywork)
  Thu → quantity (0.90, expand_daywork)
  Fri → quantity (0.90, expand_daywork)
  Total Hrs → UNMAPPED — derived total
  Rate → rate (0.95, parse_currency)
  Charge → charge (0.90, parse_currency)

### Example 11: Equipment charges from attachment — billing summary (TblInvoice)
section_type: "attachment", table_id: "equipment_charges"
Columns: ["Description", "Qty", "UOM", "Rate", "Amount"]
Sample: {"Description": "20T Franna Crane", "Qty": "18", "UOM": "Hrs", "Rate": "$250.00", "Amount": "$4,500.00"}
→ TblInvoice (aggregated equipment billing line, no per-day breakdown)
  Description → description (0.90, none) — equipment type billed
  Qty → quantity (0.95, none)
  UOM → uom (0.90, none)
  Rate → unit_price (0.95, parse_currency)
  Amount → amount (0.90, parse_currency)
"""

# Number of sample rows to include in the prompt
_MAX_SAMPLE_ROWS = 5


# =====================================================================
# Agent
# =====================================================================

class FieldMappingAgent:
    """
    Classify extracted tables and map their fields to target DB columns.

    Uses Azure OpenAI structured outputs (response_format=MappingProposal)
    exclusively. This is the only path that guarantees schema-valid
    responses — text-based JSON parsing is too lossy for production
    accuracy.
    """

    def __init__(self):
        self.storage = StorageManager()
        self._client = None

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def map_document(
        self,
        document_path: str,
        document_id: str = None,
    ) -> Dict[str, Any]:
        """
        Map all tables in an extracted document JSON.

        Args:
            document_path: Path to the final extracted JSON.
            document_id:   Identifier for logging.

        Returns:
            Report dict with mapping proposals for each table.
        """
        doc_path = Path(document_path)
        if document_id is None:
            document_id = doc_path.stem

        with open(doc_path, encoding="utf-8") as f:
            document = json.load(f)

        logger.info(f"[{document_id}] Starting field mapping")

        tables = self._find_all_tables(document)
        logger.info(f"[{document_id}] Found {len(tables)} table(s)")

        if not tables:
            return {
                "document_id": document_id,
                "tables_found": 0,
                "tables_mapped": 0,
                "proposals": [],
            }

        # Ensure client is ready
        client = self._get_client()
        if client is None:
            logger.error(
                f"[{document_id}] Azure OpenAI client not available. "
                f"Structured outputs require Azure OpenAI with the "
                f"openai SDK."
            )
            return {
                "document_id": document_id,
                "tables_found": len(tables),
                "tables_mapped": 0,
                "proposals": [],
                "error": "Azure OpenAI client not configured",
            }

        proposals: List[Dict[str, Any]] = []
        for i, table_info in enumerate(tables, 1):
            table_id = table_info.get("table_id", f"table_{i}")
            section_name = table_info.get("section_name", "Unknown")
            section_type = table_info.get("section_type", "other")
            headers = table_info.get("headers", [])
            rows = table_info.get("rows", [])
            header_keys = table_info.get("header_keys", [])

            if not headers and not rows:
                logger.info(
                    f"[{document_id}]   Skipping empty table {table_id}"
                )
                continue

            sample_rows = self._build_sample_rows(headers, rows)

            logger.info(
                f"[{document_id}]   Mapping table {i}/{len(tables)}: "
                f"'{table_id}' in '{section_name}' "
                f"({len(headers)} cols, {len(rows)} rows, "
                f"{len(header_keys)} header keys, "
                f"{len(sample_rows)} sample rows)"
            )

            proposal = self._map_table(
                client=client,
                table_columns=[str(h) for h in headers],
                sample_rows=sample_rows,
                header_keys=header_keys,
                table_id=table_id,
                section_type=section_type,
                document_id=document_id,
            )

            if proposal is not None:
                proposal_dict = proposal.model_dump()
                proposal_dict["_source"] = {
                    "table_id": table_id,
                    "section_name": section_name,
                    "section_type": section_type,
                    "column_count": len(headers),
                    "row_count": len(rows),
                }
                proposals.append(proposal_dict)

                logger.info(
                    f"[{document_id}]     -> {proposal.target_table} "
                    f"({len(proposal.mappings)} mappings, "
                    f"{len(proposal.unmapped_target_fields)} unmapped)"
                )
            else:
                logger.warning(
                    f"[{document_id}]     -> FAILED to map {table_id}"
                )

        report = {
            "document_id": document_id,
            "tables_found": len(tables),
            "tables_mapped": len(proposals),
            "proposals": proposals,
        }

        # Save
        try:
            report_path = (
                self.storage.final_dir
                / f"{document_id}_field_mapping.json"
            )
            report_path.parent.mkdir(parents=True, exist_ok=True)
            with open(report_path, "w", encoding="utf-8") as f:
                json.dump(report, f, indent=2, ensure_ascii=False)
            logger.info(f"[{document_id}] Report saved: {report_path}")
        except Exception as e:
            logger.warning(f"[{document_id}] Failed to save report: {e}")

        return report

    # ------------------------------------------------------------------
    # Materialisation — apply mappings to produce target DataFrames
    # ------------------------------------------------------------------

    def materialise_tables(
        self,
        document: Dict,
        report: Dict[str, Any],
        document_id: str = "",
    ) -> Dict[str, "pd.DataFrame"]:
        """
        Apply mapping proposals to the source data and produce target
        DataFrames ready for Excel export.

        Takes the original document (with raw table data) and the
        mapping report (with proposals), and produces one DataFrame
        per target table (TblInvoice, TblTimesheets). SKIP tables
        are excluded.

        Args:
            document:    The original extracted JSON document.
            report:      The mapping report from map_document().
            document_id: Identifier for logging.

        Returns:
            Dict with keys "TblInvoice" and/or "TblTimesheets",
            each containing a pandas DataFrame. Empty dict if no
            tables were mapped.
        """
        import pandas as pd

        proposals = report.get("proposals", [])
        if not proposals:
            return {}

        # Rebuild the raw tables keyed by table_id for lookup
        # Use (table_id, index) to avoid losing tables with duplicate IDs
        all_tables = self._find_all_tables(document)
        tables_by_id: Dict[str, list] = {}
        for t in all_tables:
            tables_by_id.setdefault(t["table_id"], []).append(t)

        # Document-level header values ONLY (from extraction_info
        # and document_header). Section-level headers are attached to
        # each raw_table under "_section_headers" and take priority.
        # This prevents values from sibling sections bleeding across.
        header_values = self._extract_header_values(document)

        # Extraction info (contract_number, po_number, etc.)
        extraction_info = document.get("extraction_info", {})

        # Build rows per target table
        target_rows: Dict[str, List[Dict]] = {
            "TblInvoice": [],
            "TblTimesheets": [],
        }

        for proposal in proposals:
            target_table = proposal.get("target_table", "")
            if target_table == "SKIP" or target_table not in target_rows:
                continue

            source = proposal.get("_source", {})
            table_id = source.get("table_id", "")
            candidates = tables_by_id.get(table_id, [])
            raw_table = candidates.pop(0) if candidates else None

            if raw_table is None:
                logger.warning(
                    f"[{document_id}] Cannot find raw table "
                    f"'{table_id}' for materialisation"
                )
                continue

            mappings = proposal.get("mappings", [])
            headers = raw_table.get("headers", [])
            rows = raw_table.get("rows", [])

            # Build the set of valid target columns for this table type.
            # Any LLM mapping to a target outside this set is dropped,
            # preventing timesheet-only fields (e.g. unique_identifier)
            # from leaking into invoice output and vice versa.
            if target_table == "TblInvoice":
                valid_targets = set(
                    INVOICE_HEADER_TARGETS + INVOICE_LINE_ITEM_TARGETS
                )
            elif target_table == "TblTimesheets":
                valid_targets = set(
                    TIMESHEET_HEADER_TARGETS + TIMESHEET_ENTRY_TARGETS
                )
            else:
                valid_targets = set()

            # Build the source->target mapping lookup
            col_mappings: Dict[str, Dict] = {}  # source_col -> {target, hint}
            hdr_mappings: Dict[str, Dict] = {}  # header_key -> {target, hint}

            for m in mappings:
                src = m.get("source", "")
                tgt = m.get("target", "")
                hint = m.get("transform_hint", "none")
                if tgt == "UNMAPPED" or not tgt:
                    continue

                # Reject mappings to fields that don't belong to this
                # table type. Prevents cross-table field pollution.
                if tgt not in valid_targets:
                    logger.warning(
                        f"[{document_id}] Dropping out-of-scope "
                        f"mapping '{src}' -> '{tgt}' for "
                        f"{target_table} (target not in schema)"
                    )
                    continue

                if src in headers:
                    col_mappings[src] = {"target": tgt, "hint": hint}
                else:
                    hdr_mappings[src] = {"target": tgt, "hint": hint}

            # Identify expand_daywork columns (day-of-week columns to pivot)
            daywork_cols: List[tuple] = []  # (col_index, col_name)
            regular_cols: Dict[str, Dict] = {}  # non-daywork col mappings
            for col_name, mapping in col_mappings.items():
                if mapping["hint"] == "expand_daywork":
                    ci = headers.index(col_name)
                    daywork_cols.append((ci, col_name))
                else:
                    regular_cols[col_name] = mapping

            if daywork_cols:
                logger.debug(
                    f"[{document_id}] Expanding {len(daywork_cols)} "
                    f"daywork columns from table '{table_id}'"
                )

            # Pre-compute shared values outside the row loop
            sec_headers = raw_table.get("_section_headers", {})
            section_type = source.get("section_type", "")
            doc_title = sec_headers.get(
                "document_title",
                sec_headers.get("invoice_title", ""),
            )
            is_draft = "draft" in doc_title.lower() if doc_title else False
            table_meta = raw_table.get("_table_metadata", {})

            # Materialise each data row
            for row in rows:
                if not isinstance(row, list):
                    continue

                # Skip header-shaped rows: if 50%+ of non-empty cells
                # match their own column name (case-insensitive), this
                # is a stray column-header row the upstream extractor
                # left in data[0]. Pivoting it produces garbage output.
                if self._looks_like_header_row(row, headers):
                    logger.debug(
                        f"[{document_id}] Skipping header-shaped row "
                        f"in table '{table_id}': {row}"
                    )
                    continue

                # Build the base output row from non-daywork columns
                base_row: Dict[str, str] = {}

                # Apply regular (non-daywork) column mappings
                for ci, col_name in enumerate(headers):
                    if ci >= len(row):
                        continue
                    mapping = regular_cols.get(col_name)
                    if mapping is None:
                        continue

                    target_col = mapping["target"]
                    hint = mapping["hint"]
                    raw_value = str(row[ci]) if row[ci] is not None else ""

                    base_row[target_col] = self._apply_transform(
                        raw_value, hint, target_col
                    )

                # Apply header key mappings (same value for every row)
                _TIMESHEET_DATE_FIELDS = {
                    "approved_date", "period_start", "period_end",
                    "work_date",
                }
                for hdr_key, mapping in hdr_mappings.items():
                    target_col = mapping["target"]
                    hint = mapping["hint"]
                    in_section = hdr_key in sec_headers
                    raw_value = str(
                        sec_headers.get(hdr_key, "")
                        or header_values.get(hdr_key, "")
                    )

                    # Sanity check: flag when a timesheet date field is
                    # populated from a document-level header rather than
                    # the section's own header — this is where stale
                    # dates can sneak in across sections.
                    if (
                        target_table == "TblTimesheets"
                        and target_col in _TIMESHEET_DATE_FIELDS
                        and raw_value
                        and not in_section
                    ):
                        logger.warning(
                            f"[{document_id}] Timesheet date field "
                            f"'{target_col}' populated from non-section "
                            f"header '{hdr_key}' = {raw_value!r} "
                            f"(possible cross-section bleed)"
                        )

                    # Don't overwrite a column-level value with a header value
                    if target_col not in base_row or not base_row[target_col]:
                        base_row[target_col] = self._apply_transform(
                            raw_value, hint, target_col
                        )

                # Add reference fields from extraction_info
                if extraction_info:
                    for ref_field in ("contract_number", "po_number",
                                      "ses_number", "sap_invoice_reference"):
                        val = extraction_info.get(ref_field, "")
                        if val:
                            base_row[ref_field] = str(val)

                # Flag whether this row comes from an invoice section
                base_row["primary_invoice_flag"] = (
                    "1" if section_type.startswith("invoice")
                    and not is_draft else "0"
                )

                # Parse metadata.notes for daily card fallback fields
                if isinstance(table_meta, dict):
                    notes = table_meta.get("notes", "")
                    if notes and isinstance(notes, str):
                        import re as _re
                        for pat, tgt in (
                            (r"(?:name|employee)[:\s]+([^\n,;]+)", "employee_name"),
                            (r"(?:date)[:\s]+([^\n,;]+)", "date"),
                            (r"(?:occupation|trade|position)[:\s]+([^\n,;]+)", "trade"),
                        ):
                            if tgt not in base_row or not base_row[tgt]:
                                match = _re.search(pat, notes, _re.IGNORECASE)
                                if match:
                                    base_row[tgt] = match.group(1).strip()

                # Add source traceability
                base_row["_source_table"] = table_id
                base_row["_source_file"] = extraction_info.get(
                    "source_file", ""
                )

                # ----- Daywork expansion -----
                if daywork_cols:
                    for dw_ci, dw_col_name in daywork_cols:
                        raw_hours = (
                            str(row[dw_ci]) if dw_ci < len(row)
                            and row[dw_ci] is not None else ""
                        )
                        hours = self._apply_transform(
                            raw_hours, "expand_daywork", "quantity"
                        )
                        # Skip zero / empty day-columns
                        if not hours:
                            continue
                        try:
                            if float(hours) == 0:
                                continue
                        except ValueError:
                            pass

                        day_row = dict(base_row)
                        day_row["quantity"] = hours

                        # Parse work_date and shift_type from column name
                        parsed = self._parse_daywork_column(
                            dw_col_name, sec_headers
                        )
                        if parsed["work_date"]:
                            day_row["work_date"] = parsed["work_date"]
                        if parsed["shift_type"] and not day_row.get("shift_type"):
                            day_row["shift_type"] = parsed["shift_type"]

                        target_rows[target_table].append(day_row)
                else:
                    # No daywork columns — emit a single row
                    target_rows[target_table].append(base_row)

        # Dedupe invoice rows that describe the same billing line
        # (e.g. labour_details + invoice_line_items overlap).
        if target_rows.get("TblInvoice"):
            target_rows["TblInvoice"] = self._dedupe_invoice_rows(
                target_rows["TblInvoice"], document_id,
            )

        # Build DataFrames
        result = {}
        for table_name, rows_list in target_rows.items():
            if rows_list:
                df = pd.DataFrame(rows_list)
                # Reorder columns: target schema first, then extras
                schema_cols = (
                    INVOICE_HEADER_TARGETS + INVOICE_LINE_ITEM_TARGETS
                    if table_name == "TblInvoice"
                    else TIMESHEET_HEADER_TARGETS + TIMESHEET_ENTRY_TARGETS
                )
                # Add reference columns
                ref_cols = [
                    "contract_number", "po_number",
                    "ses_number", "sap_invoice_reference",
                    "primary_invoice_flag",
                ]
                # orig columns and confidence scores come right after schema cols
                desc_cols = [
                    "orig_description", "description_norm_confidence",
                    "orig_staff_equipment_name", "staff_equipment_name_norm_confidence",
                    "orig_task_description", "task_description_norm_confidence",
                ]
                trace_cols = ["_source_table", "_source_file"]
                ordered = []
                for c in (
                    schema_cols + ref_cols + desc_cols + trace_cols
                ):
                    if c in df.columns and c not in ordered:
                        ordered.append(c)
                # Add any remaining columns not in the schema
                for c in df.columns:
                    if c not in ordered:
                        ordered.append(c)
                df = df[ordered]

                result[table_name] = df
                logger.info(
                    f"[{document_id}] Materialised {table_name}: "
                    f"{len(df)} rows, {len(df.columns)} columns"
                )

        # Normalise descriptions via LLM
        if result:
            # Always ensure orig columns exist even if normalisation fails
            for df in result.values():
                if "description" in df.columns and "orig_description" not in df.columns:
                    df["orig_description"] = df["description"]
                if "staff_equipment_name" in df.columns and "orig_staff_equipment_name" not in df.columns:
                    df["orig_staff_equipment_name"] = df["staff_equipment_name"]
                if "task_description" in df.columns and "orig_task_description" not in df.columns:
                    df["orig_task_description"] = df["task_description"]

            # Deduplicate staff_equipment_name (OCR variants of same person)
            try:
                self._deduplicate_staff_names(result, document_id)
            except Exception as e:
                logger.warning(
                    f"[{document_id}] Staff name dedup "
                    f"failed (non-fatal): {e}"
                )

            try:
                self._normalise_descriptions(result, document_id)
            except Exception as e:
                logger.warning(
                    f"[{document_id}] Description normalisation "
                    f"failed (non-fatal): {e}"
                )

            try:
                self._normalise_uom(result, document_id)
            except Exception as e:
                logger.warning(
                    f"[{document_id}] UOM normalisation "
                    f"failed (non-fatal): {e}"
                )

        # Clean up TblTimesheets: drop summary rows and fix rate-code names
        if "TblTimesheets" in result:
            df = result["TblTimesheets"]
            before = len(df)

            # Drop rows where quantity, rate, and charge are all empty
            # or zero. These carry no billing/time information.
            def _is_empty_val(v) -> bool:
                if v is None:
                    return True
                s = str(v).strip()
                if not s or s.lower() in ("nan", "none", "null"):
                    return True
                try:
                    return float(s) == 0.0
                except (ValueError, TypeError):
                    return False

            qty_col = (
                df["quantity"] if "quantity" in df.columns
                else pd.Series([None] * len(df))
            )
            rate_col = (
                df["rate"] if "rate" in df.columns
                else pd.Series([None] * len(df))
            )
            charge_col = (
                df["charge"] if "charge" in df.columns
                else pd.Series([None] * len(df))
            )
            all_empty = (
                qty_col.apply(_is_empty_val)
                & rate_col.apply(_is_empty_val)
                & charge_col.apply(_is_empty_val)
            )
            if all_empty.any():
                dropped = int(all_empty.sum())
                df = df[~all_empty].reset_index(drop=True)
                logger.info(
                    f"[{document_id}] Dropped {dropped} "
                    f"TblTimesheets rows with empty qty/rate/charge"
                )

            # Drop rows where staff_equipment_name is a summary label
            if "staff_equipment_name" in df.columns:
                import re as _re
                _summary = df["staff_equipment_name"].apply(
                    lambda v: isinstance(v, str)
                    and v.strip().lower() in (
                        "total", "totals", "grand total", "sub total",
                        "subtotal", "sub-total",
                    )
                )
                # Reclassify rate-card codes (e.g. "10.10") as item_description
                _rate_code = df["staff_equipment_name"].apply(
                    lambda v: isinstance(v, str)
                    and bool(_re.fullmatch(r"\d+(\.\d+)?", v.strip()))
                )
                if _rate_code.any() and "item_description" in df.columns:
                    df.loc[_rate_code, "item_description"] = (
                        df.loc[_rate_code, "staff_equipment_name"]
                    )
                    df.loc[_rate_code, "staff_equipment_name"] = None
                elif _rate_code.any():
                    df["item_description"] = None
                    df.loc[_rate_code, "item_description"] = (
                        df.loc[_rate_code, "staff_equipment_name"]
                    )
                    df.loc[_rate_code, "staff_equipment_name"] = None

                df = df[~_summary]

            # Coerce quantity to numeric; non-numeric values become NaN
            if "quantity" in df.columns:
                df["quantity"] = pd.to_numeric(
                    df["quantity"], errors="coerce"
                )

            if len(df) < before:
                logger.info(
                    f"[{document_id}] Cleaned {before - len(df)} "
                    f"summary/invalid rows from TblTimesheets"
                )
            result["TblTimesheets"] = df

        # Drop rows flagged as non-primary invoices (drafts etc.)
        # Only filter TblInvoice; timesheets don't carry this flag meaningfully.
        for table_name, df in list(result.items()):
            if table_name != "TblInvoice":
                continue
            if "primary_invoice_flag" in df.columns:
                before = len(df)
                df = df[df["primary_invoice_flag"] != "0"]
                if len(df) < before:
                    logger.info(
                        f"[{document_id}] Dropped {before - len(df)} "
                        f"non-primary rows from {table_name}"
                    )
                result[table_name] = df

        return result

    # ------------------------------------------------------------------
    # Staff / equipment name deduplication
    # ------------------------------------------------------------------

    @staticmethod
    def _deduplicate_staff_names(
        materialised: Dict[str, "pd.DataFrame"],
        document_id: str = "",
    ) -> None:
        """
        Unify OCR-garbled variants of the same person/equipment name.

        Strategy: fuzzy string matching on unique names.
        Names sharing the same first token (first name) with
        SequenceMatcher ratio >= 0.55 are clustered together.
        The most frequent name in each cluster is chosen as
        canonical; ties are broken by longest name.

        Employee numbers are NOT used for grouping because OCR
        frequently garbles digits too (e.g. 4010803 vs 4610803).

        Modifies DataFrames in-place. Original names are already
        preserved in orig_staff_equipment_name before this runs.
        """
        from difflib import SequenceMatcher

        col = "staff_equipment_name"
        # Minimum ratio when first tokens match exactly
        FIRST_MATCH_THRESHOLD = 0.55

        for table_name, df in materialised.items():
            if col not in df.columns:
                continue

            unique_names = [
                str(v).strip()
                for v in df[col].dropna().unique()
                if str(v).strip()
            ]
            if len(unique_names) < 2:
                continue

            # Build clusters via pairwise similarity
            clusters: list[list[str]] = []
            assigned: set = set()
            for i, name_a in enumerate(unique_names):
                if name_a in assigned:
                    continue
                cluster = [name_a]
                assigned.add(name_a)
                tokens_a = name_a.lower().split()
                first_a = tokens_a[0] if tokens_a else ""
                if not first_a or len(tokens_a) < 2:
                    # Single-token names (equipment etc.) — skip
                    continue
                for j in range(i + 1, len(unique_names)):
                    name_b = unique_names[j]
                    if name_b in assigned:
                        continue
                    tokens_b = name_b.lower().split()
                    first_b = tokens_b[0] if tokens_b else ""
                    if len(tokens_b) < 2:
                        continue
                    # Require matching first token (first name)
                    if first_a != first_b:
                        continue
                    ratio = SequenceMatcher(
                        None, name_a.lower(), name_b.lower()
                    ).ratio()
                    if ratio >= FIRST_MATCH_THRESHOLD:
                        cluster.append(name_b)
                        assigned.add(name_b)
                if len(cluster) > 1:
                    clusters.append(cluster)

            if not clusters:
                continue

            # Build replacement map: variant → canonical
            # Pick most frequent name; break ties with longest
            fuzzy_map: dict = {}
            for cluster in clusters:
                freq: dict = {}
                for name in cluster:
                    count = int((df[col] == name).sum())
                    freq[name] = count
                max_count = max(freq.values())
                candidates = [n for n, c in freq.items()
                              if c == max_count]
                canonical = max(candidates, key=len)
                for name in cluster:
                    if name != canonical:
                        fuzzy_map[name] = canonical

            if fuzzy_map:
                changed = 0
                for idx, row in df.iterrows():
                    old_name = row.get(col)
                    if isinstance(old_name, str):
                        key = old_name.strip()
                        if key in fuzzy_map:
                            df.at[idx, col] = fuzzy_map[key]
                            changed += 1
                if changed:
                    logger.info(
                        f"[{document_id}] {table_name}: unified "
                        f"{changed} staff names via fuzzy match "
                        f"({fuzzy_map})"
                    )

    # ------------------------------------------------------------------
    # UOM normalisation (deterministic lookup)
    # ------------------------------------------------------------------

    def _normalise_uom(
        self,
        materialised: Dict[str, "pd.DataFrame"],
        document_id: str = "",
    ) -> None:
        """
        Normalise UOM values using a deterministic lookup table.

        Cleans newlines/whitespace, looks up the lowercased value in
        config.uom_items.UOM_MAP, and replaces with the canonical code.
        The original value is preserved in 'orig_uom'.

        Modifies DataFrames in-place.
        """
        try:
            from config.uom_items import UOM_MAP
        except ImportError:
            logger.warning(
                f"[{document_id}] config.uom_items not found "
                "— skipping UOM normalisation"
            )
            return

        changed_total = 0
        for table_name, df in materialised.items():
            if "uom" not in df.columns:
                continue

            # Preserve original
            if "orig_uom" not in df.columns:
                df["orig_uom"] = df["uom"]

            def _normalise(val):
                if not isinstance(val, str):
                    return val
                # Clean OCR artefacts: collapse newlines and extra spaces
                cleaned = " ".join(val.split()).strip()
                key = cleaned.lower()
                if key in UOM_MAP:
                    return UOM_MAP[key]
                return cleaned  # keep cleaned version if not in map

            df["uom"] = df["uom"].apply(_normalise)
            changed = (df["uom"] != df["orig_uom"]).sum()
            changed_total += changed

            if changed:
                logger.info(
                    f"[{document_id}] {table_name}: normalised "
                    f"{changed} UOM values"
                )

        if changed_total:
            logger.info(
                f"[{document_id}] UOM normalisation: "
                f"{changed_total} values normalised"
            )

    # ------------------------------------------------------------------
    # Description normalisation via LLM
    # ------------------------------------------------------------------

    def _normalise_descriptions(
        self,
        materialised: Dict[str, "pd.DataFrame"],
        document_id: str = "",
    ) -> None:
        """
        Normalise description values by matching them to a reference
        list of known item names via LLM.

        For each unique description value across all target tables,
        the LLM picks the closest match from the reference list and
        returns a confidence score. If confidence >= threshold the
        description is replaced; otherwise the original is kept.

        The original value is always preserved in 'orig_description'.

        Modifies DataFrames in-place.
        """
        try:
            from config.description_items import (
                DESCRIPTION_ITEMS,
                DESCRIPTION_MATCH_THRESHOLD,
            )
        except ImportError:
            logger.warning(
                f"[{document_id}] config.description_items not found "
                "— skipping description normalisation"
            )
            return

        if not DESCRIPTION_ITEMS:
            return

        # Collect unique non-empty descriptions across all tables
        # Normalise description, staff_equipment_name, and task_description
        # (item_description contains rate-card codes, not free-text — skip it)
        # (timesheets use staff_equipment_name for role/equipment names that
        # should be normalised against the reference list)
        desc_columns = ["description", "staff_equipment_name", "task_description"]
        unique_descs: set = set()
        for df in materialised.values():
            for col in desc_columns:
                if col in df.columns:
                    for val in df[col].dropna().unique():
                        val = str(val).strip()
                        if val:
                            unique_descs.add(val)

        if not unique_descs:
            logger.info(
                f"[{document_id}] No descriptions to normalise"
            )
            return

        logger.info(
            f"[{document_id}] Normalising {len(unique_descs)} "
            f"unique descriptions against {len(DESCRIPTION_ITEMS)} "
            f"reference items"
        )

        # Cheap pre-check: if no description shares any 4+ char
        # token with any reference item, the LLM will return all
        # non-matches anyway. Skip the API round trip.
        if not self._has_reference_overlap(
            unique_descs, DESCRIPTION_ITEMS,
        ):
            logger.info(
                f"[{document_id}] No token overlap between "
                f"descriptions and reference items — skipping "
                f"LLM description normalisation"
            )
            for df in materialised.values():
                for col in desc_columns:
                    if col in df.columns:
                        df[f"{col}_norm_confidence"] = None
            return

        # Call LLM to match descriptions
        client = self._get_client()
        if client is None:
            logger.warning(
                f"[{document_id}] No Azure client — "
                "skipping description normalisation"
            )
            return

        mapping = self._match_descriptions_via_llm(
            list(unique_descs),
            DESCRIPTION_ITEMS,
            DESCRIPTION_MATCH_THRESHOLD,
            document_id,
        )

        # Always preserve the original description before any
        # normalisation so orig_description is never empty.
        for df in materialised.values():
            for col in desc_columns:
                orig_col = f"orig_{col}"
                if col in df.columns and orig_col not in df.columns:
                    df[orig_col] = df[col]

        if not mapping:
            # No mapping results — set confidence to None
            for df in materialised.values():
                for col in desc_columns:
                    if col in df.columns:
                        df[f"{col}_norm_confidence"] = None
            return

        # Apply to all DataFrames
        for table_name, df in materialised.items():
            changed_total = 0
            for col in desc_columns:
                if col not in df.columns:
                    continue

                orig_col = f"orig_{col}"
                conf_col = f"{col}_norm_confidence"

                # Replace descriptions and record confidence
                df[col] = df[col].apply(
                    lambda v: mapping[str(v).strip()][0]
                    if isinstance(v, str) and str(v).strip() in mapping
                    else v
                )
                df[conf_col] = df[orig_col].apply(
                    lambda v: mapping[str(v).strip()][1]
                    if isinstance(v, str) and str(v).strip() in mapping
                    else None
                )

                changed = (df[col] != df[orig_col]).sum()
                changed_total += changed

            if changed_total:
                logger.info(
                    f"[{document_id}] {table_name}: normalised "
                    f"{changed_total} description values"
                )

    def _match_descriptions_via_llm(
        self,
        descriptions: List[str],
        reference_items: List[str],
        threshold: float,
        document_id: str = "",
    ) -> Dict[str, tuple]:
        """
        Use the LLM to match raw descriptions to reference items.

        Sends all descriptions in a single call with the full
        reference list. Returns a dict mapping original description
        to (normalised_value, confidence) for matches above threshold,
        or (original_value, confidence) for below-threshold matches.

        Uses structured outputs for reliable JSON parsing.
        """
        try:
            from config.settings import AZURE_OPENAI_DEPLOYMENT
        except ImportError:
            AZURE_OPENAI_DEPLOYMENT = "gpt-4o"

        # Build numbered reference list for the prompt
        ref_list = "\n".join(
            f"  {i}: {item}"
            for i, item in enumerate(reference_items)
        )

        # Build numbered descriptions list
        desc_list = "\n".join(
            f"  {i}: {desc}"
            for i, desc in enumerate(descriptions)
        )

        system_prompt = f"""\
You are a product/item name matching specialist for Australian mining \
industry supply documents.

You will be given a list of raw descriptions extracted from invoices or \
timesheets, and a reference list of known standard item names.

For each raw description, find the BEST matching reference item based \
on semantic similarity. Consider:
- The same product may have different abbreviations, casing, or word order
- Minor spelling differences or extra whitespace should still match
- Partial matches are acceptable if the core product is clearly the same
- "AMMONIUM NITRATE PRILLED BUL" should match "AN, including any \
Moranbah Additional Volumes" (AN = Ammonium Nitrate)

Return a JSON array where each element has:
- "index": the description index number
- "match_index": the reference item index number (-1 if no good match)
- "confidence": float 0.0 to 1.0

Use confidence >= {threshold} only when you are confident the description \
refers to the same product/service as the reference item.

REFERENCE ITEMS:
{ref_list}

RAW DESCRIPTIONS:
{desc_list}

Respond ONLY with a JSON array, no other text."""

        try:
            response = self._get_client().chat.completions.create(
                model=AZURE_OPENAI_DEPLOYMENT,
                messages=[
                    {"role": "system", "content": system_prompt},
                ],
                temperature=0.0,
                max_completion_tokens=16384,
            )

            raw = response.choices[0].message.content.strip()
            # Strip markdown fences if present
            if raw.startswith("```"):
                raw = re.sub(
                    r'^```(?:json)?\s*', '', raw
                )
                raw = re.sub(r'\s*```$', '', raw)

            matches = json.loads(raw)

            result: Dict[str, tuple] = {}
            matched_count = 0
            for m in matches:
                idx = m.get("index", -1)
                match_idx = m.get("match_index", -1)
                confidence = m.get("confidence", 0.0)

                if 0 <= idx < len(descriptions):
                    original = descriptions[idx]
                    if (
                        0 <= match_idx < len(reference_items)
                        and confidence >= threshold
                    ):
                        normalised = reference_items[match_idx]
                        result[original] = (normalised, confidence)
                        matched_count += 1
                        logger.debug(
                            f"[{document_id}] '{original}' → "
                            f"'{normalised}' ({confidence:.2f})"
                        )
                    else:
                        # Below threshold — keep original but record confidence
                        result[original] = (original, confidence)

            logger.info(
                f"[{document_id}] Description normalisation: "
                f"{matched_count}/{len(descriptions)} matched "
                f"above threshold {threshold}"
            )
            return result

        except Exception as e:
            logger.error(
                f"[{document_id}] Description normalisation "
                f"LLM call failed: {e}"
            )
            return {}

    @staticmethod
    def _parse_daywork_column(
        col_name: str,
        sec_headers: Dict[str, str],
    ) -> Dict[str, str]:
        """
        Extract work_date and shift_type from a daywork column name.

        Supported formats (observed in real data):
          'Sat 2/11/2024'            - abbrev day + d/m/yyyy
          'Sat_2/11/2024'            - underscore variant
          'Saturday 7/12/24'         - full day + d/m/yy
          'saturday_7_12_24'         - lowercase underscored d_m_yy
          '12/10/2024 - Day Shift Hours' - date first + shift
          'Saturday Hours'           - day only (no date)
          'Sat', 'Mon'               - bare abbreviation
          'weekly_total_9_08_2025'   - prefixed d_mm_yyyy
          'Saturday 8/03/25'         - full day + d/mm/yy

        Returns dict with keys 'work_date' (ISO or '') and
        'shift_type' (e.g. 'Day', 'Night', or '').
        """
        from datetime import datetime as _dt

        result: Dict[str, str] = {"work_date": "", "shift_type": ""}

        text = col_name.strip()

        # --- Extract shift_type if present ---
        shift_match = re.search(
            r"(Day\s*Shift|Night\s*Shift|Dayshift|Nightshift"
            r"|Training|Mob-Demob|Standdown)",
            text, re.IGNORECASE,
        )
        if shift_match:
            raw_shift = shift_match.group(1).strip()
            if "night" in raw_shift.lower():
                result["shift_type"] = "Night"
            elif "day" in raw_shift.lower():
                result["shift_type"] = "Day"
            else:
                result["shift_type"] = raw_shift

        # --- Try to extract a date ---

        # Pattern 1: d/m/yyyy or d/m/yy embedded anywhere
        m = re.search(r"(\d{1,2})[/](\d{1,2})[/](\d{2,4})", text)
        if m:
            d, mo, y = m.group(1), m.group(2), m.group(3)
            if len(y) == 2:
                y = "20" + y
            try:
                dt = _dt(int(y), int(mo), int(d))
                result["work_date"] = dt.strftime("%Y-%m-%d")
                return result
            except (ValueError, OverflowError):
                pass

        # Pattern 2: underscored date  d_mm_yy  or  d_m_yy
        # e.g. 'saturday_7_12_24', 'weekly_total_9_08_2025'
        m = re.search(r"(\d{1,2})_(\d{1,2})_(\d{2,4})", text)
        if m:
            d, mo, y = m.group(1), m.group(2), m.group(3)
            if len(y) == 2:
                y = "20" + y
            try:
                dt = _dt(int(y), int(mo), int(d))
                result["work_date"] = dt.strftime("%Y-%m-%d")
                return result
            except (ValueError, OverflowError):
                pass

        # Pattern 3: bare day name with no date — derive from
        # period_end (week_ending) in section headers
        _DAY_INDEX = {
            "mon": 0, "monday": 0,
            "tue": 1, "tues": 1, "tuesday": 1,
            "wed": 2, "wednesday": 2,
            "thu": 3, "thur": 3, "thurs": 3, "thursday": 3,
            "fri": 4, "friday": 4,
            "sat": 5, "saturday": 5,
            "sun": 6, "sunday": 6,
        }
        # Strip trailing non-alpha (e.g. "Saturday Hours" -> "Saturday")
        bare = re.sub(r"[\s_]*(hours|shift).*", "", text, flags=re.IGNORECASE).strip()
        day_key = bare.lower().rstrip("_")
        if day_key in _DAY_INDEX:
            # Try to derive date from period_end / week_ending
            week_end_str = (
                sec_headers.get("week_ending", "")
                or sec_headers.get("period_end", "")
            )
            if week_end_str:
                week_end_str = week_end_str.strip()
                we_dt = None
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y",
                            "%d-%m-%Y", "%d.%m.%Y"):
                    try:
                        we_dt = _dt.strptime(week_end_str, fmt)
                        break
                    except ValueError:
                        continue
                if we_dt:
                    # week_ending is typically Fr or the last day of
                    # the pay week.  Compute offset from that day.
                    we_dow = we_dt.weekday()  # Mon=0 … Sun=6
                    target_dow = _DAY_INDEX[day_key]
                    from datetime import timedelta
                    # Go backwards from week-end day to find the target
                    # day within the same 7-day window
                    delta = (we_dow - target_dow) % 7
                    work_dt = we_dt - timedelta(days=delta)
                    result["work_date"] = work_dt.strftime("%Y-%m-%d")

        return result

    @staticmethod
    def _looks_like_header_row(row: list, headers: list) -> bool:
        """
        True if a row's values match its column headers.

        Used to detect and skip stray header rows that upstream
        extractors sometimes leave in data[0]. Comparison is
        case-insensitive and whitespace-stripped. Returns True when
        at least 50% of non-empty cells match their column header.
        """
        if not row or not headers:
            return False
        matches = 0
        non_empty = 0
        for ci, cell in enumerate(row):
            if ci >= len(headers):
                break
            cell_str = str(cell).strip() if cell is not None else ""
            header_str = str(headers[ci]).strip()
            if not cell_str:
                continue
            non_empty += 1
            if cell_str.lower() == header_str.lower():
                matches += 1
        return non_empty > 0 and (matches / non_empty) >= 0.5

    @staticmethod
    def _dedupe_invoice_rows(
        rows: list, document_id: str = "",
    ) -> list:
        """
        Remove duplicate TblInvoice rows within the same document.

        Two rows are duplicates when they share the same normalised
        (description, amount, unit_price, quantity). When duplicates
        exist, prefer the row with a non-empty item_description
        (richer data — usually means a rate-card line was matched).
        """
        if not rows:
            return rows

        def _key(row):
            return (
                str(row.get("description", "")).strip().lower(),
                str(row.get("amount", "")).strip(),
                str(row.get("unit_price", "")).strip(),
                str(row.get("quantity", "")).strip(),
            )

        groups: dict = {}
        key_order: list = []
        for r in rows:
            k = _key(r)
            if k not in groups:
                groups[k] = []
                key_order.append(k)
            groups[k].append(r)

        deduped: list = []
        for k in key_order:
            group = groups[k]
            if len(group) == 1:
                deduped.append(group[0])
                continue
            best = None
            for r in group:
                if str(r.get("item_description", "")).strip():
                    best = r
                    break
            if best is None:
                best = group[0]
            deduped.append(best)
            if document_id:
                logger.info(
                    f"[{document_id}] Deduped {len(group)} invoice "
                    f"rows for key {k}"
                )
        return deduped

    @staticmethod
    def _has_reference_overlap(
        descriptions, reference_items: list,
        min_token_length: int = 4,
    ) -> bool:
        """
        True if any description shares a token (>= min_token_length
        characters) with any reference item.

        Used as a cheap pre-filter before calling the LLM for
        description normalisation. When the sets are disjoint the LLM
        returns all non-matches anyway — save the API round trip.
        """
        def _tokens(s: str) -> set:
            s = re.sub(r'[^a-z0-9\s]', ' ', str(s).lower())
            return {t for t in s.split() if len(t) >= min_token_length}

        ref_tokens: set = set()
        for item in reference_items:
            ref_tokens.update(_tokens(item))
        if not ref_tokens:
            return False

        for desc in descriptions:
            if _tokens(desc) & ref_tokens:
                return True
        return False

    @staticmethod
    def _apply_transform(
        value: str, hint: str, target_field: str = ""
    ) -> str:
        """
        Apply a transform hint to a raw value.

        Handles:
          - parse_currency: strip $, A$, AUD, commas -> clean number
          - parse_date: normalise to ISO 8601 (YYYY-MM-DD)
          - none/expand_daywork: strip thousand separators from numeric
            target fields, pass through everything else
        """
        if not value or not value.strip():
            return ""

        value = value.strip()

        if hint == "parse_currency":
            cleaned = value
            # Remove currency prefixes: A$, $, €, £, ¥
            cleaned = re.sub(r'^[A]?\$', '', cleaned)
            # Remove currency code prefixes: AUD, USD, etc.
            cleaned = re.sub(
                r'^(AUD|USD|EUR|GBP|NZD|CAD|SGD)\s*', '',
                cleaned,
            )
            cleaned = cleaned.replace(' ', '').strip()

            # OCR-safe comma/decimal handling.
            # If there's no dot and the LAST comma has fewer than 3
            # digits after it, the OCR produced e.g. "2,585,00" where
            # the last comma is actually the decimal point. Treat
            # that comma as a decimal; strip earlier commas as
            # thousand separators. Otherwise treat all commas as
            # thousand separators.
            if '.' not in cleaned and ',' in cleaned:
                last_comma = cleaned.rfind(',')
                tail = cleaned[last_comma + 1:]
                tail_digits = re.sub(r'[^\d]', '', tail)
                if 0 < len(tail_digits) < 3:
                    cleaned = (
                        cleaned[:last_comma].replace(',', '')
                        + '.' + tail
                    )
                else:
                    cleaned = cleaned.replace(',', '')
            else:
                cleaned = cleaned.replace(',', '')

            # Keep only digits, dot, minus
            cleaned = re.sub(r'[^\d.\-]', '', cleaned)
            return cleaned if cleaned else value

        elif hint == "parse_date":
            from datetime import datetime as _dt

            # Strip day-name and W/E prefixes
            cleaned = re.sub(
                r'^(?:W/E|w/e|Mon|Tue|Wed|Thu|Fri|Sat|Sun)'
                r'\s*[-]?\s*',
                '', value,
            ).strip()

            # Try common date formats -> normalise to YYYY-MM-DD
            _formats = [
                "%d.%m.%Y",      # 06.11.2024
                "%d/%m/%Y",      # 06/11/2024
                "%d-%m-%Y",      # 06-11-2024
                "%d %B %Y",      # 06 November 2024
                "%d %b %Y",      # 06 Nov 2024
                "%Y-%m-%d",      # 2024-11-06 (already ISO)
                "%Y/%m/%d",      # 2024/11/06
                "%d.%m.%y",      # 06.11.24
                "%d/%m/%y",      # 06/11/24
                "%d-%m-%y",      # 06-11-24
                "%B %d, %Y",     # November 06, 2024
                "%b %d, %Y",     # Nov 06, 2024
                "%m/%d/%Y",      # 11/06/2024 (US format, try last)
            ]

            for fmt in _formats:
                try:
                    dt = _dt.strptime(cleaned, fmt)
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue

            # No format matched — return cleaned with separators
            # normalised to dashes
            cleaned = cleaned.replace('.', '-').replace('/', '-')
            return cleaned if cleaned else value

        # For "none" and "expand_daywork": if the target field is
        # numeric, strip thousand separators (commas)
        _NUMERIC_TARGETS = {
            "quantity", "unit_price", "amount", "net_amount", "gst",
            "gross_amount", "rate", "charge", "discount_rebate",
            "price_units",
        }
        if target_field in _NUMERIC_TARGETS:
            cleaned = value.replace(',', '')
            try:
                float(cleaned)
                return cleaned
            except ValueError:
                return value

        return value

    @staticmethod
    def _extract_header_values(document: Dict) -> Dict[str, str]:
        """
        Extract header key-value pairs from ONLY document-level sources.

        Specifically: document_header and extraction_info. Section-level
        headers are NOT included here — they should only be used within
        their own section (via raw_table["_section_headers"]), to
        prevent values from one section leaking into tables of another.

        Returns a flat dict of {key: value_string}.
        """
        values: Dict[str, str] = {}

        def _add_from_dict(d: Dict) -> None:
            if not isinstance(d, dict):
                return
            for k, v in d.items():
                if k in ("sections", "tables", "failed_tables",
                         "detection_summary"):
                    continue
                if k in values:
                    continue
                if isinstance(v, dict):
                    text = v.get("text", "")
                    if text:
                        values[k] = str(text)
                elif isinstance(v, str) and v.strip():
                    values[k] = v.strip()
                elif v is not None:
                    values[k] = str(v)

        _add_from_dict(document.get("document_header", {}))
        _add_from_dict(document.get("extraction_info", {}))
        return values

    # ------------------------------------------------------------------
    # Excel writer
    # ------------------------------------------------------------------

    @staticmethod
    def write_excel(
        target_dfs: Dict[str, "pd.DataFrame"],
        output_path: Path,
        document_id: str = "",
    ) -> Path:
        """
        Write materialised target DataFrames to a formatted Excel file.

        One sheet per target table (TblInvoice, TblTimesheets).
        """
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side,
        )
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        first_sheet = True

        thin = Side(style="thin", color="FFCCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
        hdr_fill = PatternFill("solid", fgColor="FF2B5797")
        hdr_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        body_font = Font(name="Arial", size=10)
        wrap = Alignment(vertical="top", wrap_text=True)
        alt_fill = PatternFill("solid", fgColor="FFF5F7FA")

        for table_name, df in target_dfs.items():
            if df.empty:
                continue

            if first_sheet:
                ws = wb.active
                ws.title = table_name
                first_sheet = False
            else:
                ws = wb.create_sheet(table_name)

            # Header row
            for ci, col in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = hdr_align
                cell.border = border

            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 28

            # Data rows
            for ri, (_, row) in enumerate(df.iterrows(), 2):
                for ci, col in enumerate(df.columns, 1):
                    val = row[col]
                    if val is None or (isinstance(val, float) and str(val) == "nan"):
                        val = ""
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font = body_font
                    cell.alignment = wrap
                    cell.border = border

                if ri % 2 == 0:
                    for ci in range(1, len(df.columns) + 1):
                        ws.cell(row=ri, column=ci).fill = alt_fill

            # Auto-width
            for ci, col in enumerate(df.columns, 1):
                max_len = len(str(col))
                for ri in range(2, min(len(df) + 2, 50)):
                    cell_val = str(
                        ws.cell(row=ri, column=ci).value or ""
                    )
                    max_len = max(max_len, len(cell_val))
                ws.column_dimensions[get_column_letter(ci)].width = (
                    min(max_len + 4, 45)
                )

        if first_sheet:
            # No sheets were created — add an empty one
            ws = wb.active
            ws.title = "No Data"
            ws.cell(row=1, column=1, value="No tables were mapped.")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path

    # ------------------------------------------------------------------
    # LLM call — structured outputs only
    # ------------------------------------------------------------------

    def _map_table(
        self,
        client,
        table_columns: List[str],
        sample_rows: List[Dict[str, str]],
        header_keys: List[str],
        table_id: str,
        section_type: str,
        document_id: str,
    ) -> Optional[MappingProposal]:
        """
        Call Azure OpenAI structured outputs to classify and map a table.

        Uses client.beta.chat.completions.parse(response_format=MappingProposal)
        which guarantees schema-valid output directly as a Pydantic model.
        """
        from config.settings import (
            AZURE_OPENAI_DEPLOYMENT,
            MAX_RETRIES,
            RETRY_DELAY,
        )
        import time

        user_prompt = self._build_user_prompt(
            table_columns, sample_rows, header_keys,
            table_id, section_type,
        )

        last_error = None
        for attempt in range(MAX_RETRIES):
            try:
                response = client.beta.chat.completions.parse(
                    model=AZURE_OPENAI_DEPLOYMENT,
                    messages=[
                        {"role": "system", "content": _SYSTEM_PROMPT},
                        {"role": "user", "content": user_prompt},
                    ],
                    response_format=MappingProposal,
                    temperature=0.0,
                    max_completion_tokens=16384,
                )

                proposal: MappingProposal = response.choices[0].message.parsed

                if proposal is not None:
                    return proposal

                logger.warning(
                    f"[{document_id}] Structured output returned None "
                    f"for {table_id}, attempt {attempt + 1}"
                )

            except Exception as e:
                last_error = e
                logger.warning(
                    f"[{document_id}] Mapping attempt "
                    f"{attempt + 1}/{MAX_RETRIES} failed for "
                    f"{table_id}: {e}"
                )
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY * (attempt + 1))

        logger.error(
            f"[{document_id}] All {MAX_RETRIES} attempts failed for "
            f"{table_id}: {last_error}"
        )
        return None

    # ------------------------------------------------------------------
    # Sample rows
    # ------------------------------------------------------------------

    @staticmethod
    def _build_sample_rows(
        headers: List, rows: List,
    ) -> List[Dict[str, str]]:
        """
        Build sample rows as {column: value} dicts.

        Multiple rows let the LLM see data patterns — monetary values
        suggest invoice, worker names + hours suggest timesheet,
        subtotal/total rows suggest skip.
        """
        samples = []
        if not headers or not rows:
            return samples

        for row in rows[:_MAX_SAMPLE_ROWS]:
            if not isinstance(row, list):
                continue
            sample = {}
            for ci, h in enumerate(headers):
                if ci < len(row):
                    s = str(row[ci])
                    sample[str(h)] = (
                        s[:120] + "..." if len(s) > 120 else s
                    )
            if sample:
                samples.append(sample)

        return samples

    # ------------------------------------------------------------------
    # Prompt builder
    # ------------------------------------------------------------------

    def _build_user_prompt(
        self,
        table_columns: List[str],
        sample_rows: List[Dict[str, str]],
        header_keys: List[str],
        table_id: str,
        section_type: str,
    ) -> str:
        """Build the user prompt with readable table format, field
        descriptions, and all context needed for accurate mapping."""

        # Format sample data as a readable table
        sample_text = self._format_sample_table(table_columns, sample_rows)

        # Format target schemas with descriptions
        inv_hdr = self._format_targets(INVOICE_HEADER_TARGETS)
        inv_line = self._format_targets(INVOICE_LINE_ITEM_TARGETS)
        ts_hdr = self._format_targets(TIMESHEET_HEADER_TARGETS)
        ts_entry = self._format_targets(TIMESHEET_ENTRY_TARGETS)

        # Format header keys with explanation
        if header_keys:
            hk_text = (
                "These are metadata fields from the document header "
                "(NOT table columns). Map them if they correspond to a "
                "target field.\n"
                + "\n".join(f"  - {k}" for k in header_keys)
            )
        else:
            hk_text = "  (none)"

        return (
            f"## Table to classify and map\n\n"
            f"- **table_id**: {table_id}\n"
            f"- **section_type**: {section_type}\n"
            f"- **columns**: {len(table_columns)}\n"
            f"- **data rows**: {len(sample_rows)} shown below\n\n"
            f"### Section header keys\n"
            f"{hk_text}\n\n"
            f"### Table columns\n"
            f"{table_columns}\n\n"
            f"### Sample data\n"
            f"{sample_text}\n\n"
            f"---\n\n"
            f"## Target schemas\n\n"
            f"**If TblInvoice — Header fields (one value per invoice):**\n"
            f"{inv_hdr}\n\n"
            f"**If TblInvoice — Line item fields (one value per row):**\n"
            f"{inv_line}\n\n"
            f"**If TblTimesheets — Header fields (one value per timesheet):**\n"
            f"{ts_hdr}\n\n"
            f"**If TblTimesheets — Entry fields (one value per row):**\n"
            f"{ts_entry}\n\n"
            f"Produce a MappingProposal.\n"
        )

    @staticmethod
    def _format_sample_table(
        columns: List[str], sample_rows: List[Dict[str, str]]
    ) -> str:
        """Format sample rows as a readable ASCII table so the LLM
        can clearly see column-value alignment."""
        if not sample_rows or not columns:
            return "(no sample data available)"

        # Compute column widths
        col_strs = [str(c) for c in columns]
        widths = [len(c) for c in col_strs]
        for row in sample_rows:
            for ci, col in enumerate(col_strs):
                val = str(row.get(col, ""))
                if len(val) > 40:
                    val = val[:37] + "..."
                if ci < len(widths):
                    widths[ci] = max(widths[ci], len(val))

        # Cap widths
        widths = [min(w, 40) for w in widths]

        # Build header
        header = " | ".join(
            c.ljust(widths[i]) for i, c in enumerate(col_strs)
        )
        separator = "-+-".join("-" * widths[i] for i in range(len(col_strs)))

        # Build rows
        row_lines = []
        for row in sample_rows:
            cells = []
            for ci, col in enumerate(col_strs):
                val = str(row.get(col, ""))
                if len(val) > 40:
                    val = val[:37] + "..."
                cells.append(val.ljust(widths[ci]))
            row_lines.append(" | ".join(cells))

        return f"  {header}\n  {separator}\n" + "\n".join(
            f"  {line}" for line in row_lines
        )

    @staticmethod
    def _format_targets(targets: List[str]) -> str:
        """Format target fields with their descriptions."""
        lines = []
        for name in targets:
            desc = _TARGET_FIELD_DESCRIPTIONS.get(name, "")
            if desc:
                lines.append(f"  - {name}: {desc}")
            else:
                lines.append(f"  - {name}")
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Table discovery
    # ------------------------------------------------------------------

    def _find_all_tables(self, document: Dict) -> List[Dict]:
        """
        Find all tables in the document with context.

        Handles multiple document formats:

        Format A — pdf_table_extractor output:
          sections[].tables[] with columns[].name, data[].values
          Header in sections[].supply_document_header or invoice_header
          or timesheet_header

        Format B — docuporter pipeline (flat sections):
          sections[].content[] with {type: "table", headers, rows}

        Format C — docuporter pipeline (data wrapper):
          sections[].data.content[] with {type: "table", headers, rows}

        Format D — old bucketed format:
          section_type_key[].content[] with {type: "table", headers, rows}
        """
        tables = []

        sections = document.get("sections", [])
        if not isinstance(sections, list):
            sections = []

        # Extract document-level header keys (shared across all sections)
        doc_level_keys = self._extract_doc_level_header_keys(document)

        for sec in sections:
            if not isinstance(sec, dict):
                continue

            section_name = sec.get(
                "section_name", sec.get("heading", "")
            )
            section_type = sec.get("section_type", "other")

            # ── Collect per-section header values ────────────────────
            sec_headers = {}
            for sk, sv in sec.items():
                if sk.endswith("_header") and isinstance(sv, dict):
                    for k, v in sv.items():
                        if k in sec_headers:
                            continue
                        if k in ("sections", "tables", "failed_tables",
                                 "detection_summary"):
                            continue
                        if isinstance(v, dict):
                            text = v.get("text", "")
                            if text:
                                sec_headers[k] = str(text)
                        elif isinstance(v, str) and v.strip():
                            sec_headers[k] = v.strip()
                        elif v is not None:
                            sec_headers[k] = str(v)

            # ── Per-section header keys (doc-level + this section) ───
            _seen_hk = set(doc_level_keys)
            header_keys = list(doc_level_keys)
            for k in sec_headers:
                if k not in _seen_hk:
                    header_keys.append(k)
                    _seen_hk.add(k)

            # ── Skip draft invoice sections early (saves LLM calls) ──
            _doc_title = sec_headers.get("document_title", "")
            if _doc_title and "draft" in _doc_title.lower():
                logger.info(
                    f"Skipping draft section "
                    f"(document_title={_doc_title!r})"
                )
                continue

            # ── Format A: sections[].tables[] (pdf_table_extractor) ──
            raw_tables = sec.get("tables", [])
            if isinstance(raw_tables, list) and raw_tables:
                for tbl in raw_tables:
                    if not isinstance(tbl, dict):
                        continue
                    converted = self._convert_extractor_table(
                        tbl, section_name, section_type, header_keys
                    )
                    if converted is not None:
                        converted["_section_headers"] = sec_headers
                        # Preserve per-table metadata (notes, page info)
                        tbl_meta = tbl.get("metadata", {})
                        tbl_ext = tbl.get("extraction_info", {})
                        converted["_table_metadata"] = tbl_meta
                        converted["_extraction_info"] = tbl_ext
                        tables.append(converted)

            # ── Format B/C: content array with typed blocks ──────────
            content = sec.get("content", [])
            if not content:
                data = sec.get("data", {})
                if isinstance(data, dict):
                    content = data.get("content", [])

            if isinstance(content, list) and content:
                pre_count = len(tables)
                self._collect_tables(
                    content, section_name, section_type,
                    header_keys, tables,
                )
                # Attach section headers to newly added tables
                for t in tables[pre_count:]:
                    t["_section_headers"] = sec_headers

        if tables:
            return tables

        # ── Format D: old bucketed format ────────────────────────────
        skip_keys = {"document_id", "document_header", "sections",
                     "extraction_info"}
        for key, value in document.items():
            if key in skip_keys:
                continue

            if isinstance(value, list):
                for item in value:
                    if not isinstance(item, dict):
                        continue
                    name = item.get(
                        "heading", item.get("section_name", key)
                    )
                    stype = item.get("section_type", key)

                    # Build section-level headers for Format D items
                    d_sec_headers: dict[str, str] = {}
                    for hdr_key in header_keys:
                        hdr_block = item.get(hdr_key, {})
                        if isinstance(hdr_block, dict):
                            for k, v in hdr_block.items():
                                if k not in d_sec_headers and v not in (
                                    None, "", "N/A",
                                ):
                                    d_sec_headers[k] = str(v)

                    # Check for tables[] array
                    raw_tables = item.get("tables", [])
                    if isinstance(raw_tables, list):
                        for tbl in raw_tables:
                            if isinstance(tbl, dict):
                                converted = self._convert_extractor_table(
                                    tbl, name, stype, header_keys
                                )
                                if converted is not None:
                                    converted["_section_headers"] = d_sec_headers
                                    tables.append(converted)

                    # Check for content array
                    content = item.get("content", [])
                    if not content:
                        data = item.get("data", {})
                        if isinstance(data, dict):
                            content = data.get("content", [])
                    if isinstance(content, list) and content:
                        pre_count = len(tables)
                        self._collect_tables(
                            content, name, stype,
                            header_keys, tables,
                        )
                        for t in tables[pre_count:]:
                            t["_section_headers"] = d_sec_headers

        return tables

    def _convert_extractor_table(
        self,
        tbl: Dict,
        section_name: str,
        section_type: str,
        header_keys: List[str],
    ) -> Optional[Dict]:
        """
        Convert a pdf_table_extractor table dict into the standardised
        format used by the mapper.

        Extractor format:
          {
            "table_id": "item_material_pricing",
            "columns": [{"name": "Item No.", "data_type": "text", ...}],
            "data": [{"values": {"Item No.": "010", ...}}],
            ...
          }

        Mapper format:
          {
            "table_id": "...",
            "headers": ["Item No.", ...],
            "rows": [["010", ...], ...],
            ...
          }
        """
        table_id = tbl.get("table_id", "")
        columns = tbl.get("columns", [])
        data_rows = tbl.get("data", [])
        title = tbl.get("title", "")

        # Skip tables explicitly marked as draft in their title.
        # e.g. title="INVOICE DRAFT - (no title)" — these are the
        # pipeline's preview of what will be invoiced, NOT the real
        # invoice, and including them creates duplicate/wrong rows.
        if title and "draft" in title.lower():
            logger.info(
                f"Skipping draft table '{table_id}' in "
                f"'{section_name}' (title={title!r})"
            )
            return None

        # Extract column names
        if isinstance(columns, list) and columns:
            headers = []
            for col in columns:
                if isinstance(col, dict):
                    headers.append(str(col.get("name", "")))
                else:
                    headers.append(str(col))
        else:
            headers = []

        # Convert data rows from {values: {col: val}} to [[val, val, ...]]
        rows = []
        if isinstance(data_rows, list):
            for dr in data_rows:
                if not isinstance(dr, dict):
                    continue
                # Skip non-data rows (total, subtotal, header, etc.).
                # Only rows with row_type='data' represent real billable
                # line items; totals and subtotals are summary aggregates
                # the upstream extractor already broke out for us.
                # If row_type is missing entirely, keep the row (legacy
                # extractor format).
                _row_type = dr.get("row_type")
                if _row_type is not None and _row_type != "data":
                    continue
                values = dr.get("values", {})
                if isinstance(values, dict) and headers:
                    row = [str(values.get(h, "")) for h in headers]
                    rows.append(row)
                elif isinstance(values, list):
                    rows.append([str(v) for v in values])

        if not headers and not rows:
            return None

        if not table_id:
            table_id = title if title else f"{section_name}_table"

        table_id = re.sub(r'[^\w\-]', '_', table_id)[:60]

        return {
            "table_id": table_id,
            "section_name": section_name,
            "section_type": section_type,
            "headers": headers,
            "rows": rows,
            "caption": title,
            "header_keys": header_keys,
        }

    def _collect_tables(
        self,
        content: List,
        section_name: str,
        section_type: str,
        header_keys: List[str],
        tables: List[Dict],
    ) -> None:
        """
        Recursively collect table blocks from a content array
        (docuporter pipeline format).

        Handles:
          - type: "table" blocks (headers + rows)
          - type: "subsection" blocks (recurse into content)
          - Untyped blocks with a content array (recurse)
        """
        if not isinstance(content, list):
            return

        for block in content:
            if not isinstance(block, dict):
                continue

            btype = block.get("type", "")

            if btype == "table":
                headers = block.get("headers", [])
                rows = block.get("rows", [])
                caption = block.get("caption", "")

                # Skip tables explicitly marked as draft in their caption.
                if caption and "draft" in caption.lower():
                    logger.info(
                        f"Skipping draft table in '{section_name}' "
                        f"(caption={caption!r})"
                    )
                    continue

                table_id = caption if caption else f"{section_name}_table"
                table_id = re.sub(r'[^\w\-]', '_', table_id)[:60]

                tables.append({
                    "table_id": table_id,
                    "section_name": section_name,
                    "section_type": section_type,
                    "headers": [str(h) for h in headers],
                    "rows": rows,
                    "caption": caption,
                    "header_keys": header_keys,
                })

            elif btype == "subsection":
                sub_name = block.get("heading", section_name)
                inner = block.get("content", [])
                self._collect_tables(
                    inner, sub_name, section_type,
                    header_keys, tables,
                )

            else:
                # Untyped block — check if it wraps content with tables
                inner = block.get("content", [])
                if isinstance(inner, list) and inner:
                    sub_name = block.get("heading", section_name)
                    self._collect_tables(
                        inner, sub_name, section_type,
                        header_keys, tables,
                    )

    @staticmethod
    def _extract_doc_level_header_keys(document: Dict) -> List[str]:
        """Extract header keys from document-level headers only
        (document_header, extraction_info). Section-level headers
        are handled per-section in _find_all_tables()."""
        keys: List[str] = []
        seen: set = set()

        def _add(d: Dict) -> None:
            if not isinstance(d, dict):
                return
            for k, v in d.items():
                if k in ("sections", "tables", "failed_tables",
                         "detection_summary", "extraction_info"):
                    continue
                if k in seen:
                    continue
                if isinstance(v, dict):
                    if v.get("text"):
                        keys.append(k)
                        seen.add(k)
                elif isinstance(v, str) and v.strip():
                    keys.append(k)
                    seen.add(k)

        _add(document.get("document_header", {}))
        _add(document.get("extraction_info", {}))
        return keys

    @staticmethod
    def _extract_header_keys(document: Dict) -> List[str]:
        """
        Extract header keys from all possible header locations.

        Searches:
          - document_header (docuporter pipeline)
          - extraction_info (pdf_table_extractor)
          - sections[].supply_document_header
          - sections[].invoice_header
          - sections[].timesheet_header
          - sections[].*_header (any key ending in _header)
        """
        keys = []
        seen = set()

        def _add_keys_from_dict(d: Dict) -> None:
            if not isinstance(d, dict):
                return
            for k, v in d.items():
                if k in ("sections", "tables", "failed_tables",
                         "detection_summary", "extraction_info"):
                    continue
                if k in seen:
                    continue
                if isinstance(v, dict):
                    if v.get("text"):
                        keys.append(k)
                        seen.add(k)
                elif isinstance(v, str) and v.strip():
                    keys.append(k)
                    seen.add(k)

        # Top-level document_header
        doc_header = document.get("document_header", {})
        _add_keys_from_dict(doc_header)

        # Top-level extraction_info
        ext_info = document.get("extraction_info", {})
        _add_keys_from_dict(ext_info)

        # Section-level headers (supply_document_header, invoice_header, etc.)
        sections = document.get("sections", [])
        if isinstance(sections, list):
            for sec in sections:
                if not isinstance(sec, dict):
                    continue
                for sec_key, sec_val in sec.items():
                    if (sec_key.endswith("_header")
                            and isinstance(sec_val, dict)):
                        _add_keys_from_dict(sec_val)

        return keys

    # ------------------------------------------------------------------
    # Azure OpenAI client
    # ------------------------------------------------------------------

    def _get_client(self):
        """Get or create the Azure OpenAI client for structured outputs."""
        if self._client is not None:
            return self._client

        try:
            from config.settings import (
                AZURE_OPENAI_ENDPOINT,
                AZURE_OPENAI_API_KEY,
                AZURE_OPENAI_API_VERSION,
                AZURE_OPENAI_TIMEOUT,
            )
            from openai import AzureOpenAI

            if not AZURE_OPENAI_ENDPOINT or not AZURE_OPENAI_API_KEY:
                logger.error(
                    "Azure OpenAI not configured — "
                    "set AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY"
                )
                return None

            self._client = AzureOpenAI(
                azure_endpoint=AZURE_OPENAI_ENDPOINT,
                api_key=AZURE_OPENAI_API_KEY,
                api_version=AZURE_OPENAI_API_VERSION,
                timeout=AZURE_OPENAI_TIMEOUT,
                http_client=httpx.Client(verify=False),
            )
            logger.info("Azure OpenAI client initialised")
            return self._client

        except ImportError:
            logger.error(
                "openai package not installed — pip install openai"
            )
            return None
        except Exception as e:
            logger.error(f"Failed to create Azure OpenAI client: {e}")
            return None