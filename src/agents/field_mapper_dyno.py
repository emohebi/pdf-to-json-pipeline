"""
Field Mapping Agent — classifies extracted tables and maps source fields
to target database columns using LLM structured outputs.

Takes an extracted document JSON, finds all tables, and for each table:
  1. Classifies it as TblInvoice, TblTimesheets, or SKIP
  2. Maps every source field (table columns + section header keys) to the
     most appropriate target database column
  3. Tags each mapping with a transform hint (parse_currency, parse_date, etc.)

Uses Azure OpenAI structured outputs (response_format) to return validated
Pydantic models directly. This is critical for accuracy — structured output
mode constrains the LLM to produce exactly the fields and types defined in
the Pydantic model, avoiding JSON parsing errors entirely.

This agent is standalone — it is NOT part of the main extraction pipeline.
"""
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional
import pandas as pd
import httpx
from enum import Enum
from pydantic import BaseModel, Field

from src.utils import setup_logger, StorageManager

logger = setup_logger("field_mapper")


# =====================================================================
# Pydantic models for structured LLM output
# =====================================================================

class TransformHint(str, Enum):
    """Hint for the normalisation pipeline on how to treat a mapped field."""
    NONE = "none"
    EXPAND_DAYWORK = "expand_daywork"
    PARSE_CURRENCY = "parse_currency"
    PARSE_DATE = "parse_date"


class FieldMapping(BaseModel):
    """A single source -> target field mapping produced by the LLM."""
    source: str = Field(
        description="Source column name or header key exactly as it appears in the extracted JSON."
    )
    target: str = Field(
        description="Target database column name (e.g. 'vendor_name', 'quantity') or 'UNMAPPED' if no match."
    )
    confidence: float = Field(
        ge=0.0, le=1.0,
        description="Confidence score 0.0-1.0."
    )
    transform_hint: TransformHint = Field(
        description="Processing hint for the normalisation pipeline."
    )
    reasoning: str = Field(
        description="Brief explanation of why this mapping was chosen."
    )


class MappingProposal(BaseModel):
    """Complete mapping proposal for one table — returned by the LLM as structured output."""
    target_table: Literal["TblInvoice", "TblTimesheets", "SKIP"] = Field(
        description=(
            "Classification: TblInvoice for invoice data (including supply/delivery documents), "
            "TblTimesheets for timesheet/labour data, "
            "SKIP for tables that should be ignored "
            "(totals_summary, how_to_pay, remittance, tax_summary, signatures)."
        )
    )
    classification_reasoning: str = Field(
        description="Why this table was classified as invoice, timesheet, or skip."
    )
    mappings: list[FieldMapping] = Field(
        description="One mapping per source field (columns + header keys)."
    )
    unmapped_target_fields: list[str] = Field(
        description="Target schema fields that no source field was mapped to (will be NULL in DB)."
    )


# =====================================================================
# Target field schemas — mirror the DDL columns exactly
# =====================================================================

INVOICE_HEADER_TARGETS: List[str] = [
    "vendor_name",
    "invoice_number",
    "invoice_date",
    "due_date",
    "currency",
    "net_amount",
    "gst",
    "gross_amount",
    "description",
    "pricing_type",
    "primary_invoice_flag",
    "discount_rebate",
    "pr_number",
]

INVOICE_LINE_ITEM_TARGETS: List[str] = [
    "description",
    "quantity",
    "uom",
    "unit_price",
    "amount",
    "item_description",
]

TIMESHEET_HEADER_TARGETS: List[str] = [
    "vendor_name",
    "period_start",
    "period_end",
]

TIMESHEET_ENTRY_TARGETS: List[str] = [
    "work_date",
    "staff_equipment_name",
    "quantity",
    "uom",
    "shift_type",
    "rate",
    "charge",
    "position_make",
    "location",
    "approved_date",
    "unique_identifier",
    "item_description",
]

# Descriptions for each target field — included in the prompt so the LLM
# knows what kind of values each column expects and common source-column
# aliases that map to it.
_TARGET_FIELD_DESCRIPTIONS = {
    # Invoice header
    "vendor_name": (
        "Supplier/vendor company name. "
        "Common source names: Supplier, Vendor, Company, From, Billed By, Contractor"
    ),
    "invoice_number": (
        "Invoice or document reference number. "
        "Common source names: Invoice No, Inv #, Invoice Ref, Document No, Tax Invoice Number"
    ),
    "invoice_date": (
        "Date the invoice was issued OR the service period start date. "
        "For SES (Service Entry Sheet) documents, the service period start date "
        "(period_from) is the invoice date. "
        "Common source names: Date, Invoice Date, Tax Date, Document Date, Issued, "
        "Period From, Period Start, Service Date, Service Period From"
    ),
    "due_date": (
        "Payment due date OR the service period end date. "
        "For SES (Service Entry Sheet) documents, the service period end date "
        "(period_to) is the due date. "
        "Common source names: Due Date, Payment Due, Pay By, Terms Date, "
        "Period To, Period End, Service Period To"
    ),
    "currency": (
        "ISO 4217 currency code ONLY (AUD, USD, GBP, NZD). "
        "Must be a 3-letter code, NOT descriptive text like 'Australian Dollars'. "
        "Common source names: Currency, Ccy, Cur"
    ),
    "net_amount": (
        "Invoice total BEFORE tax (ex-GST). This is a HEADER-LEVEL total, not a line item. "
        "Common source names: Subtotal, Net, Total Ex GST, Amount Ex Tax, Net Amount"
    ),
    "gst": (
        "GST or tax amount on the invoice. "
        "Common source names: GST, Tax, VAT, Tax Amount, GST Amount"
    ),
    "gross_amount": (
        "Invoice total INCLUDING tax (inc-GST). This is a HEADER-LEVEL total, not a line item. "
        "Common source names: Total, Total Inc GST, Gross, Amount Inc Tax, Invoice Total, Grand Total, Total Due"
    ),
    "description": (
        "General invoice description or purpose at the HEADER level. "
        "NOT for individual line item descriptions — use line-level 'description' for those. "
        "Common source names: Description, Particulars, Re, Subject, Regarding, Scope"
    ),
    "pricing_type": (
        "A CATEGORY label: LABOUR, MATERIALS, EQUIPMENT, SUBCONTRACTOR, or CONSUMABLES. "
        "Do NOT map worker names, role titles, or item descriptions to this field. "
        "Common source names: Type, Category, Cost Type, Pricing Category"
    ),
    "primary_invoice_flag": (
        "Whether this is the primary invoice (Y/N). Rarely present in source data."
    ),
    "discount_rebate": (
        "Any discount or rebate amount. "
        "Common source names: Discount, Rebate, Adjustment, Credit"
    ),
    "pr_number": (
        "Purchase requisition number. "
        "Common source names: PR, PR No, Requisition, PR Number"
    ),
    # Invoice line items
    "quantity": (
        "Quantity of items, hours, or units for a LINE ITEM. "
        "Common source names: Qty, Quantity, No., Units, Hrs, Hours, Vol, Volume, Count"
    ),
    "uom": (
        "Unit of measure for a line item. "
        "Common source names: UOM, Unit, Per, Basis, Measure, U/M, Each"
    ),
    "unit_price": (
        "Price PER SINGLE UNIT. This is the rate, NOT the line total. "
        "Common source names: Rate, Price, Unit Price, Unit Rate, $/Unit, Unit Cost, Price Each"
    ),
    "amount": (
        "LINE ITEM total = quantity x unit_price. This is a per-row total, NOT the invoice total. "
        "Common source names: Amount, Total, Ext., Extension, Value, Line Total, Amt, Net"
    ),
    "item_description": (
        "Pricing schedule item CODE or reference (e.g. '10.10 Supervisor Mech,DS', 'Item 3.2.1'). "
        "This is an identifier that links back to a contract rate card, NOT free-text description. "
        "Common source names: Item, Item Code, Schedule Item, Ref, Code, Line Ref, Rate Code"
    ),
    # Timesheet header
    "period_start": (
        "Timesheet period start date. "
        "Common source names: Period Start, W/E, Week Ending, From, Start Date, Week Commencing"
    ),
    "period_end": (
        "Timesheet period end date. "
        "Common source names: Period End, To, End Date, Week Ending"
    ),
    # Timesheet entries
    "work_date": (
        "Date the work was performed. "
        "Common source names: Date, Work Date, Day, Shift Date"
    ),
    "staff_equipment_name": (
        "Worker's full name or equipment identifier. "
        "Common source names: Name, Employee, Worker, Staff, Personnel, Operator, Equipment"
    ),
    "shift_type": (
        "Day shift (DS) or night shift (NS). "
        "Common source names: Shift, D/N, Shift Type, DS/NS"
    ),
    "rate": (
        "Hourly or daily PAY RATE for a timesheet entry. "
        "Common source names: Rate, Hourly Rate, $/Hr, Pay Rate, Charge Rate, Daily Rate"
    ),
    "charge": (
        "Total charge for a timesheet entry = quantity x rate. "
        "Common source names: Charge, Total, Amount, Cost, Ext, Value"
    ),
    "position_make": (
        "Job role/position title (for labour) or equipment make/model (for plant). "
        "Common source names: Position, Role, Classification, Trade, Title, Make, Model, Equipment Type"
    ),
    "location": (
        "Work site, pit, or location name. "
        "Common source names: Location, Site, Area, Pit, Mine, Project"
    ),
    "approved_date": (
        "Date the timesheet was approved or signed off. "
        "Common source names: Approved, Approved Date, Sign Off, Authorised"
    ),
    "unique_identifier": (
        "Unique row identifier, employee ID, or reference number. "
        "Common source names: ID, Emp ID, Employee No, Ref, Badge, TS No"
    ),
}


# =====================================================================
# System prompt
# =====================================================================

_SYSTEM_PROMPT = """\
You are a data-mapping specialist for Australian mining industry invoices and timesheets.

## YOUR TASK

Given a table extracted from a PDF document, you must:

1. **CLASSIFY** the table as TblInvoice, TblTimesheets, or SKIP.
2. **MAP** every source field to the best target database column, or "UNMAPPED".
3. **TAG** each mapping with a transform_hint.

## CLASSIFICATION RULES

**TblInvoice** — The table contains invoice LINE ITEMS with some combination of:
  descriptions, quantities, unit prices, amounts/totals, UOM.
  This INCLUDES supply/delivery documents, purchase orders, and credit notes.

**TblTimesheets** — The table contains timesheet/labour entries with some combination of:
  worker names, dates, hours, shifts (DS/NS), rates, charges, positions/roles.
  IMPORTANT: section_type "other" can contain timesheets — classify by DATA CONTENT, not section_type.

**SKIP** — The table should be IGNORED. Skip these:
  - Summary/totals tables (e.g. "Total Ex GST / GST / Total Inc GST" with no line items)
  - How-to-pay / remittance / bank details tables
  - Tax summaries (e.g. "Tax Code / Net / Tax" breakdown)
  - Signature blocks / approval tables
  - Tables with fewer than 2 data rows AND no meaningful line items
  - Metadata tables (key-value pairs like "Invoice No: 12345 / Date: 01/01/2025")

## MAPPING RULES

### Source fields
You will receive TWO types of source fields:
  - **Table columns**: Column headers from the extracted table (e.g. "Description", "Qty", "Amount")
  - **Header keys**: Metadata fields from the document/section header (e.g. "title", "date", "organisation").
    These are NOT table columns — they come from the document's metadata block.
    Map them if they correspond to a target field (e.g. header key "date" -> invoice_date).

You must produce ONE mapping entry for every source field (columns + header keys).

### Disambiguation rules (critical for accuracy)

**description vs item_description:**
  - "description" = free-text description of what was supplied/done (e.g. "Labour hire for shutdown week 12")
  - "item_description" = a schedule/rate-card CODE that references a pricing agreement
    (e.g. "10.10 Supervisor Mech,DS", "Item 3.2.1", "SOR-001")
  - If the source column contains free text → map to "description"
  - If the source column contains codes/references → map to "item_description"
  - When in doubt, check sample data: codes are short, alphanumeric, with dots/dashes

**amount vs net_amount vs gross_amount:**
  - "amount" = a LINE-LEVEL total (one per row, = qty x unit_price)
  - "net_amount" = the INVOICE-LEVEL total before tax (appears once, not per row)
  - "gross_amount" = the INVOICE-LEVEL total after tax (appears once, not per row)
  - If the column has different values per row → it's "amount" (line level)
  - If the column has one value for the whole table → it's "net_amount" or "gross_amount"

**document_total / Document Total — STRICT RULE:**
  - If the document shows BOTH a subtotal AND a GST line AND a total (3 separate values),
    then: subtotal → net_amount, GST → gst, total → gross_amount.
  - If the document shows ONLY ONE total with NO separate GST line (common in supply
    documents, delivery documents, and inter-company transfers), then that single total
    is the NET amount (ex-GST). Map it to "net_amount", NOT "gross_amount".
  - "gross_amount" should ONLY be used when you can confirm GST/tax IS included in the value.
  - When in doubt, prefer "net_amount" — most Australian mining supply documents
    show amounts ex-GST.

**unit_price vs rate vs amount:**
  - "unit_price" (invoice) = price per single unit BEFORE multiplication
  - "rate" (timesheet) = hourly/daily charge rate BEFORE multiplication
  - "amount" / "charge" = the RESULT of multiplication (qty x price or hrs x rate)
  - Look at sample data: if Column A x Column B = Column C, then A and B are qty/price, C is amount

**pricing_type — STRICT RULE:**
  Map ONLY if the source contains category labels like LABOUR, MATERIALS, EQUIPMENT.
  Do NOT map worker role names (e.g. "Boilermaker", "Supervisor N/S") to pricing_type.
  Do NOT map item descriptions to pricing_type.

**currency — STRICT RULE:**
  Map ONLY if the source contains ISO 4217 codes (AUD, USD, EUR, GBP).
  Do NOT map disclaimer text like "All amounts shown are in Australian Dollars".
  Do NOT map currency symbols ($, A$) — those are handled by parse_currency transform.

### Fields to NEVER map
Do NOT produce mappings for these — they are extracted by a separate deterministic process:
  - contract_number / contract_no / contract_ref
  - po_number / purchase_order / PO / customer_po / customer_p_o
  - ses_number / SES / service_entry_sheet
  - sap_invoice_reference / sap_ref / sap_invoice

If a source field matches these, map it to "UNMAPPED" with reasoning "handled by reference extractor".

### Fields to NEVER map — additional rules

**sap_invoice_reference / sap_ref:**
  ALWAYS map to UNMAPPED. This is a SAP system reference number — it is NOT an
  invoice number. Even if it looks like it could be an invoice number, it is
  handled separately by the reference extractor. Do NOT map it to invoice_number.

**item_no / Item No. / line number / sequence number:**
  ALWAYS map to UNMAPPED. This is a line SEQUENCE number (e.g. "010", "020", "030")
  that identifies the row position within the table. It is NOT a pricing schedule
  item code. Do NOT map it to item_description.
  - item_description is for schedule CODES like "10.10 Supervisor Mech,DS" or "SOR-001"
  - A line sequence number like "010" or "1" is just a row counter — mark it UNMAPPED.

**quantity_ordered vs quantity_supplied:**
  In supply/delivery documents, there are often TWO quantity columns:
  - "Quantity Ordered" = what was requested (the order amount)
  - "Quantity Supplied" = what was actually delivered
  Map ONLY "Quantity Supplied" (or "Qty Supplied", "Qty Delivered") to "quantity".
  Map "Quantity Ordered" (or "Qty Ordered", "Order Qty") to UNMAPPED — it is
  reference information, not the billable quantity.

**document_title / document_type:**
  ALWAYS map to UNMAPPED. Values like "SUPPLY DOCUMENT", "TAX INVOICE",
  "TIMESHEET" are structural labels that describe the document type, NOT a
  meaningful description of what was supplied or invoiced. Do NOT map to
  "description".

**header keys that duplicate table columns:**
  When a header key provides the same information as a table column (e.g.
  header "amount" = "53,896.68" and table column "Price" = "53,896.68"),
  prefer the TABLE COLUMN mapping. Map the header key to UNMAPPED with
  reasoning "duplicates table column [column_name]".

**price_units / price_unit / pricing_units / price_uom:**
  ALWAYS map to UNMAPPED. These indicate the pricing basis (e.g. "per 1000 KG").
  The pipeline uses these values automatically to normalise unit_price to a
  per-single-unit value during materialisation. Do NOT map to any target field.

### Transform hints
  - **"parse_currency"** — The value contains currency formatting: "$1,234.56", "A$7,573.50", "$-500.00"
  - **"parse_date"** — The value contains a date in any format: "04.01.2026", "25/02/2025", "Sat 29-11-2025", "W/E 10.10.2025"
  - **"expand_daywork"** — The column represents a specific day in a daywork timesheet (Mon, Tue, Wed, Thu, Fri, Sat, Sun).
    Each day column contains hours worked on that day. Map each day column to "quantity" with hint "expand_daywork".
    The pipeline will pivot these into separate rows per day.
  - **"none"** — No special processing needed.

### Confidence calibration
  - **0.95–1.0** — Exact or near-exact name match with matching data type (e.g. "Invoice Date" -> invoice_date)
  - **0.80–0.94** — Strong semantic match, confirmed by sample data (e.g. "Ext." with "$1,234" values -> amount)
  - **0.60–0.79** — Reasonable match but some ambiguity (e.g. "Total" could be amount or gross_amount)
  - **0.40–0.59** — Uncertain, multiple plausible targets (e.g. "Value" with no clear context)
  - **Below 0.40** — Use UNMAPPED instead. Do not force a low-confidence mapping.

## WORKED EXAMPLES

### Example 1: Invoice line items
Columns: ["Item No", "Description", "Qty", "Unit", "Rate", "Amount"]
Sample: {"Item No": "3.1", "Description": "Mobilisation", "Qty": "1", "Unit": "EA", "Rate": "$5,000.00", "Amount": "$5,000.00"}
→ TblInvoice
  Item No → item_description (0.85, none) — schedule reference code like "3.1"
  Description → description (0.95, none) — free-text line description
  Qty → quantity (0.95, none)
  Unit → uom (0.95, none)
  Rate → unit_price (0.90, parse_currency)
  Amount → amount (0.95, parse_currency)

NOTE: "Item No" maps to item_description ONLY because the sample value "3.1" is a
schedule reference code. If the value were "010" or "1" (a line sequence number),
it should be UNMAPPED instead.

### Example 2: Timesheet daywork
Columns: ["Name", "Position", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Total Hrs", "Rate", "Amount"]
Sample: {"Name": "J Smith", "Position": "Boilermaker", "Mon": "10", "Tue": "10", ..., "Rate": "$85.50", "Amount": "$855.00"}
→ TblTimesheets
  Name → staff_equipment_name (0.95, none)
  Position → position_make (0.90, none) — NOT pricing_type
  Mon → quantity (0.90, expand_daywork)
  Tue → quantity (0.90, expand_daywork)
  ... (each day column → quantity with expand_daywork)
  Total Hrs → UNMAPPED — derived field, not stored
  Rate → rate (0.95, parse_currency)
  Amount → charge (0.90, parse_currency)

### Example 3: SKIP table
Columns: ["", "Amount"]
Rows: [["Subtotal Ex GST", "$45,000.00"], ["GST", "$4,500.00"], ["Total Inc GST", "$49,500.00"]]
→ SKIP — summary totals table, not line items

### Example 4: Supply document
Columns: ["Item No.", "Material Description", "Quantity Ordered", "Quantity Supplied", "UOM", "Unit Price", "Price"]
Sample: {"Item No.": "010", "Material Description": "AMMONIUM NITRATE PRILLED BUL", "Quantity Ordered": "5000,000.00", "Quantity Supplied": "59,530.00", "UOM": "KG", "Unit Price": "905.37", "Price": "53,896.68"}
Header keys include: document_title, sap_invoice_reference, vendor_name, doc_number, date, customer_po, amount, price_units, price_uom
→ TblInvoice
  Item No. → UNMAPPED — line sequence number "010", not a schedule code
  Material Description → description (0.95, none) — what was supplied
  Quantity Ordered → UNMAPPED — order qty, not the delivered/billable qty
  Quantity Supplied → quantity (0.95, none) — actual delivered quantity
  UOM → uom (0.95, none)
  Unit Price → unit_price (0.95, parse_currency) — note: pipeline will auto-adjust using price_units
  Price → amount (0.95, parse_currency) — line total
  document_title → UNMAPPED — structural label "SUPPLY DOCUMENT", not a description
  sap_invoice_reference → UNMAPPED — handled by reference extractor
  vendor_name → vendor_name (0.95, none)
  doc_number → invoice_number (0.90, none) — document reference number
  date → invoice_date (0.90, parse_date)
  customer_po → UNMAPPED — handled by reference extractor (PO number)
  amount → UNMAPPED — duplicates table column "Price"
  price_units → UNMAPPED — pricing basis handled by pipeline during materialisation
  price_uom → UNMAPPED — pricing basis handled by pipeline during materialisation

### Example 5: SES (Service Entry Sheet) document
Columns: ["Service Item No", "Service Description", "GL Account", "WBS Element/Cost Centre", "Qty", "UOM", "Price/unit (Tax excl.)", "Tax (%)", "Tax Amt", "Cost excl tax"]
Sample: {"Service Item No": "10", "Service Description": "FY25 Q1 AN & ANE", "GL Account": "4200010", "WBS Element/Cost Centre": "I-MN19-0031.01.01", "Qty": "119620.200", "UOM": "KG", "Price/unit (Tax excl.)": "0.91", "Tax (%)": "10.0", "Tax Amt": "10833.36", "Cost excl tax": "108333.56"}
Header keys include: vendor_name, reference_number, po_number, period_from, period_to, currency, net_amount, tax_amount, gross_amount, ses_type, po_short_description, claim_submission_date
→ TblInvoice
  Service Item No → UNMAPPED — line sequence number, not a schedule code
  Service Description → description (0.95, none) — what service was provided
  GL Account → UNMAPPED — internal accounting code
  WBS Element/Cost Centre → UNMAPPED — internal cost centre
  Qty → quantity (0.95, none)
  UOM → uom (0.95, none)
  Price/unit (Tax excl.) → unit_price (0.95, parse_currency)
  Tax (%) → UNMAPPED — percentage, not a monetary amount
  Tax Amt → UNMAPPED — line-level tax, use header tax_amount for gst
  Cost excl tax → amount (0.95, parse_currency) — line total ex-tax
  vendor_name → vendor_name (0.95, none)
  reference_number → invoice_number (0.85, none) — SES reference number
  po_number → UNMAPPED — handled by reference extractor
  period_from → invoice_date (0.95, parse_date) — service period start = invoice date
  period_to → due_date (0.90, parse_date) — service period end = due date
  currency → currency (0.95, none) — already ISO code "AUD"
  net_amount → net_amount (0.95, parse_currency) — header-level total ex-tax
  tax_amount → gst (0.95, parse_currency) — header-level GST
  gross_amount → gross_amount (0.95, parse_currency) — header-level total inc-tax
  ses_type → UNMAPPED — structural metadata ("Create"), not a description
  po_short_description → UNMAPPED — PO reference text, not line description
  claim_submission_date → UNMAPPED — internal submission date, not invoice date
"""

# Number of sample rows to include in the prompt
_MAX_SAMPLE_ROWS = 5


# =====================================================================
# Agent
# =====================================================================

class FieldMappingAgent:
    """
    Classify extracted tables and map their fields to target DB columns.

    Uses Azure OpenAI structured outputs (response_format=MappingProposal)
    exclusively. This is the only path that guarantees schema-valid
    responses — text-based JSON parsing is too lossy for production
    accuracy.
    """

    def __init__(self):
        self.storage = StorageManager()
        self._client = None

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def map_document(
        self,
        document_path: str,
        document_id: str = None,
    ) -> Dict[str, Any]:
        """
        Map all tables in an extracted document JSON.

        Args:
            document_path: Path to the final extracted JSON.
            document_id:   Identifier for logging.

        Returns:
            Report dict with mapping proposals for each table.
        """
        doc_path = Path(document_path)
        if document_id is None:
            document_id = doc_path.stem

        with open(doc_path, encoding="utf-8") as f:
            document = json.load(f)

        logger.info(f"[{document_id}] Starting field mapping")

        tables = self._find_all_tables(document)
        logger.info(f"[{document_id}] Found {len(tables)} table(s)")

        if not tables:
            return {
                "document_id": document_id,
                "tables_found": 0,
                "tables_mapped": 0,
                "proposals": [],
            }

        # Ensure client is ready
        client = self._get_client()
        if client is None:
            logger.error(
                f"[{document_id}] Azure OpenAI client not available. "
                f"Structured outputs require Azure OpenAI with the "
                f"openai SDK."
            )
            return {
                "document_id": document_id,
                "tables_found": len(tables),
                "tables_mapped": 0,
                "proposals": [],
                "error": "Azure OpenAI client not configured",
            }

        proposals: List[Dict[str, Any]] = []
        for i, table_info in enumerate(tables, 1):
            table_id = table_info.get("table_id", f"table_{i}")
            section_name = table_info.get("section_name", "Unknown")
            section_type = table_info.get("section_type", "other")
            headers = table_info.get("headers", [])
            rows = table_info.get("rows", [])
            header_keys = table_info.get("header_keys", [])

            if not headers and not rows:
                logger.info(
                    f"[{document_id}]   Skipping empty table {table_id}"
                )
                continue

            sample_rows = self._build_sample_rows(headers, rows)

            logger.info(
                f"[{document_id}]   Mapping table {i}/{len(tables)}: "
                f"'{table_id}' in '{section_name}' "
                f"({len(headers)} cols, {len(rows)} rows, "
                f"{len(header_keys)} header keys, "
                f"{len(sample_rows)} sample rows)"
            )

            proposal = self._map_table(
                client=client,
                table_columns=[str(h) for h in headers],
                sample_rows=sample_rows,
                header_keys=header_keys,
                table_id=table_id,
                section_type=section_type,
                document_id=document_id,
            )

            if proposal is not None:
                proposal_dict = proposal.model_dump()
                proposal_dict["_source"] = {
                    "table_id": table_id,
                    "section_name": section_name,
                    "section_type": section_type,
                    "column_count": len(headers),
                    "row_count": len(rows),
                }
                proposals.append(proposal_dict)

                logger.info(
                    f"[{document_id}]     -> {proposal.target_table} "
                    f"({len(proposal.mappings)} mappings, "
                    f"{len(proposal.unmapped_target_fields)} unmapped)"
                )
            else:
                logger.warning(
                    f"[{document_id}]     -> FAILED to map {table_id}"
                )

        report = {
            "document_id": document_id,
            "tables_found": len(tables),
            "tables_mapped": len(proposals),
            "proposals": proposals,
        }

        # Save
        try:
            report_path = (
                self.storage.final_dir
                / f"{document_id}_field_mapping.json"
            )
            report_path.parent.mkdir(parents=True, exist_ok=True)
            with open(report_path, "w", encoding="utf-8") as f:
                json.dump(report, f, indent=2, ensure_ascii=False)
            logger.info(f"[{document_id}] Report saved: {report_path}")
        except Exception as e:
            logger.warning(f"[{document_id}] Failed to save report: {e}")

        return report

    # ------------------------------------------------------------------
    # Materialisation — apply mappings to produce target DataFrames
    # ------------------------------------------------------------------

    def materialise_tables(
        self,
        document: Dict,
        report: Dict[str, Any],
        document_id: str = "",
    ) -> Dict[str, "pd.DataFrame"]:
        """
        Apply mapping proposals to the source data and produce target
        DataFrames ready for Excel export.

        Takes the original document (with raw table data) and the
        mapping report (with proposals), and produces one DataFrame
        per target table (TblInvoice, TblTimesheets). SKIP tables
        are excluded.

        Args:
            document:    The original extracted JSON document.
            report:      The mapping report from map_document().
            document_id: Identifier for logging.

        Returns:
            Dict with keys "TblInvoice" and/or "TblTimesheets",
            each containing a pandas DataFrame. Empty dict if no
            tables were mapped.
        """
        import pandas as pd

        proposals = report.get("proposals", [])
        if not proposals:
            return {}

        # Rebuild the raw tables keyed by table_id for lookup
        all_tables = self._find_all_tables(document)
        tables_by_id: Dict[str, Dict] = {}
        for t in all_tables:
            tables_by_id[t["table_id"]] = t

        # Collect header values from all header locations
        header_values = self._extract_header_values(document)

        # Extraction info (contract_number, po_number, etc.)
        extraction_info = document.get("extraction_info", {})

        # Build rows per target table
        target_rows: Dict[str, List[Dict]] = {
            "TblInvoice": [],
            "TblTimesheets": [],
        }

        for proposal in proposals:
            target_table = proposal.get("target_table", "")
            if target_table == "SKIP" or target_table not in target_rows:
                continue

            source = proposal.get("_source", {})
            table_id = source.get("table_id", "")
            raw_table = tables_by_id.get(table_id)

            if raw_table is None:
                logger.warning(
                    f"[{document_id}] Cannot find raw table "
                    f"'{table_id}' for materialisation"
                )
                continue

            mappings = proposal.get("mappings", [])
            headers = raw_table.get("headers", [])
            rows = raw_table.get("rows", [])

            # Build the source->target mapping lookup
            col_mappings: Dict[str, Dict] = {}  # source_col -> {target, hint}
            hdr_mappings: Dict[str, Dict] = {}  # header_key -> {target, hint}

            for m in mappings:
                src = m.get("source", "")
                tgt = m.get("target", "")
                hint = m.get("transform_hint", "none")
                if tgt == "UNMAPPED" or not tgt:
                    continue

                if src in headers:
                    col_mappings[src] = {"target": tgt, "hint": hint}
                else:
                    hdr_mappings[src] = {"target": tgt, "hint": hint}

            # Materialise each data row
            for row in rows:
                if not isinstance(row, list):
                    continue

                output_row: Dict[str, str] = {}

                # Apply column mappings
                for ci, col_name in enumerate(headers):
                    if ci >= len(row):
                        continue
                    mapping = col_mappings.get(col_name)
                    if mapping is None:
                        continue

                    target_col = mapping["target"]
                    hint = mapping["hint"]
                    raw_value = str(row[ci]) if row[ci] is not None else ""

                    output_row[target_col] = self._apply_transform(
                        raw_value, hint, target_col
                    )

                # Apply header key mappings (same value for every row)
                for hdr_key, mapping in hdr_mappings.items():
                    target_col = mapping["target"]
                    hint = mapping["hint"]
                    raw_value = str(header_values.get(hdr_key, ""))

                    # Don't overwrite a column-level value with a header value
                    if target_col not in output_row or not output_row[target_col]:
                        output_row[target_col] = self._apply_transform(
                            raw_value, hint, target_col
                        )

                # Add reference fields from extraction_info
                if extraction_info:
                    for ref_field in ("contract_number", "po_number",
                                      "ses_number", "sap_invoice_reference"):
                        val = extraction_info.get(ref_field, "")
                        if val:
                            output_row[ref_field] = str(val)

                # Add source traceability
                output_row["_source_table"] = table_id
                output_row["_source_file"] = extraction_info.get(
                    "source_file", ""
                )

                target_rows[target_table].append(output_row)

        # Build DataFrames
        result = {}
        for table_name, rows_list in target_rows.items():
            if rows_list:
                df = pd.DataFrame(rows_list)
                # Reorder columns: target schema first, then extras
                schema_cols = (
                    INVOICE_HEADER_TARGETS + INVOICE_LINE_ITEM_TARGETS
                    if table_name == "TblInvoice"
                    else TIMESHEET_HEADER_TARGETS + TIMESHEET_ENTRY_TARGETS
                )
                # Add reference columns
                ref_cols = [
                    "contract_number", "po_number",
                    "ses_number", "sap_invoice_reference",
                ]
                # orig_description comes right after schema cols
                desc_cols = ["orig_description", "description_norm_confidence"]
                trace_cols = ["_source_table", "_source_file"]
                ordered = []
                for c in (
                    schema_cols + ref_cols + desc_cols + trace_cols
                ):
                    if c in df.columns and c not in ordered:
                        ordered.append(c)
                # Add any remaining columns not in the schema
                for c in df.columns:
                    if c not in ordered:
                        ordered.append(c)
                df = df[ordered]

                result[table_name] = df
                logger.info(
                    f"[{document_id}] Materialised {table_name}: "
                    f"{len(df)} rows, {len(df.columns)} columns"
                )

        # Normalise descriptions via LLM
        if result:
            # Always ensure orig_description exists even if normalisation fails
            for df in result.values():
                if "description" in df.columns and "orig_description" not in df.columns:
                    df["orig_description"] = df["description"]
            try:
                self._normalise_descriptions(result, document_id)
            except Exception as e:
                logger.warning(
                    f"[{document_id}] Description normalisation "
                    f"failed (non-fatal): {e}"
                )

        return result

    # ------------------------------------------------------------------
    # Description normalisation via LLM
    # ------------------------------------------------------------------

    def _normalise_descriptions(
        self,
        materialised: Dict[str, "pd.DataFrame"],
        document_id: str = "",
    ) -> None:
        """
        Normalise description values by matching them to a reference
        list of known item names via LLM.

        For each unique description value across all target tables,
        the LLM picks the closest match from the reference list and
        returns a confidence score. If confidence >= threshold the
        description is replaced; otherwise the original is kept.

        The original value is always preserved in 'orig_description'.

        Modifies DataFrames in-place.
        """
        try:
            from config.description_items import (
                DESCRIPTION_ITEMS,
                DESCRIPTION_MATCH_THRESHOLD,
            )
        except ImportError:
            logger.warning(
                f"[{document_id}] config.description_items not found "
                "— skipping description normalisation"
            )
            return

        if not DESCRIPTION_ITEMS:
            return

        # Collect unique non-empty descriptions across all tables
        unique_descs: set = set()
        for df in materialised.values():
            if "description" in df.columns:
                for val in df["description"].dropna().unique():
                    val = str(val).strip()
                    if val:
                        unique_descs.add(val)

        if not unique_descs:
            logger.info(
                f"[{document_id}] No descriptions to normalise"
            )
            return

        logger.info(
            f"[{document_id}] Normalising {len(unique_descs)} "
            f"unique descriptions against {len(DESCRIPTION_ITEMS)} "
            f"reference items"
        )

        # Call LLM to match descriptions
        client = self._get_client()
        if client is None:
            logger.warning(
                f"[{document_id}] No Azure client — "
                "skipping description normalisation"
            )
            return

        mapping = self._match_descriptions_via_llm(
            list(unique_descs),
            DESCRIPTION_ITEMS,
            DESCRIPTION_MATCH_THRESHOLD,
            document_id,
        )

        # Always preserve the original description before any
        # normalisation so orig_description is never empty.
        for df in materialised.values():
            if "description" in df.columns:
                df["orig_description"] = df["description"]

        if not mapping:
            # No mapping results — set confidence to None
            for df in materialised.values():
                if "description" in df.columns:
                    df["description_norm_confidence"] = None
            return

        # Apply to all DataFrames
        for table_name, df in materialised.items():
            if "description" not in df.columns:
                continue

            # Replace descriptions and record confidence
            df["description"] = df["description"].apply(
                lambda v: mapping[str(v).strip()][0]
                if isinstance(v, str) and str(v).strip() in mapping
                else v
            )
            df["description_norm_confidence"] = df["orig_description"].apply(
                lambda v: mapping[str(v).strip()][1]
                if isinstance(v, str) and str(v).strip() in mapping
                else None
            )

            changed = (
                df["description"] != df["orig_description"]
            ).sum()
            logger.info(
                f"[{document_id}] {table_name}: normalised "
                f"{changed}/{len(df)} description values"
            )

    def _match_descriptions_via_llm(
        self,
        descriptions: List[str],
        reference_items: List[str],
        threshold: float,
        document_id: str = "",
    ) -> Dict[str, tuple]:
        """
        Use the LLM to match raw descriptions to reference items.

        Sends all descriptions in a single call with the full
        reference list. Returns a dict mapping original description
        to (normalised_value, confidence) for matches above threshold,
        or (original_value, confidence) for below-threshold matches.

        Uses structured outputs for reliable JSON parsing.
        """
        try:
            from config.settings import AZURE_OPENAI_DEPLOYMENT
        except ImportError:
            AZURE_OPENAI_DEPLOYMENT = "gpt-4o"

        # Build numbered reference list for the prompt
        ref_list = "\n".join(
            f"  {i}: {item}"
            for i, item in enumerate(reference_items)
        )

        # Build numbered descriptions list
        desc_list = "\n".join(
            f"  {i}: {desc}"
            for i, desc in enumerate(descriptions)
        )

        system_prompt = f"""\
You are a product/item name matching specialist for Australian mining \
industry supply documents.

You will be given a list of raw descriptions extracted from invoices or \
timesheets, and a reference list of known standard item names.

For each raw description, find the BEST matching reference item based \
on semantic similarity. Consider:
- The same product may have different abbreviations, casing, or word order
- Minor spelling differences or extra whitespace should still match
- Partial matches are acceptable if the core product is clearly the same
- "AMMONIUM NITRATE PRILLED BUL" should match "AN, including any \
Moranbah Additional Volumes" (AN = Ammonium Nitrate)

Return a JSON array where each element has:
- "index": the description index number
- "match_index": the reference item index number (-1 if no good match)
- "confidence": float 0.0 to 1.0

Use confidence >= {threshold} only when you are confident the description \
refers to the same product/service as the reference item.

REFERENCE ITEMS:
{ref_list}

RAW DESCRIPTIONS:
{desc_list}

Respond ONLY with a JSON array, no other text."""

        try:
            response = self._get_client().chat.completions.create(
                model=AZURE_OPENAI_DEPLOYMENT,
                messages=[
                    {"role": "system", "content": system_prompt},
                ],
                temperature=0.0,
                max_completion_tokens=4096,
            )

            raw = response.choices[0].message.content.strip()
            # Strip markdown fences if present
            if raw.startswith("```"):
                raw = re.sub(
                    r'^```(?:json)?\s*', '', raw
                )
                raw = re.sub(r'\s*```$', '', raw)

            matches = json.loads(raw)

            result: Dict[str, tuple] = {}
            matched_count = 0
            for m in matches:
                idx = m.get("index", -1)
                match_idx = m.get("match_index", -1)
                confidence = m.get("confidence", 0.0)

                if 0 <= idx < len(descriptions):
                    original = descriptions[idx]
                    if (
                        0 <= match_idx < len(reference_items)
                        and confidence >= threshold
                    ):
                        normalised = reference_items[match_idx]
                        result[original] = (normalised, confidence)
                        matched_count += 1
                        logger.debug(
                            f"[{document_id}] '{original}' → "
                            f"'{normalised}' ({confidence:.2f})"
                        )
                    else:
                        # Below threshold — keep original but record confidence
                        result[original] = (original, confidence)

            logger.info(
                f"[{document_id}] Description normalisation: "
                f"{matched_count}/{len(descriptions)} matched "
                f"above threshold {threshold}"
            )
            return result

        except Exception as e:
            logger.error(
                f"[{document_id}] Description normalisation "
                f"LLM call failed: {e}"
            )
            return {}

    @staticmethod
    def _apply_transform(
        value: str, hint: str, target_field: str = ""
    ) -> str:
        """
        Apply a transform hint to a raw value.

        Handles:
          - parse_currency: strip $, A$, AUD, commas -> clean number
          - parse_date: normalise to ISO 8601 (YYYY-MM-DD)
          - none/expand_daywork: strip thousand separators from numeric
            target fields, pass through everything else
        """
        if not value or not value.strip():
            return ""

        value = value.strip()

        if hint == "parse_currency":
            cleaned = value
            # Remove currency prefixes: A$, $, €, £, ¥
            cleaned = re.sub(r'^[A]?\$', '', cleaned)
            # Remove currency code prefixes: AUD, USD, etc.
            cleaned = re.sub(
                r'^(AUD|USD|EUR|GBP|NZD|CAD|SGD)\s*', '',
                cleaned,
            )
            # Remove thousand separators (commas) but keep decimal
            cleaned = cleaned.replace(',', '')
            # Remove spaces
            cleaned = cleaned.replace(' ', '')
            # Keep only digits, dot, minus
            cleaned = re.sub(r'[^\d.\-]', '', cleaned)
            return cleaned if cleaned else value

        elif hint == "parse_date":
            from datetime import datetime as _dt

            # Strip day-name and W/E prefixes
            cleaned = re.sub(
                r'^(?:W/E|w/e|Mon|Tue|Wed|Thu|Fri|Sat|Sun)'
                r'\s*[-]?\s*',
                '', value,
            ).strip()

            # Try common date formats -> normalise to YYYY-MM-DD
            _formats = [
                "%d.%m.%Y",      # 06.11.2024
                "%d/%m/%Y",      # 06/11/2024
                "%d-%m-%Y",      # 06-11-2024
                "%d %B %Y",      # 06 November 2024
                "%d %b %Y",      # 06 Nov 2024
                "%Y-%m-%d",      # 2024-11-06 (already ISO)
                "%Y/%m/%d",      # 2024/11/06
                "%d.%m.%y",      # 06.11.24
                "%d/%m/%y",      # 06/11/24
                "%d-%m-%y",      # 06-11-24
                "%B %d, %Y",     # November 06, 2024
                "%b %d, %Y",     # Nov 06, 2024
                "%m/%d/%Y",      # 11/06/2024 (US format, try last)
            ]

            for fmt in _formats:
                try:
                    dt = _dt.strptime(cleaned, fmt)
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue

            # No format matched — return cleaned with separators
            # normalised to dashes
            cleaned = cleaned.replace('.', '-').replace('/', '-')
            return cleaned if cleaned else value

        # For "none" and "expand_daywork": if the target field is
        # numeric, strip thousand separators (commas)
        _NUMERIC_TARGETS = {
            "quantity", "unit_price", "amount", "net_amount", "gst",
            "gross_amount", "rate", "charge", "discount_rebate",
            "price_units",
        }
        if target_field in _NUMERIC_TARGETS:
            cleaned = value.replace(',', '')
            try:
                float(cleaned)
                return cleaned
            except ValueError:
                return value

        return value

    @staticmethod
    def _extract_header_values(document: Dict) -> Dict[str, str]:
        """
        Extract ALL header key-value pairs from all header locations.
        Returns a flat dict of {key: value_string}.
        """
        values: Dict[str, str] = {}

        def _add_from_dict(d: Dict) -> None:
            if not isinstance(d, dict):
                return
            for k, v in d.items():
                if k in ("sections", "tables", "failed_tables",
                         "detection_summary"):
                    continue
                if k in values:
                    continue
                if isinstance(v, dict):
                    text = v.get("text", "")
                    if text:
                        values[k] = str(text)
                elif isinstance(v, str) and v.strip():
                    values[k] = v.strip()
                elif v is not None:
                    values[k] = str(v)

        # document_header
        _add_from_dict(document.get("document_header", {}))
        # extraction_info
        _add_from_dict(document.get("extraction_info", {}))
        # Section-level headers
        sections = document.get("sections", [])
        if isinstance(sections, list):
            for sec in sections:
                if isinstance(sec, dict):
                    for sk, sv in sec.items():
                        if sk.endswith("_header") and isinstance(sv, dict):
                            _add_from_dict(sv)

        return values

    # ------------------------------------------------------------------
    # Excel writer
    # ------------------------------------------------------------------

    @staticmethod
    def write_excel(
        target_dfs: Dict[str, "pd.DataFrame"],
        output_path: Path,
        document_id: str = "",
    ) -> Path:
        """
        Write materialised target DataFrames to a formatted Excel file.

        One sheet per target table (TblInvoice, TblTimesheets).
        """
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side,
        )
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        first_sheet = True

        thin = Side(style="thin", color="FFCCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
        hdr_fill = PatternFill("solid", fgColor="FF2B5797")
        hdr_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        body_font = Font(name="Arial", size=10)
        wrap = Alignment(vertical="top", wrap_text=True)
        alt_fill = PatternFill("solid", fgColor="FFF5F7FA")

        for table_name, df in target_dfs.items():
            if df.empty:
                continue

            if first_sheet:
                ws = wb.active
                ws.title = table_name
                first_sheet = False
            else:
                ws = wb.create_sheet(table_name)

            # Header row
            for ci, col in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = hdr_align
                cell.border = border

            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 28

            # Data rows
            for ri, (_, row) in enumerate(df.iterrows(), 2):
                for ci, col in enumerate(df.columns, 1):
                    val = row[col]
                    if val is None or (isinstance(val, float) and str(val) == "nan"):
                        val = ""
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font = body_font
                    cell.alignment = wrap
                    cell.border = border

                if ri % 2 == 0:
                    for ci in range(1, len(df.columns) + 1):
                        ws.cell(row=ri, column=ci).fill = alt_fill

            # Auto-width
            for ci, col in enumerate(df.columns, 1):
                max_len = len(str(col))
                for ri in range(2, min(len(df) + 2, 50)):
                    cell_val = str(
                        ws.cell(row=ri, column=ci).value or ""
                    )
                    max_len = max(max_len, len(cell_val))
                ws.column_dimensions[get_column_letter(ci)].width = (
                    min(max_len + 4, 45)
                )

        if first_sheet:
            # No sheets were created — add an empty one
            ws = wb.active
            ws.title = "No Data"
            ws.cell(row=1, column=1, value="No tables were mapped.")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path

    # ------------------------------------------------------------------
    # LLM call — structured outputs only
    # ------------------------------------------------------------------

    def _map_table(
        self,
        client,
        table_columns: List[str],
        sample_rows: List[Dict[str, str]],
        header_keys: List[str],
        table_id: str,
        section_type: str,
        document_id: str,
    ) -> Optional[MappingProposal]:
        """
        Call Azure OpenAI structured outputs to classify and map a table.

        Uses client.beta.chat.completions.parse(response_format=MappingProposal)
        which guarantees schema-valid output directly as a Pydantic model.
        """
        from config.settings import (
            AZURE_OPENAI_DEPLOYMENT,
            MAX_RETRIES,
            RETRY_DELAY,
        )
        import time

        user_prompt = self._build_user_prompt(
            table_columns, sample_rows, header_keys,
            table_id, section_type,
        )

        last_error = None
        for attempt in range(MAX_RETRIES):
            try:
                response = client.beta.chat.completions.parse(
                    model=AZURE_OPENAI_DEPLOYMENT,
                    messages=[
                        {"role": "system", "content": _SYSTEM_PROMPT},
                        {"role": "user", "content": user_prompt},
                    ],
                    response_format=MappingProposal,
                    temperature=0.0,
                    max_completion_tokens=4096,
                )

                proposal: MappingProposal = response.choices[0].message.parsed

                if proposal is not None:
                    return proposal

                logger.warning(
                    f"[{document_id}] Structured output returned None "
                    f"for {table_id}, attempt {attempt + 1}"
                )

            except Exception as e:
                last_error = e
                logger.warning(
                    f"[{document_id}] Mapping attempt "
                    f"{attempt + 1}/{MAX_RETRIES} failed for "
                    f"{table_id}: {e}"
                )
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY * (attempt + 1))

        logger.error(
            f"[{document_id}] All {MAX_RETRIES} attempts failed for "
            f"{table_id}: {last_error}"
        )
        return None

    # ------------------------------------------------------------------
    # Sample rows
    # ------------------------------------------------------------------

    @staticmethod
    def _build_sample_rows(
        headers: List, rows: List,
    ) -> List[Dict[str, str]]:
        """
        Build sample rows as {column: value} dicts.

        Multiple rows let the LLM see data patterns — monetary values
        suggest invoice, worker names + hours suggest timesheet,
        subtotal/total rows suggest skip.
        """
        samples = []
        if not headers or not rows:
            return samples

        for row in rows[:_MAX_SAMPLE_ROWS]:
            if not isinstance(row, list):
                continue
            sample = {}
            for ci, h in enumerate(headers):
                if ci < len(row):
                    s = str(row[ci])
                    sample[str(h)] = (
                        s[:120] + "..." if len(s) > 120 else s
                    )
            if sample:
                samples.append(sample)

        return samples

    # ------------------------------------------------------------------
    # Prompt builder
    # ------------------------------------------------------------------

    def _build_user_prompt(
        self,
        table_columns: List[str],
        sample_rows: List[Dict[str, str]],
        header_keys: List[str],
        table_id: str,
        section_type: str,
    ) -> str:
        """Build the user prompt with readable table format, field
        descriptions, and all context needed for accurate mapping."""

        # Format sample data as a readable table
        sample_text = self._format_sample_table(table_columns, sample_rows)

        # Format target schemas with descriptions
        inv_hdr = self._format_targets(INVOICE_HEADER_TARGETS)
        inv_line = self._format_targets(INVOICE_LINE_ITEM_TARGETS)
        ts_hdr = self._format_targets(TIMESHEET_HEADER_TARGETS)
        ts_entry = self._format_targets(TIMESHEET_ENTRY_TARGETS)

        # Format header keys with explanation
        if header_keys:
            hk_text = (
                "These are metadata fields from the document header "
                "(NOT table columns). Map them if they correspond to a "
                "target field.\n"
                + "\n".join(f"  - {k}" for k in header_keys)
            )
        else:
            hk_text = "  (none)"

        return (
            f"## Table to classify and map\n\n"
            f"- **table_id**: {table_id}\n"
            f"- **section_type**: {section_type}\n"
            f"- **columns**: {len(table_columns)}\n"
            f"- **data rows**: {len(sample_rows)} shown below\n\n"
            f"### Section header keys\n"
            f"{hk_text}\n\n"
            f"### Table columns\n"
            f"{table_columns}\n\n"
            f"### Sample data\n"
            f"{sample_text}\n\n"
            f"---\n\n"
            f"## Target schemas\n\n"
            f"**If TblInvoice — Header fields (one value per invoice):**\n"
            f"{inv_hdr}\n\n"
            f"**If TblInvoice — Line item fields (one value per row):**\n"
            f"{inv_line}\n\n"
            f"**If TblTimesheets — Header fields (one value per timesheet):**\n"
            f"{ts_hdr}\n\n"
            f"**If TblTimesheets — Entry fields (one value per row):**\n"
            f"{ts_entry}\n\n"
            f"Produce a MappingProposal.\n"
        )

    @staticmethod
    def _format_sample_table(
        columns: List[str], sample_rows: List[Dict[str, str]]
    ) -> str:
        """Format sample rows as a readable ASCII table so the LLM
        can clearly see column-value alignment."""
        if not sample_rows or not columns:
            return "(no sample data available)"

        # Compute column widths
        col_strs = [str(c) for c in columns]
        widths = [len(c) for c in col_strs]
        for row in sample_rows:
            for ci, col in enumerate(col_strs):
                val = str(row.get(col, ""))
                if len(val) > 40:
                    val = val[:37] + "..."
                if ci < len(widths):
                    widths[ci] = max(widths[ci], len(val))

        # Cap widths
        widths = [min(w, 40) for w in widths]

        # Build header
        header = " | ".join(
            c.ljust(widths[i]) for i, c in enumerate(col_strs)
        )
        separator = "-+-".join("-" * widths[i] for i in range(len(col_strs)))

        # Build rows
        row_lines = []
        for row in sample_rows:
            cells = []
            for ci, col in enumerate(col_strs):
                val = str(row.get(col, ""))
                if len(val) > 40:
                    val = val[:37] + "..."
                cells.append(val.ljust(widths[ci]))
            row_lines.append(" | ".join(cells))

        return f"  {header}\n  {separator}\n" + "\n".join(
            f"  {line}" for line in row_lines
        )

    @staticmethod
    def _format_targets(targets: List[str]) -> str:
        """Format target fields with their descriptions."""
        lines = []
        for name in targets:
            desc = _TARGET_FIELD_DESCRIPTIONS.get(name, "")
            if desc:
                lines.append(f"  - {name}: {desc}")
            else:
                lines.append(f"  - {name}")
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Table discovery
    # ------------------------------------------------------------------

    def _find_all_tables(self, document: Dict) -> List[Dict]:
        """
        Find all tables in the document with context.

        Handles multiple document formats:

        Format A — pdf_table_extractor output:
          sections[].tables[] with columns[].name, data[].values
          Header in sections[].supply_document_header or invoice_header
          or timesheet_header

        Format B — docuporter pipeline (flat sections):
          sections[].content[] with {type: "table", headers, rows}

        Format C — docuporter pipeline (data wrapper):
          sections[].data.content[] with {type: "table", headers, rows}

        Format D — old bucketed format:
          section_type_key[].content[] with {type: "table", headers, rows}
        """
        tables = []

        sections = document.get("sections", [])
        if not isinstance(sections, list):
            sections = []

        # Determine header keys once (searches all possible header locations)
        header_keys = self._extract_header_keys(document)

        for sec in sections:
            if not isinstance(sec, dict):
                continue

            section_name = sec.get(
                "section_name", sec.get("heading", "")
            )
            section_type = sec.get("section_type", "other")

            # ── Format A: sections[].tables[] (pdf_table_extractor) ──
            raw_tables = sec.get("tables", [])
            if isinstance(raw_tables, list) and raw_tables:
                for tbl in raw_tables:
                    if not isinstance(tbl, dict):
                        continue
                    converted = self._convert_extractor_table(
                        tbl, section_name, section_type, header_keys
                    )
                    if converted is not None:
                        tables.append(converted)

            # ── Format B/C: content array with typed blocks ──────────
            content = sec.get("content", [])
            if not content:
                data = sec.get("data", {})
                if isinstance(data, dict):
                    content = data.get("content", [])

            if isinstance(content, list) and content:
                self._collect_tables(
                    content, section_name, section_type,
                    header_keys, tables,
                )

        if tables:
            return tables

        # ── Format D: old bucketed format ────────────────────────────
        skip_keys = {"document_id", "document_header", "sections",
                     "extraction_info"}
        for key, value in document.items():
            if key in skip_keys:
                continue

            if isinstance(value, list):
                for item in value:
                    if not isinstance(item, dict):
                        continue
                    name = item.get(
                        "heading", item.get("section_name", key)
                    )
                    stype = item.get("section_type", key)

                    # Check for tables[] array
                    raw_tables = item.get("tables", [])
                    if isinstance(raw_tables, list):
                        for tbl in raw_tables:
                            if isinstance(tbl, dict):
                                converted = self._convert_extractor_table(
                                    tbl, name, stype, header_keys
                                )
                                if converted is not None:
                                    tables.append(converted)

                    # Check for content array
                    content = item.get("content", [])
                    if not content:
                        data = item.get("data", {})
                        if isinstance(data, dict):
                            content = data.get("content", [])
                    if isinstance(content, list) and content:
                        self._collect_tables(
                            content, name, stype,
                            header_keys, tables,
                        )

        return tables

    def _convert_extractor_table(
        self,
        tbl: Dict,
        section_name: str,
        section_type: str,
        header_keys: List[str],
    ) -> Optional[Dict]:
        """
        Convert a pdf_table_extractor table dict into the standardised
        format used by the mapper.

        Extractor format:
          {
            "table_id": "item_material_pricing",
            "columns": [{"name": "Item No.", "data_type": "text", ...}],
            "data": [{"values": {"Item No.": "010", ...}}],
            ...
          }

        Mapper format:
          {
            "table_id": "...",
            "headers": ["Item No.", ...],
            "rows": [["010", ...], ...],
            ...
          }
        """
        table_id = tbl.get("table_id", "")
        columns = tbl.get("columns", [])
        data_rows = tbl.get("data", [])
        title = tbl.get("title", "")

        # Extract column names
        if isinstance(columns, list) and columns:
            headers = []
            for col in columns:
                if isinstance(col, dict):
                    headers.append(str(col.get("name", "")))
                else:
                    headers.append(str(col))
        else:
            headers = []

        # Convert data rows from {values: {col: val}} to [[val, val, ...]]
        rows = []
        if isinstance(data_rows, list):
            for dr in data_rows:
                if not isinstance(dr, dict):
                    continue
                values = dr.get("values", {})
                if isinstance(values, dict) and headers:
                    row = [str(values.get(h, "")) for h in headers]
                    rows.append(row)
                elif isinstance(values, list):
                    rows.append([str(v) for v in values])

        if not headers and not rows:
            return None

        if not table_id:
            table_id = title if title else f"{section_name}_table"

        table_id = re.sub(r'[^\w\-]', '_', table_id)[:60]

        return {
            "table_id": table_id,
            "section_name": section_name,
            "section_type": section_type,
            "headers": headers,
            "rows": rows,
            "caption": title,
            "header_keys": header_keys,
        }

    def _collect_tables(
        self,
        content: List,
        section_name: str,
        section_type: str,
        header_keys: List[str],
        tables: List[Dict],
    ) -> None:
        """
        Recursively collect table blocks from a content array
        (docuporter pipeline format).

        Handles:
          - type: "table" blocks (headers + rows)
          - type: "subsection" blocks (recurse into content)
          - Untyped blocks with a content array (recurse)
        """
        if not isinstance(content, list):
            return

        for block in content:
            if not isinstance(block, dict):
                continue

            btype = block.get("type", "")

            if btype == "table":
                headers = block.get("headers", [])
                rows = block.get("rows", [])
                caption = block.get("caption", "")

                table_id = caption if caption else f"{section_name}_table"
                table_id = re.sub(r'[^\w\-]', '_', table_id)[:60]

                tables.append({
                    "table_id": table_id,
                    "section_name": section_name,
                    "section_type": section_type,
                    "headers": [str(h) for h in headers],
                    "rows": rows,
                    "caption": caption,
                    "header_keys": header_keys,
                })

            elif btype == "subsection":
                sub_name = block.get("heading", section_name)
                inner = block.get("content", [])
                self._collect_tables(
                    inner, sub_name, section_type,
                    header_keys, tables,
                )

            else:
                # Untyped block — check if it wraps content with tables
                inner = block.get("content", [])
                if isinstance(inner, list) and inner:
                    sub_name = block.get("heading", section_name)
                    self._collect_tables(
                        inner, sub_name, section_type,
                        header_keys, tables,
                    )

    @staticmethod
    def _extract_header_keys(document: Dict) -> List[str]:
        """
        Extract header keys from all possible header locations.

        Searches:
          - document_header (docuporter pipeline)
          - extraction_info (pdf_table_extractor)
          - sections[].supply_document_header
          - sections[].invoice_header
          - sections[].timesheet_header
          - sections[].*_header (any key ending in _header)
        """
        keys = []
        seen = set()

        def _add_keys_from_dict(d: Dict) -> None:
            if not isinstance(d, dict):
                return
            for k, v in d.items():
                if k in ("sections", "tables", "failed_tables",
                         "detection_summary", "extraction_info"):
                    continue
                if k in seen:
                    continue
                if isinstance(v, dict):
                    if v.get("text"):
                        keys.append(k)
                        seen.add(k)
                elif isinstance(v, str) and v.strip():
                    keys.append(k)
                    seen.add(k)

        # Top-level document_header
        doc_header = document.get("document_header", {})
        _add_keys_from_dict(doc_header)

        # Top-level extraction_info
        ext_info = document.get("extraction_info", {})
        _add_keys_from_dict(ext_info)

        # Section-level headers (supply_document_header, invoice_header, etc.)
        sections = document.get("sections", [])
        if isinstance(sections, list):
            for sec in sections:
                if not isinstance(sec, dict):
                    continue
                for sec_key, sec_val in sec.items():
                    if (sec_key.endswith("_header")
                            and isinstance(sec_val, dict)):
                        _add_keys_from_dict(sec_val)

        return keys

    # ------------------------------------------------------------------
    # Azure OpenAI client
    # ------------------------------------------------------------------

    def _get_client(self):
        """Get or create the Azure OpenAI client for structured outputs."""
        if self._client is not None:
            return self._client

        try:
            from config.settings import (
                AZURE_OPENAI_ENDPOINT,
                AZURE_OPENAI_API_KEY,
                AZURE_OPENAI_API_VERSION,
                AZURE_OPENAI_TIMEOUT,
            )
            from openai import AzureOpenAI

            if not AZURE_OPENAI_ENDPOINT or not AZURE_OPENAI_API_KEY:
                logger.error(
                    "Azure OpenAI not configured — "
                    "set AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY"
                )
                return None

            self._client = AzureOpenAI(
                azure_endpoint=AZURE_OPENAI_ENDPOINT,
                api_key=AZURE_OPENAI_API_KEY,
                api_version=AZURE_OPENAI_API_VERSION,
                timeout=AZURE_OPENAI_TIMEOUT,
                http_client=httpx.Client(verify=False),
            )
            logger.info("Azure OpenAI client initialised")
            return self._client

        except ImportError:
            logger.error(
                "openai package not installed — pip install openai"
            )
            return None
        except Exception as e:
            logger.error(f"Failed to create Azure OpenAI client: {e}")
            return None