"""
Verification report writer.

Generates an Excel (.xlsx) report from verification results, with
colour-coded verdict cells, and a JSON accuracy summary file.
"""
import json
from pathlib import Path
from typing import Any, Dict, List

from src.utils import setup_logger

logger = setup_logger("verification_reporter")

# Verdict colour map (openpyxl ARGB hex, no leading #)
VERDICT_COLOURS = {
    "CORRECT":       "FF92D050",   # green
    "INCORRECT":     "FFFF0000",   # red
    "MISSING":       "FFFFC000",   # amber/orange
    "UNVERIFIABLE":  "FFD9D9D9",   # light grey
}
HEADER_COLOUR = "FF4472C4"         # blue


def write_excel_report(
    report: Dict[str, Any],
    output_path: Path,
    extra_columns: List[str] = None,
) -> Path:
    """
    Write an Excel report with one row per verification result.

    Columns:
        - All original columns from the extraction file
        - verdict       (colour-coded)
        - reasoning

    Plus a Summary sheet with overall accuracy stats.

    Args:
        report:        The dict returned by ExtractionVerifier.verify().
        output_path:   Where to write the .xlsx file.
        extra_columns: Additional column names beyond info/page/verdict/reasoning
                       to include (inferred automatically if None).

    Returns:
        Path to the written file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required. Install with: pip install openpyxl"
        )

    results = report.get("results", [])
    accuracy = report.get("accuracy", {})
    document_id = report.get("document_id", "")

    # Infer all column names preserving insertion order
    col_set: dict = {}  # ordered set via dict keys
    for r in results:
        for k in r.keys():
            col_set[k] = None
    # Ensure verdict + reasoning are last
    for special in ("verdict", "reasoning"):
        col_set.pop(special, None)
    all_cols = list(col_set.keys()) + ["verdict", "reasoning"]

    wb = Workbook()

    # ── Sheet 1: Results ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Verification Results"

    thin = Side(style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(name="Arial", bold=True, color="FFFFFFFF", size=11)
    header_fill = PatternFill("solid", start_color=HEADER_COLOUR)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write header row
    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 30

    # Write data rows
    body_font = Font(name="Arial", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")

    for row_idx, result in enumerate(results, 2):
        verdict = result.get("verdict", "UNVERIFIABLE")
        verdict_colour = VERDICT_COLOURS.get(verdict, VERDICT_COLOURS["UNVERIFIABLE"])

        for col_idx, col_name in enumerate(all_cols, 1):
            value = result.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.border = border

            if col_name == "verdict":
                cell.fill = PatternFill("solid", start_color=verdict_colour)
                cell.alignment = center_align
                cell.font = Font(
                    name="Arial", size=10, bold=True,
                    color="FFFFFFFF" if verdict in ("CORRECT", "INCORRECT") else "FF000000",
                )
            else:
                cell.alignment = wrap_align

    # Set column widths
    col_widths = {
        "verdict": 14,
        "reasoning": 55,
    }
    default_widths = {
        "information": 50,
        "page_number": 12,
        "page": 12,
    }
    for col_idx, col_name in enumerate(all_cols, 1):
        col_letter = get_column_letter(col_idx)
        width = (
            col_widths.get(col_name)
            or default_widths.get(col_name)
            or 25
        )
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"

    # ── Sheet 2: Summary ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    summary_rows = [
        ("Document", document_id),
        ("Total Items", accuracy.get("total", 0)),
        ("Verifiable Items", accuracy.get("verifiable", 0)),
        ("", ""),
        ("CORRECT", accuracy.get("correct", 0)),
        ("INCORRECT", accuracy.get("incorrect", 0)),
        ("MISSING", accuracy.get("missing", 0)),
        ("UNVERIFIABLE", accuracy.get("unverifiable", 0)),
        ("", ""),
        ("Accuracy (%)", accuracy.get("accuracy_pct", 0.0)),
    ]

    label_font = Font(name="Arial", bold=True, size=11)
    value_font = Font(name="Arial", size=11)

    for row_idx, (label, value) in enumerate(summary_rows, 1):
        lc = ws2.cell(row=row_idx, column=1, value=label)
        vc = ws2.cell(row=row_idx, column=2, value=value)
        lc.font = label_font
        vc.font = value_font

        # Colour the verdict rows
        colour = VERDICT_COLOURS.get(label.upper())
        if colour:
            vc.fill = PatternFill("solid", start_color=colour)
            vc.font = Font(
                name="Arial", size=11,
                color="FFFFFFFF" if label.upper() in ("CORRECT", "INCORRECT") else "FF000000",
            )

        if label == "Accuracy (%)":
            vc.number_format = "0.00%"
            vc.value = (accuracy.get("accuracy_pct", 0.0) or 0) / 100

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 30

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Excel report saved: {output_path}")
    return output_path


def write_json_summary(
    report: Dict[str, Any],
    output_path: Path,
) -> Path:
    """Write the accuracy summary as a JSON file."""
    summary = {
        "document_id": report.get("document_id", ""),
        "accuracy": report.get("accuracy", {}),
    }
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)
    logger.info(f"JSON summary saved: {output_path}")
    return output_path
