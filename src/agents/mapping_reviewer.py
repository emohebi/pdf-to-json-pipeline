"""
Mapping Review Agent — verifies the accuracy of field mapping results
using a second LLM call as a QA checker.

For each mapped table, the reviewer receives:
  1. The raw source table (headers, rows, header values)
  2. The mapping proposal (classification + field mappings)
  3. The materialised output rows (transformed target data)

It asks the LLM to act as a QA auditor and check:
  - Classification correctness (TblInvoice vs TblTimesheets vs SKIP)
  - Mapping correctness (each source→target assignment)
  - Value correctness (transformed values match the source)
  - Completeness (unmapped fields that should have been mapped)

Uses Azure OpenAI structured outputs for reliable, schema-valid verdicts.

This agent is standalone — it is NOT part of the main pipeline.
"""
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional

import httpx
from enum import Enum
from pydantic import BaseModel, Field

from src.utils import setup_logger, StorageManager

logger = setup_logger("mapping_reviewer")


# =====================================================================
# Pydantic models for structured review output
# =====================================================================

class Verdict(str, Enum):
    CORRECT = "correct"
    INCORRECT = "incorrect"
    MISSING = "missing"
    UNNECESSARY = "unnecessary"


class ClassificationVerdict(BaseModel):
    """Verdict on whether the table was classified correctly."""
    verdict: Verdict = Field(
        description="'correct' if classification is right, 'incorrect' if wrong."
    )
    actual_classification: str = Field(
        description="What the mapper classified this table as."
    )
    expected_classification: str = Field(
        description="What the classification SHOULD be (same as actual if correct)."
    )
    reasoning: str = Field(
        description="Why the classification is correct or incorrect."
    )


class FieldMappingVerdict(BaseModel):
    """Verdict on a single source→target mapping."""
    source: str = Field(description="Source field name.")
    mapped_target: str = Field(
        description="What the mapper assigned as target (or 'UNMAPPED')."
    )
    verdict: Verdict = Field(
        description=(
            "'correct' if mapping is right, "
            "'incorrect' if wrong target, "
            "'missing' if should have been mapped but was UNMAPPED, "
            "'unnecessary' if was mapped but should be UNMAPPED."
        )
    )
    expected_target: str = Field(
        description="What the target SHOULD be (same as mapped_target if correct, or the correct target)."
    )
    reasoning: str = Field(description="Brief explanation.")


class ValueVerdict(BaseModel):
    """Verdict on a single transformed output value."""
    target_field: str = Field(description="Target column name.")
    output_value: str = Field(description="The value in the materialised output.")
    verdict: Verdict = Field(
        description=(
            "'correct' if value faithfully represents the source, "
            "'incorrect' ONLY if the numeric value differs from the "
            "source. Never mark incorrect based on cross-field "
            "arithmetic or pricing basis adjustments."
        )
    )
    expected_value: str = Field(
        description=(
            "What the value SHOULD be based on the raw source field "
            "value and formatting transform only. Must equal "
            "output_value when verdict is correct."
        )
    )
    reasoning: str = Field(description="Brief explanation.")


class TableReviewResult(BaseModel):
    """Complete review result for one table."""
    table_id: str = Field(description="Identifier of the reviewed table.")
    classification_verdict: ClassificationVerdict = Field(
        description="Verdict on the table classification."
    )
    mapping_verdicts: list[FieldMappingVerdict] = Field(
        description="One verdict per source field mapping."
    )
    value_verdicts: list[ValueVerdict] = Field(
        description="Verdicts on output values (only for incorrect or notable values)."
    )
    overall_accuracy_pct: float = Field(
        ge=0.0, le=100.0,
        description="Overall accuracy percentage (0-100)."
    )
    summary: str = Field(
        description="1-2 sentence summary of the review findings."
    )


# =====================================================================
# System prompt
# =====================================================================

_REVIEW_SYSTEM_PROMPT = """\
You are a data-mapping QA auditor for Australian mining industry invoices and timesheets.

You are reviewing the output of an automated field mapper. The mapper:
1. Classified a table as TblInvoice, TblTimesheets, or SKIP
2. Mapped each source field to a target database column
3. Transformed the values (currency parsing, date parsing, etc.)

Your job is to CHECK the mapper's work by examining the raw source data,
the proposed mappings, and the materialised output.

## CLASSIFICATION CHECK

Verify the table type:
- TblInvoice — tables with invoice line items: descriptions, quantities, prices, amounts.
  Includes supply documents, delivery documents, purchase orders, credit notes.
- TblTimesheets — tables with timesheet data: worker names, dates, hours, shifts, rates.
- SKIP — summary/totals tables, bank details, remittance, tax summaries, signatures.

## MAPPING CHECK

For each source→target mapping, verify:
- Is the target column semantically correct for the source field?
- Key distinctions to check:
  * "description" (free text) vs "item_description" (schedule code like "10.10 Supervisor")
  * "amount" (line total per row) vs "net_amount" (invoice total ex-GST) vs "gross_amount" (invoice total inc-GST)
  * "unit_price" (price per unit) vs "amount" (line total = qty x unit_price)
  * "rate" (timesheet hourly/daily rate) vs "charge" (timesheet line total)
  * "pricing_type" must be a CATEGORY (LABOUR, MATERIALS, etc.) — NOT a role name or description
  * "currency" must be an ISO code (AUD, USD) — NOT descriptive text

Mark as:
- "correct" — mapping is right
- "incorrect" — wrong target, specify what it should be
- "missing" — marked UNMAPPED but should have been mapped, specify the correct target
- "unnecessary" — mapped to a target but should be UNMAPPED (no good match exists)

## VALUE CHECK — FIDELITY ONLY

Your job is to check that each source value was correctly transferred to its
target field with the right formatting transform applied. You are NOT checking
business logic or cross-field arithmetic.

For each output value, verify:
- Does the output value faithfully represent the source value?
- Currency: "$1,234.56" should become "1234.56" (no $ or commas)
- Dates: should be normalised to YYYY-MM-DD (ISO 8601)
- Numeric values with thousand separators: "59,530.00" should become "59530.00"
- Empty values: should be empty string, not "None" or "nan"

**ABSOLUTE PROHIBITIONS — NEVER DO THESE:**
- NEVER calculate quantity × unit_price and compare to amount
- NEVER adjust unit_price based on price_units, price_uom, or any pricing basis
- NEVER flag a value as incorrect because it "doesn't match" another field's value
- NEVER suggest a value should be divided, multiplied, or otherwise arithmetically
  transformed based on other fields in the row
- Cross-field arithmetic and business logic are COMPLETELY out of scope

**price_units example — DO NOT FLAG AS INCORRECT:**
  Source has: unit_price = 891.96, price_units = 1000 (meaning "per 1000 KG")
  Output has: unit_price = 891.96
  This is CORRECT. The source says 891.96, the output says 891.96.
  Do NOT suggest it should be 0.89196 or any other adjusted value.
  The mapper's job is to transfer the value faithfully, NOT to do arithmetic.

**Numeric equivalence — STRICT RULE:**
  Trailing zeros, leading zeros after a decimal, and differences in decimal
  places do NOT make a value incorrect. These are all CORRECT and equivalent:
    "36660.00" = "36660" = "36660.0"
    "0.89" = "0.890" = "0.89000"
    "891.96" = "891.960"
    "1000" = "1000.00"
  When comparing numeric values, compare the NUMERIC VALUE, not the string
  representation. If the numbers are mathematically equal, the value is CORRECT.

A value is CORRECT if it accurately represents what was in the source field,
with the appropriate formatting transform applied. A value is INCORRECT only
if it does not match the source — e.g. wrong number, wrong date, garbled text,
or transform not applied (currency symbols still present, date not normalised).

Only report values that are INCORRECT. Do not list every correct value.

**FINAL HARD RULE:** If the output value matches the source value (after
formatting transform), the verdict is CORRECT. Period. No exceptions.
Do not override this verdict based on price_units, price_uom, or any
other field. If got = expected, it is CORRECT.

## REFERENCE FIELDS — CRITICAL (do NOT flag as errors)

The following fields are populated by a SEPARATE deterministic process AFTER
mapping, NOT by the mapper itself. They are injected from extraction_info by
the materialise_tables() function:
  - contract_number
  - po_number
  - ses_number
  - sap_invoice_reference

This means:
1. The mapper correctly marks these source fields as UNMAPPED (they should NOT
   be mapped by the LLM).
2. The values WILL appear in the materialised output even though the mapping
   says UNMAPPED. This is CORRECT and EXPECTED behaviour.
3. Do NOT mark these mappings as "unnecessary" or "missing" — they are CORRECT
   as UNMAPPED because the pipeline handles them separately.
4. Do NOT mark the presence of these values in the output as "incorrect" —
   they were correctly injected by the pipeline, not by the mapping.

If you see contract_number, po_number, ses_number, or sap_invoice_reference
in the materialised output despite being UNMAPPED, this is normal pipeline
behaviour. Mark the mapping as "correct" and skip value checking for these fields.

## DOCUMENT TOTAL / NET vs GROSS

For Australian supply documents, delivery documents, and inter-company transfers,
the document total is typically shown BEFORE tax (ex-GST). Only map to
"gross_amount" when a separate GST line confirms tax IS included in the total.
When only one total is shown with no GST breakdown, it should map to "net_amount".

## ACCURACY SCORE

Calculate overall_accuracy_pct as:
  (correct_classifications + correct_mappings + correct_values) / total_checks * 100

Be strict but fair. Minor formatting differences are acceptable.
"""

# Number of sample rows to show in the review prompt
_MAX_REVIEW_ROWS = 5


# =====================================================================
# Agent
# =====================================================================

class MappingReviewAgent:
    """
    Review the accuracy of field mapping results using a second
    LLM call as a QA checker.

    Uses Azure OpenAI structured outputs for schema-valid verdicts.
    """

    def __init__(self):
        self.storage = StorageManager()
        self._client = None

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def review_mapping(
        self,
        document: Dict,
        mapping_report: Dict,
        materialised_data: Dict[str, Any] = None,
        document_id: str = "",
    ) -> Dict[str, Any]:
        """
        Review all mapped tables in a mapping report.

        Args:
            document:          The original extracted JSON document.
            mapping_report:    The mapping report from FieldMappingAgent.
            materialised_data: Optional dict of {table_name: DataFrame}
                               from materialise_tables(). If not provided,
                               value checks are skipped.
            document_id:       Identifier for logging.

        Returns:
            Review report dict with per-table verdicts and overall stats.
        """
        if not document_id:
            document_id = mapping_report.get("document_id", "unknown")

        logger.info(f"[{document_id}] Starting mapping review")

        proposals = mapping_report.get("proposals", [])
        if not proposals:
            logger.info(f"[{document_id}] No proposals to review")
            return self._empty_report(document_id)

        client = self._get_client()
        if client is None:
            logger.error(
                f"[{document_id}] Azure OpenAI client not available"
            )
            return self._empty_report(document_id, error="Azure OpenAI not configured")

        # Find raw tables for lookup
        from src.agents.field_mapper import FieldMappingAgent
        mapper = FieldMappingAgent()
        all_tables = mapper._find_all_tables(document)
        tables_by_id = {t["table_id"]: t for t in all_tables}

        # Collect header values
        header_values = mapper._extract_header_values(document)

        # Build materialised rows lookup
        mat_rows_by_table: Dict[str, List[Dict]] = {}
        if materialised_data:
            try:
                for tbl_name, df in materialised_data.items():
                    for _, row in df.iterrows():
                        source_table = row.get("_source_table", "")
                        if source_table:
                            mat_rows_by_table.setdefault(
                                source_table, []
                            ).append(row.to_dict())
            except Exception:
                pass

        # Review each proposal
        table_reviews: List[Dict] = []
        for i, proposal in enumerate(proposals, 1):
            source_info = proposal.get("_source", {})
            table_id = source_info.get("table_id", f"table_{i}")
            target_table = proposal.get("target_table", "")

            logger.info(
                f"[{document_id}]   Reviewing table {i}/{len(proposals)}: "
                f"'{table_id}' -> {target_table}"
            )

            raw_table = tables_by_id.get(table_id)
            mat_rows = mat_rows_by_table.get(table_id, [])

            review = self._review_table(
                client=client,
                proposal=proposal,
                raw_table=raw_table,
                header_values=header_values,
                materialised_rows=mat_rows,
                document_id=document_id,
            )

            if review is not None:
                review_dict = review.model_dump()
                review_dict["table_id"] = table_id
                table_reviews.append(review_dict)

                logger.info(
                    f"[{document_id}]     Accuracy: "
                    f"{review.overall_accuracy_pct:.1f}% — "
                    f"{review.summary[:80]}"
                )
            else:
                table_reviews.append({
                    "table_id": table_id,
                    "error": "Review failed",
                })

        # Compute aggregate stats
        stats = self._compute_stats(table_reviews)

        report = {
            "document_id": document_id,
            "tables_reviewed": len(table_reviews),
            "table_reviews": table_reviews,
            "statistics": stats,
        }

        # Save
        try:
            report_path = (
                self.storage.final_dir
                / f"{document_id}_mapping_review.json"
            )
            report_path.parent.mkdir(parents=True, exist_ok=True)
            with open(report_path, "w", encoding="utf-8") as f:
                json.dump(report, f, indent=2, ensure_ascii=False)
            logger.info(f"[{document_id}] Review saved: {report_path}")
        except Exception as e:
            logger.warning(f"[{document_id}] Failed to save review: {e}")

        return report

    # ------------------------------------------------------------------
    # Per-table review
    # ------------------------------------------------------------------

    def _review_table(
        self,
        client,
        proposal: Dict,
        raw_table: Optional[Dict],
        header_values: Dict[str, str],
        materialised_rows: List[Dict],
        document_id: str,
    ) -> Optional[TableReviewResult]:
        """Review a single table's mapping using the LLM."""
        from config.settings import (
            AZURE_OPENAI_DEPLOYMENT,
            MAX_RETRIES,
            RETRY_DELAY,
        )
        import time

        user_prompt = self._build_review_prompt(
            proposal, raw_table, header_values, materialised_rows,
        )

        last_error = None
        for attempt in range(MAX_RETRIES):
            try:
                response = client.beta.chat.completions.parse(
                    model=AZURE_OPENAI_DEPLOYMENT,
                    messages=[
                        {"role": "system", "content": _REVIEW_SYSTEM_PROMPT},
                        {"role": "user", "content": user_prompt},
                    ],
                    response_format=TableReviewResult,
                    temperature=0.0,
                    max_completion_tokens=10240,
                )

                result = response.choices[0].message.parsed
                if result is not None:
                    return result

            except Exception as e:
                last_error = e
                logger.warning(
                    f"[{document_id}] Review attempt "
                    f"{attempt + 1}/{MAX_RETRIES} failed: {e}"
                )
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY * (attempt + 1))

        logger.error(
            f"[{document_id}] All review attempts failed: {last_error}"
        )
        return None

    # ------------------------------------------------------------------
    # Prompt builder
    # ------------------------------------------------------------------

    def _build_review_prompt(
        self,
        proposal: Dict,
        raw_table: Optional[Dict],
        header_values: Dict[str, str],
        materialised_rows: List[Dict],
    ) -> str:
        """Build the review prompt showing source, mappings, and output."""
        source_info = proposal.get("_source", {})
        table_id = source_info.get("table_id", "?")
        target_table = proposal.get("target_table", "?")
        classification_reasoning = proposal.get(
            "classification_reasoning", ""
        )
        mappings = proposal.get("mappings", [])
        unmapped_targets = proposal.get("unmapped_target_fields", [])

        # Format raw source table
        if raw_table:
            headers = raw_table.get("headers", [])
            rows = raw_table.get("rows", [])
            source_table_text = self._format_source_table(headers, rows)
        else:
            source_table_text = "(raw table data not available)"

        # Format header values (only non-empty ones)
        hdr_text = "\n".join(
            f"  {k}: {v}" for k, v in header_values.items()
            if v and v.strip()
        ) or "  (none)"

        # Format the mappings
        mapping_lines = []
        for m in mappings:
            src = m.get("source", "")
            tgt = m.get("target", "UNMAPPED")
            conf = m.get("confidence", 0)
            hint = m.get("transform_hint", "none")
            reason = m.get("reasoning", "")
            mapping_lines.append(
                f"  {src:30s} -> {tgt:25s} "
                f"(conf={conf:.2f}, hint={hint}) {reason}"
            )
        mappings_text = "\n".join(mapping_lines) or "  (no mappings)"

        # Format materialised output
        if materialised_rows:
            output_lines = []
            for ri, row in enumerate(materialised_rows[:_MAX_REVIEW_ROWS], 1):
                # Filter out internal fields for display
                display_row = {
                    k: v for k, v in row.items()
                    if not k.startswith("_") and v and str(v).strip()
                    and str(v) != "nan"
                }
                output_lines.append(f"  Row {ri}: {display_row}")
            output_text = "\n".join(output_lines)
            if len(materialised_rows) > _MAX_REVIEW_ROWS:
                output_text += (
                    f"\n  ... ({len(materialised_rows)} rows total)"
                )
        else:
            output_text = "  (no materialised output available)"

        return (
            f"## Table to review\n\n"
            f"- **table_id**: {table_id}\n"
            f"- **classification**: {target_table}\n"
            f"- **classification_reasoning**: {classification_reasoning}\n\n"
            f"### Raw source table\n"
            f"{source_table_text}\n\n"
            f"### Header values (from document/section metadata)\n"
            f"{hdr_text}\n\n"
            f"### Mapper's field mappings\n"
            f"{mappings_text}\n\n"
            f"### Unmapped target fields\n"
            f"  {unmapped_targets}\n\n"
            f"### Materialised output (after transforms)\n"
            f"{output_text}\n\n"
            f"---\n\n"
            f"Review the classification, every mapping, and the output "
            f"values. Produce a TableReviewResult.\n"
        )

    @staticmethod
    def _format_source_table(
        headers: List[str], rows: List
    ) -> str:
        """Format the raw source table as a readable ASCII table."""
        if not headers:
            return "  (no headers)"

        col_strs = [str(h) for h in headers]
        widths = [len(c) for c in col_strs]
        display_rows = rows[:_MAX_REVIEW_ROWS]

        for row in display_rows:
            if isinstance(row, list):
                for ci in range(min(len(row), len(widths))):
                    val = str(row[ci])
                    if len(val) > 35:
                        val = val[:32] + "..."
                    widths[ci] = max(widths[ci], len(val))

        widths = [min(w, 35) for w in widths]

        header_line = " | ".join(
            c.ljust(widths[i]) for i, c in enumerate(col_strs)
        )
        sep_line = "-+-".join("-" * w for w in widths)

        row_lines = []
        for row in display_rows:
            if isinstance(row, list):
                cells = []
                for ci in range(len(col_strs)):
                    val = str(row[ci]) if ci < len(row) else ""
                    if len(val) > 35:
                        val = val[:32] + "..."
                    cells.append(val.ljust(widths[ci]))
                row_lines.append(" | ".join(cells))

        result = f"  {header_line}\n  {sep_line}\n"
        result += "\n".join(f"  {line}" for line in row_lines)

        if len(rows) > _MAX_REVIEW_ROWS:
            result += f"\n  ... ({len(rows)} rows total)"

        return result

    # ------------------------------------------------------------------
    # Statistics
    # ------------------------------------------------------------------

    @staticmethod
    def _compute_stats(table_reviews: List[Dict]) -> Dict[str, Any]:
        """Compute aggregate statistics across all reviewed tables."""
        completed = [
            r for r in table_reviews if "error" not in r
        ]

        if not completed:
            return {
                "tables_reviewed": 0,
                "average_accuracy_pct": 0.0,
                "classification_correct": 0,
                "classification_incorrect": 0,
                "mappings_correct": 0,
                "mappings_incorrect": 0,
                "mappings_missing": 0,
                "mappings_unnecessary": 0,
                "values_correct": 0,
                "values_incorrect": 0,
            }

        accuracies = [
            r.get("overall_accuracy_pct", 0) for r in completed
        ]

        cls_correct = sum(
            1 for r in completed
            if r.get("classification_verdict", {}).get("verdict") == "correct"
        )
        cls_incorrect = len(completed) - cls_correct

        m_correct = m_incorrect = m_missing = m_unnecessary = 0
        for r in completed:
            for mv in r.get("mapping_verdicts", []):
                v = mv.get("verdict", "")
                if v == "correct":
                    m_correct += 1
                elif v == "incorrect":
                    m_incorrect += 1
                elif v == "missing":
                    m_missing += 1
                elif v == "unnecessary":
                    m_unnecessary += 1

        v_correct = v_incorrect = 0
        for r in completed:
            for vv in r.get("value_verdicts", []):
                v = vv.get("verdict", "")
                if v == "correct":
                    v_correct += 1
                elif v == "incorrect":
                    v_incorrect += 1

        return {
            "tables_reviewed": len(completed),
            "average_accuracy_pct": round(
                sum(accuracies) / len(accuracies), 1
            ),
            "classification_correct": cls_correct,
            "classification_incorrect": cls_incorrect,
            "mappings_correct": m_correct,
            "mappings_incorrect": m_incorrect,
            "mappings_missing": m_missing,
            "mappings_unnecessary": m_unnecessary,
            "values_correct": v_correct,
            "values_incorrect": v_incorrect,
        }

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _empty_report(
        document_id: str, error: str = None
    ) -> Dict[str, Any]:
        report = {
            "document_id": document_id,
            "tables_reviewed": 0,
            "table_reviews": [],
            "statistics": {
                "tables_reviewed": 0,
                "average_accuracy_pct": 0.0,
            },
        }
        if error:
            report["error"] = error
        return report

    # ------------------------------------------------------------------
    # Excel report writer
    # ------------------------------------------------------------------

    @staticmethod
    def write_review_excel(
        report: Dict[str, Any],
        output_path: Path,
    ) -> Path:
        """
        Write review results to a formatted Excel workbook.

        Sheets:
          1. Summary        — overall stats and per-table accuracy
          2. Classification — one row per table with verdict
          3. Mappings       — one row per field mapping with verdict
          4. Values         — one row per incorrect value
        """
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side,
        )
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Shared styles
        thin = Side(style="thin", color="FFCCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font = Font(
            name="Arial", bold=True, color="FFFFFFFF", size=10,
        )
        hdr_fill = PatternFill("solid", fgColor="FF2B5797")
        hdr_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
        body_font = Font(name="Arial", size=10)
        wrap = Alignment(vertical="top", wrap_text=True)
        center = Alignment(horizontal="center", vertical="top")

        # Verdict colours
        verdict_fills = {
            "correct": PatternFill("solid", fgColor="FF92D050"),
            "incorrect": PatternFill("solid", fgColor="FFFF6666"),
            "missing": PatternFill("solid", fgColor="FFFFC000"),
            "unnecessary": PatternFill("solid", fgColor="FFD9D9D9"),
        }

        document_id = report.get("document_id", "")
        table_reviews = report.get("table_reviews", [])
        stats = report.get("statistics", {})

        # ── Sheet 1: Summary ─────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Summary"

        label_font = Font(name="Arial", bold=True, size=11)
        val_font = Font(name="Arial", size=11)

        summary_rows = [
            ("Document", document_id),
            ("Tables Reviewed", stats.get("tables_reviewed", 0)),
            ("Average Accuracy %", stats.get("average_accuracy_pct", 0)),
            ("", ""),
            ("CLASSIFICATION", ""),
            ("Correct", stats.get("classification_correct", 0)),
            ("Incorrect", stats.get("classification_incorrect", 0)),
            ("", ""),
            ("FIELD MAPPINGS", ""),
            ("Correct", stats.get("mappings_correct", 0)),
            ("Incorrect", stats.get("mappings_incorrect", 0)),
            ("Missing", stats.get("mappings_missing", 0)),
            ("Unnecessary", stats.get("mappings_unnecessary", 0)),
            ("", ""),
            ("VALUE CHECKS", ""),
            ("Correct", stats.get("values_correct", 0)),
            ("Incorrect", stats.get("values_incorrect", 0)),
        ]

        # Per-table accuracy rows
        if table_reviews:
            summary_rows.append(("", ""))
            summary_rows.append(("PER-TABLE ACCURACY", ""))
            for tr in table_reviews:
                tid = tr.get("table_id", "?")
                acc = tr.get("overall_accuracy_pct", 0)
                summary_rows.append((tid, f"{acc:.1f}%"))

        for ri, (label, value) in enumerate(summary_rows, 1):
            lc = ws1.cell(row=ri, column=1, value=label)
            vc = ws1.cell(row=ri, column=2, value=value)
            lc.font = label_font
            vc.font = val_font

        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 30

        # ── Sheet 2: Classification ──────────────────────────────
        ws2 = wb.create_sheet("Classification")

        # Detect combined mode: _document_id present on reviews
        is_combined = any(
            tr.get("_document_id") for tr in table_reviews
        )

        cls_cols = (
            ["Source File", "Table ID", "Verdict",
             "Actual", "Expected", "Reasoning"]
            if is_combined
            else ["Table ID", "Verdict", "Actual",
                  "Expected", "Reasoning"]
        )
        for ci, col in enumerate(cls_cols, 1):
            c = ws2.cell(row=1, column=ci, value=col)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = border

        ws2.freeze_panes = "A2"

        for ri, tr in enumerate(table_reviews, 2):
            cls_v = tr.get("classification_verdict", {})
            verdict = cls_v.get("verdict", "")

            row_data = [
                tr.get("table_id", ""),
                verdict,
                cls_v.get("actual_classification", ""),
                cls_v.get("expected_classification", ""),
                cls_v.get("reasoning", ""),
            ]
            if is_combined:
                row_data.insert(0, tr.get("_document_id", ""))

            verdict_ci = 3 if is_combined else 2
            reason_ci = len(row_data)

            for ci, val in enumerate(row_data, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.font = body_font
                cell.border = border
                cell.alignment = (
                    wrap if ci == reason_ci else center
                )
                if ci == verdict_ci:
                    fill = verdict_fills.get(verdict)
                    if fill:
                        cell.fill = fill

        if is_combined:
            ws2.column_dimensions["A"].width = 40
            ws2.column_dimensions["B"].width = 30
            ws2.column_dimensions["C"].width = 14
            ws2.column_dimensions["D"].width = 18
            ws2.column_dimensions["E"].width = 18
            ws2.column_dimensions["F"].width = 60
        else:
            ws2.column_dimensions["A"].width = 30
            ws2.column_dimensions["B"].width = 14
            ws2.column_dimensions["C"].width = 18
            ws2.column_dimensions["D"].width = 18
            ws2.column_dimensions["E"].width = 60

        # ── Sheet 3: Mappings ────────────────────────────────────
        ws3 = wb.create_sheet("Mappings")
        map_cols = (
            ["Source File", "Table ID", "Source Field",
             "Mapped Target", "Verdict", "Expected Target",
             "Reasoning"]
            if is_combined
            else ["Table ID", "Source Field", "Mapped Target",
                  "Verdict", "Expected Target", "Reasoning"]
        )
        for ci, col in enumerate(map_cols, 1):
            c = ws3.cell(row=1, column=ci, value=col)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = border

        ws3.freeze_panes = "A2"

        mr = 2
        for tr in table_reviews:
            table_id = tr.get("table_id", "")
            doc_id = tr.get("_document_id", "")
            for mv in tr.get("mapping_verdicts", []):
                verdict = mv.get("verdict", "")
                row_data = [
                    table_id,
                    mv.get("source", ""),
                    mv.get("mapped_target", ""),
                    verdict,
                    mv.get("expected_target", ""),
                    mv.get("reasoning", ""),
                ]
                if is_combined:
                    row_data.insert(0, doc_id)

                verdict_ci = 5 if is_combined else 4
                reason_ci = len(row_data)

                for ci, val in enumerate(row_data, 1):
                    cell = ws3.cell(row=mr, column=ci, value=val)
                    cell.font = body_font
                    cell.border = border
                    cell.alignment = (
                        wrap if ci == reason_ci else center
                    )
                    if ci == verdict_ci:
                        fill = verdict_fills.get(verdict)
                        if fill:
                            cell.fill = fill
                mr += 1

        if is_combined:
            ws3.column_dimensions["A"].width = 40
            ws3.column_dimensions["B"].width = 25
            ws3.column_dimensions["C"].width = 30
            ws3.column_dimensions["D"].width = 25
            ws3.column_dimensions["E"].width = 14
            ws3.column_dimensions["F"].width = 25
            ws3.column_dimensions["G"].width = 55
        else:
            ws3.column_dimensions["A"].width = 25
            ws3.column_dimensions["B"].width = 30
            ws3.column_dimensions["C"].width = 25
            ws3.column_dimensions["D"].width = 14
            ws3.column_dimensions["E"].width = 25
            ws3.column_dimensions["F"].width = 55

        # ── Sheet 4: Values ──────────────────────────────────────
        ws4 = wb.create_sheet("Values")
        val_cols = (
            ["Source File", "Table ID", "Target Field",
             "Output Value", "Verdict", "Expected Value",
             "Reasoning"]
            if is_combined
            else ["Table ID", "Target Field", "Output Value",
                  "Verdict", "Expected Value", "Reasoning"]
        )
        for ci, col in enumerate(val_cols, 1):
            c = ws4.cell(row=1, column=ci, value=col)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = border

        ws4.freeze_panes = "A2"

        vr = 2
        for tr in table_reviews:
            table_id = tr.get("table_id", "")
            doc_id = tr.get("_document_id", "")
            for vv in tr.get("value_verdicts", []):
                verdict = vv.get("verdict", "")
                row_data = [
                    table_id,
                    vv.get("target_field", ""),
                    vv.get("output_value", ""),
                    verdict,
                    vv.get("expected_value", ""),
                    vv.get("reasoning", ""),
                ]
                if is_combined:
                    row_data.insert(0, doc_id)

                verdict_ci = 5 if is_combined else 4
                wrap_cols = (
                    {3, 5, 6, 7} if is_combined
                    else {3, 5, 6}
                )

                for ci, val in enumerate(row_data, 1):
                    cell = ws4.cell(row=vr, column=ci, value=val)
                    cell.font = body_font
                    cell.border = border
                    cell.alignment = (
                        wrap if ci in wrap_cols else center
                    )
                    if ci == verdict_ci:
                        fill = verdict_fills.get(verdict)
                        if fill:
                            cell.fill = fill
                vr += 1

        if vr == 2:
            ws4.cell(row=2, column=1, value="No value issues found")

        if is_combined:
            ws4.column_dimensions["A"].width = 40
            ws4.column_dimensions["B"].width = 25
            ws4.column_dimensions["C"].width = 25
            ws4.column_dimensions["D"].width = 30
            ws4.column_dimensions["E"].width = 14
            ws4.column_dimensions["F"].width = 30
            ws4.column_dimensions["G"].width = 55
        else:
            ws4.column_dimensions["A"].width = 25
            ws4.column_dimensions["B"].width = 25
            ws4.column_dimensions["C"].width = 30
            ws4.column_dimensions["D"].width = 14
            ws4.column_dimensions["E"].width = 30
            ws4.column_dimensions["F"].width = 55

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path

    # ------------------------------------------------------------------
    # Azure OpenAI client
    # ------------------------------------------------------------------

    def _get_client(self):
        """Get or create the Azure OpenAI client."""
        if self._client is not None:
            return self._client

        try:
            from config.settings import (
                AZURE_OPENAI_ENDPOINT,
                AZURE_OPENAI_API_KEY,
                AZURE_OPENAI_API_VERSION,
                AZURE_OPENAI_TIMEOUT,
            )
            from openai import AzureOpenAI

            if not AZURE_OPENAI_ENDPOINT or not AZURE_OPENAI_API_KEY:
                logger.error(
                    "Azure OpenAI not configured — "
                    "set AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY"
                )
                return None

            self._client = AzureOpenAI(
                azure_endpoint=AZURE_OPENAI_ENDPOINT,
                api_key=AZURE_OPENAI_API_KEY,
                api_version=AZURE_OPENAI_API_VERSION,
                timeout=AZURE_OPENAI_TIMEOUT,
                http_client=httpx.Client(verify=False),
            )
            logger.info("Azure OpenAI client initialised for review")
            return self._client

        except ImportError:
            logger.error("openai package not installed")
            return None
        except Exception as e:
            logger.error(f"Failed to create Azure client: {e}")
            return None