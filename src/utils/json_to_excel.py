#!/usr/bin/env python3
"""
Convert extracted section JSON (new ordered content-block schema)
to a formatted Excel workbook.

The JSON has:
{
  "document_id": "...",
  "document_header": {...},
  "sections": [
    {
      "section_name": "...",
      "section_type": "...",
      "page_range": [start, end],
      "heading": "...",
      "heading_level": "...",
      "content": [
        {"type": "paragraph", "text": "..."},
        {"type": "table", "caption": "...", "headers": [...], "rows": [...]},
        {"type": "subsection", "heading": "...", "content": [...]}
      ]
    }
  ]
}

Output Excel has:
  Sheet 1 "Document"   — flattened rows of all content in reading order
  Sheet 2 "Header"     — document metadata
  Sheet 3 "Tables"     — every table extracted, one per block

Usage:
    python scripts/json_to_excel.py --input output/final/doc.json
    python scripts/json_to_excel.py --input doc.json --output doc.xlsx
"""
import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)


# ======================================================================
# Colours
# ======================================================================

COLORS = {
    "header_bg":      "FF2B5797",  # dark blue
    "header_font":    "FFFFFFFF",  # white
    "section_bg":     "FFD6E4F0",  # light blue
    "subsection_bg":  "FFEDF2F8",  # lighter blue
    "table_bg":       "FFF2F2F2",  # light grey
    "table_header_bg":"FFDCE6F1",  # blue-grey
    "alt_row":        "FFF9F9F9",  # subtle grey
}

THIN_BORDER = Border(
    left=Side(style="thin", color="FFCCCCCC"),
    right=Side(style="thin", color="FFCCCCCC"),
    top=Side(style="thin", color="FFCCCCCC"),
    bottom=Side(style="thin", color="FFCCCCCC"),
)


# ======================================================================
# Flattening: content tree → flat rows
# ======================================================================

def flatten_document(document: Dict) -> List[Dict]:
    """
    Flatten the entire document into ordered rows.

    Each row is a dict with:
      - section_name, section_type, page_range
      - depth (0 = section, 1 = subsection, 2 = sub-sub, ...)
      - block_type (paragraph, table, subsection_heading, table_row)
      - heading (for section/subsection headings)
      - text (for paragraphs)
      - table_caption, table_headers, table_row_data (for tables)
    """
    rows: List[Dict] = []
    sections = document.get("sections", [])

    # Handle old format
    if not sections:
        sections = _convert_old_format(document)

    for sec in sections:
        section_name = sec.get("section_name", sec.get("heading", ""))
        section_type = sec.get("section_type", "section")
        page_range = sec.get("page_range", [])
        page_str = (
            f"{page_range[0]}-{page_range[1]}"
            if isinstance(page_range, list) and len(page_range) == 2
            else str(page_range)
        )

        # Section heading row
        rows.append({
            "section_name": section_name,
            "section_type": section_type,
            "pages": page_str,
            "depth": 0,
            "block_type": "section_heading",
            "heading": sec.get("heading", section_name),
            "heading_level": sec.get("heading_level", "1"),
            "text": "",
        })

        # Flatten content
        content = sec.get("content", [])
        _flatten_content(content, rows, section_name, section_type, page_str, depth=0)

    return rows


def _flatten_content(
    content: List, rows: List[Dict],
    section_name: str, section_type: str, pages: str,
    depth: int,
) -> None:
    """Recursively flatten content blocks into rows."""
    if not isinstance(content, list):
        return

    for block in content:
        if not isinstance(block, dict):
            if isinstance(block, str) and block.strip():
                rows.append(_make_row(
                    section_name, section_type, pages, depth,
                    "paragraph", text=block,
                ))
            continue

        btype = block.get("type", "")

        if btype == "paragraph":
            text = block.get("text", "")
            if text and text.strip():
                rows.append(_make_row(
                    section_name, section_type, pages, depth,
                    "paragraph", text=text,
                ))

        elif btype == "table":
            caption = block.get("caption", "")
            headers = block.get("headers", [])
            table_rows = block.get("rows", [])

            # Table caption row
            rows.append(_make_row(
                section_name, section_type, pages, depth,
                "table_caption",
                text=f"[Table: {caption}]" if caption else "[Table]",
                table_caption=caption,
                table_headers=headers,
            ))

            # Header row
            if headers:
                rows.append(_make_row(
                    section_name, section_type, pages, depth,
                    "table_header",
                    text=" | ".join(str(h) for h in headers),
                    table_headers=headers,
                ))

            # Data rows
            for row_data in table_rows:
                if isinstance(row_data, list):
                    rows.append(_make_row(
                        section_name, section_type, pages, depth,
                        "table_row",
                        text=" | ".join(str(c) for c in row_data),
                        table_row_data=row_data,
                    ))

        elif btype == "subsection":
            heading = block.get("heading", "")
            level = block.get("heading_level", "")

            if heading and heading.strip():
                rows.append(_make_row(
                    section_name, section_type, pages, depth + 1,
                    "subsection_heading",
                    heading=heading, heading_level=level,
                ))

            inner = block.get("content", [])
            _flatten_content(inner, rows, section_name, section_type, pages, depth + 1)

        else:
            # Unknown block type — extract any text
            text = block.get("text", "")
            if text and text.strip():
                rows.append(_make_row(
                    section_name, section_type, pages, depth,
                    "paragraph", text=text,
                ))


def _make_row(
    section_name, section_type, pages, depth, block_type,
    heading="", heading_level="", text="",
    table_caption="", table_headers=None, table_row_data=None,
) -> Dict:
    return {
        "section_name": section_name,
        "section_type": section_type,
        "pages": pages,
        "depth": depth,
        "block_type": block_type,
        "heading": heading,
        "heading_level": heading_level,
        "text": text,
        "table_caption": table_caption,
        "table_headers": table_headers or [],
        "table_row_data": table_row_data or [],
    }


def _convert_old_format(document: Dict) -> List[Dict]:
    """Convert old bucketed format to flat sections list."""
    sections = []
    skip = {"document_id", "document_header"}
    for key, val in document.items():
        if key in skip:
            continue
        if isinstance(val, list):
            for item in val:
                if isinstance(item, dict):
                    sections.append({
                        "section_name": item.get("heading", item.get("section", key)),
                        "section_type": key,
                        "page_range": [],
                        "heading": item.get("heading", ""),
                        "heading_level": item.get("heading_level", ""),
                        "content": item.get("content", []),
                    })
    return sections


# ======================================================================
# Excel writing
# ======================================================================

def write_excel(document: Dict, output_path: Path) -> Path:
    """Write the document to a formatted Excel workbook."""
    wb = Workbook()
    doc_id = document.get("document_id", "document")

    # -- Sheet 1: Document content --
    _write_document_sheet(wb, document)

    # -- Sheet 2: Header / metadata --
    _write_header_sheet(wb, document)

    # -- Sheet 3: Tables --
    _write_tables_sheet(wb, document)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def _write_document_sheet(wb: Workbook, document: Dict) -> None:
    """Main sheet: all content in reading order."""
    ws = wb.active
    ws.title = "Document"

    # Header row
    columns = ["Section", "Type", "Pages", "Depth", "Block", "Content"]
    header_font = Font(name="Arial", bold=True, color=COLORS["header_font"], size=10)
    header_fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Flatten and write rows
    rows = flatten_document(document)

    section_font = Font(name="Arial", bold=True, size=11)
    subsection_font = Font(name="Arial", bold=True, size=10)
    body_font = Font(name="Arial", size=10)
    table_font = Font(name="Arial", size=9, italic=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for row_idx, row in enumerate(rows, 2):
        block_type = row["block_type"]
        depth = row["depth"]

        # Column A: Section name (only on section headings)
        if block_type == "section_heading":
            ws.cell(row=row_idx, column=1, value=row["section_name"])
        # Column B: Type
        ws.cell(row=row_idx, column=2, value=row["section_type"] if block_type == "section_heading" else "")
        # Column C: Pages
        ws.cell(row=row_idx, column=3, value=row["pages"] if block_type == "section_heading" else "")
        # Column D: Depth
        ws.cell(row=row_idx, column=4, value=depth)
        # Column E: Block type
        ws.cell(row=row_idx, column=5, value=block_type)

        # Column F: Content
        indent = "  " * depth
        if block_type == "section_heading":
            content_text = row["heading"]
            font = section_font
        elif block_type == "subsection_heading":
            content_text = f"{indent}{row['heading']}"
            font = subsection_font
        elif block_type in ("table_caption", "table_header"):
            content_text = f"{indent}{row['text']}"
            font = table_font
        elif block_type == "table_row":
            content_text = f"{indent}{row['text']}"
            font = table_font
        else:
            content_text = f"{indent}{row['text']}"
            font = body_font

        content_cell = ws.cell(row=row_idx, column=6, value=content_text)
        content_cell.font = font
        content_cell.alignment = wrap

        # Row styling
        if block_type == "section_heading":
            fill = PatternFill("solid", fgColor=COLORS["section_bg"])
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = fill
                ws.cell(row=row_idx, column=c).font = section_font
        elif block_type == "subsection_heading":
            fill = PatternFill("solid", fgColor=COLORS["subsection_bg"])
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = fill
        elif block_type in ("table_caption", "table_header", "table_row"):
            fill = PatternFill("solid", fgColor=COLORS["table_bg"])
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = fill
        elif row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor=COLORS["alt_row"])
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = fill

        for c in range(1, 7):
            ws.cell(row=row_idx, column=c).border = THIN_BORDER
            if c != 6:
                ws.cell(row=row_idx, column=c).alignment = Alignment(vertical="top")

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 100


def _write_header_sheet(wb: Workbook, document: Dict) -> None:
    """Sheet 2: document header metadata."""
    ws = wb.create_sheet("Header")
    header = document.get("document_header", {})
    doc_id = document.get("document_id", "")

    header_font = Font(name="Arial", bold=True, color=COLORS["header_font"], size=10)
    header_fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    label_font = Font(name="Arial", bold=True, size=10)
    value_font = Font(name="Arial", size=10)

    for col_idx, col_name in enumerate(["Field", "Value"], 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = THIN_BORDER

    row = 2
    # Document ID
    ws.cell(row=row, column=1, value="document_id").font = label_font
    ws.cell(row=row, column=2, value=doc_id).font = value_font
    row += 1

    for key, val in header.items():
        if key == "sections":
            ws.cell(row=row, column=1, value="sections").font = label_font
            if isinstance(val, list):
                ws.cell(row=row, column=2, value=", ".join(str(s) for s in val)).font = value_font
            row += 1
            continue

        ws.cell(row=row, column=1, value=key).font = label_font

        if isinstance(val, dict):
            text = val.get("text", val.get("orig_text", str(val)))
        elif isinstance(val, str):
            text = val
        else:
            text = str(val)

        ws.cell(row=row, column=2, value=text).font = value_font
        row += 1

    for r in range(2, row):
        for c in (1, 2):
            ws.cell(row=r, column=c).border = THIN_BORDER

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80


def _write_tables_sheet(wb: Workbook, document: Dict) -> None:
    """Sheet 3: all tables, each with its own header row."""
    ws = wb.create_sheet("Tables")

    header_font = Font(name="Arial", bold=True, color=COLORS["header_font"], size=10)
    header_fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    th_font = Font(name="Arial", bold=True, size=9)
    th_fill = PatternFill("solid", fgColor=COLORS["table_header_bg"])
    td_font = Font(name="Arial", size=9)
    caption_font = Font(name="Arial", bold=True, size=10, color="FF333333")

    tables = _collect_all_tables(document)

    if not tables:
        ws.cell(row=1, column=1, value="No tables found in document")
        return

    row = 1
    for tidx, tbl in enumerate(tables, 1):
        # Caption / source
        caption = tbl.get("caption", "")
        source = tbl.get("_source_section", "")
        label = f"Table {tidx}"
        if caption:
            label += f": {caption}"
        if source:
            label += f"  (from: {source})"

        cell = ws.cell(row=row, column=1, value=label)
        cell.font = caption_font
        row += 1

        headers = tbl.get("headers", [])
        data_rows = tbl.get("rows", [])
        max_cols = max(
            len(headers),
            max((len(r) for r in data_rows if isinstance(r, list)), default=0),
            1,
        )

        # Header row
        if headers:
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=ci, value=str(h))
                cell.font = th_font
                cell.fill = th_fill
                cell.border = THIN_BORDER
            row += 1

        # Data rows
        for dr in data_rows:
            if isinstance(dr, list):
                for ci, cv in enumerate(dr, 1):
                    cell = ws.cell(row=row, column=ci, value=str(cv) if cv else "")
                    cell.font = td_font
                    cell.border = THIN_BORDER
                row += 1

        row += 1  # blank row between tables

    # Auto-width (approximate)
    for col in range(1, 20):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _collect_all_tables(document: Dict) -> List[Dict]:
    """Collect all table blocks from the document, with source info."""
    tables = []
    sections = document.get("sections", [])
    for sec in sections:
        section_name = sec.get("section_name", sec.get("heading", ""))
        content = sec.get("content", [])
        _collect_tables_from_content(content, section_name, tables)
    return tables


def _collect_tables_from_content(
    content: List, source: str, acc: List[Dict],
) -> None:
    if not isinstance(content, list):
        return
    for block in content:
        if not isinstance(block, dict):
            continue
        if block.get("type") == "table":
            tbl = dict(block)
            tbl["_source_section"] = source
            acc.append(tbl)
        elif block.get("type") == "subsection":
            sub_source = block.get("heading", source)
            _collect_tables_from_content(
                block.get("content", []), sub_source, acc,
            )


# ======================================================================
# CLI
# ======================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Convert extracted section JSON to Excel"
    )
    # parser.add_argument(
    #     "--input", "-i", required=True,
    #     help="Path to the final output JSON from the pipeline",
    # )
    parser.add_argument(
        "--output", "-o",
        help="Output Excel path. Default: same name with .xlsx extension",
    )
    args = parser.parse_args()
    args.input = "output/9100000695 Monadelphous Extension (2 years) + Rate Review Amending Deed Mono Signed & Executed/final/9100000695_Monadelphous_Extension_2_years_Rate_Review_Amending_Deed_Mono_Signed.json"
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: File not found: {input_path}")
        sys.exit(1)

    with open(input_path, encoding="utf-8") as f:
        document = json.load(f)

    if not isinstance(document, dict):
        print(f"ERROR: Expected JSON object, got {type(document).__name__}")
        sys.exit(1)

    output_path = Path(args.output) if args.output else input_path.with_suffix(".xlsx")

    result = write_excel(document, output_path)
    doc_id = document.get("document_id", "?")
    sections = document.get("sections", [])
    rows = flatten_document(document)

    print(f"Document:   {doc_id}")
    print(f"Sections:   {len(sections)}")
    print(f"Rows:       {len(rows)}")
    print(f"Output:     {result}")


if __name__ == "__main__":
    main()
